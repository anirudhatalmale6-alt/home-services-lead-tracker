from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, PieChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import random

wb = Workbook()

# Styles
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_row_fill = PatternFill("solid", fgColor="D6EAF8")
success_fill = PatternFill("solid", fgColor="D4EDDA")
warning_fill = PatternFill("solid", fgColor="FFF3CD")
danger_fill = PatternFill("solid", fgColor="F8D7DA")
profit_fill = PatternFill("solid", fgColor="C6EFCE")
loss_fill = PatternFill("solid", fgColor="FFC7CE")
currency_format = '₹#,##0.00'
date_format = 'DD-MMM-YYYY'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
thick_border = Border(
    left=Side(style='medium'), right=Side(style='medium'),
    top=Side(style='medium'), bottom=Side(style='medium')
)
center_align = Alignment(horizontal='center', vertical='center')

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

def style_data_rows(ws, start_row, end_row, cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = alt_row_fill

def auto_width(ws, cols):
    for col in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

# Common data
services = ["Deep Cleaning", "Regular Cleaning", "Pest Control", "Painting", "Plumbing", "Electrical", "Carpentry", "AC Service"]
project_types = ["Individual", "Apartment Bulk", "Commercial"]
payment_modes = ["Cash", "UPI", "Bank Transfer", "Card"]
payment_status = ["Received", "Pending"]
expense_cats = ["Staff Salary", "Vendor Payment", "Travel & Fuel", "Marketing", "Office Expenses", "Miscellaneous"]
machine_cats = ["Cleaning Machine", "Spray Equipment", "Power Tools", "Safety Equipment"]
base_date = datetime(2025, 1, 1)

# ============ HOW TO USE ============
ws_how = wb.active
ws_how.title = "HOW TO USE"
instructions = [
    ["", "MIS SYSTEM - USER GUIDE", ""],
    [""],
    ["SHEET", "PURPOSE", "HOW TO USE"],
    ["Income Tracker", "Record all service revenue with GST", "Enter each payment with service type"],
    ["Expense Tracker", "Track all business expenses with GST", "Log every expense with category"],
    ["Employee Salaries", "Track employee salary payments", "Monthly salary with PF/ESI deductions"],
    ["Contractor Payments", "Track contractor payouts with TDS", "Gross amount, TDS deduction, net payable"],
    ["Stock Purchases", "Track inventory purchases", "Vendor purchases with GST"],
    ["Machines & Equipment", "Asset register", "Add machines with unique ID"],
    ["Machine Maintenance", "Track repairs/servicing", "Log maintenance with GST"],
    ["Chemicals Stock", "Chemical inventory", "Track opening, added, used, closing stock"],
    ["Accessories Stock", "Consumables inventory", "Track stock levels"],
    ["Stock Transactions", "Stock movement log", "Record purchases and usage"],
    ["Income Summary", "Revenue reports", "Auto-calculated from Income Tracker"],
    ["Expense Summary", "Expense reports", "Auto-calculated from Expense Tracker"],
    ["GST Summary", "GST collected vs paid", "For GST filing reference"],
    ["TDS Summary", "TDS deducted from contractors", "For TDS filing reference"],
    ["Profitability Report", "Profit analysis", "Shows actual NET PROFIT"],
    ["Dashboard", "Visual overview", "Key metrics and charts"],
]
for row_data in instructions:
    ws_how.append(row_data)
ws_how.merge_cells('A1:C1')
ws_how['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws_how['A1'].alignment = center_align
style_header(ws_how, 3, 3)
for row in range(4, 20):
    for col in range(1, 4):
        ws_how.cell(row=row, column=col).border = thin_border
ws_how.column_dimensions['A'].width = 25
ws_how.column_dimensions['B'].width = 35
ws_how.column_dimensions['C'].width = 50

# ============ INCOME TRACKER ============
ws_income = wb.create_sheet("Income Tracker")
income_headers = ["Date", "Invoice No", "Customer Name", "Service Type", "Project Type", "Project Name",
                  "Base Amount (₹)", "GST %", "GST Amount (₹)", "Total Amount (₹)", "Payment Mode", "Payment Status", "Notes"]
ws_income.append(income_headers)
style_header(ws_income, 1, len(income_headers))

customers = ["Rajesh Kumar", "Priya Sharma", "Amit Patel", "Sunita Reddy", "Vikram Singh", "Deepa Nair", "Karthik Iyer", "Meena Gupta", "Rahul Verma", "Anjali Menon", "Suresh Rao", "Lakshmi Pillai", "Arun Krishnan", "Kavitha Srinivasan", "Manoj Das"]

for i in range(15):
    row_num = i + 2
    date = base_date + timedelta(days=random.randint(0, 28))
    invoice = f"INV-2025-{1001+i}"
    customer = customers[i]
    service = random.choice(services)
    proj_type = random.choice(project_types)
    proj_name = "Green Valley Apartments" if proj_type == "Apartment Bulk" else ("Tech Park" if proj_type == "Commercial" else "")
    base_amount = random.randint(2000, 25000) if proj_type == "Individual" else random.randint(50000, 200000)
    gst_rate = 0.18
    mode = random.choice(payment_modes)
    status = random.choice(payment_status)

    ws_income.cell(row=row_num, column=1).value = date
    ws_income.cell(row=row_num, column=2).value = invoice
    ws_income.cell(row=row_num, column=3).value = customer
    ws_income.cell(row=row_num, column=4).value = service
    ws_income.cell(row=row_num, column=5).value = proj_type
    ws_income.cell(row=row_num, column=6).value = proj_name
    ws_income.cell(row=row_num, column=7).value = base_amount
    ws_income.cell(row=row_num, column=8).value = gst_rate
    ws_income.cell(row=row_num, column=9).value = f"=G{row_num}*H{row_num}"
    ws_income.cell(row=row_num, column=10).value = f"=G{row_num}+I{row_num}"
    ws_income.cell(row=row_num, column=11).value = mode
    ws_income.cell(row=row_num, column=12).value = status

style_data_rows(ws_income, 2, 16, len(income_headers))
ws_income.column_dimensions['A'].width = 12
ws_income.column_dimensions['B'].width = 15
ws_income.column_dimensions['C'].width = 18
ws_income.column_dimensions['D'].width = 16
ws_income.column_dimensions['E'].width = 14
ws_income.column_dimensions['F'].width = 22
ws_income.column_dimensions['G'].width = 14
ws_income.column_dimensions['H'].width = 8
ws_income.column_dimensions['I'].width = 14
ws_income.column_dimensions['J'].width = 14

for row in range(2, 17):
    ws_income.cell(row=row, column=1).number_format = date_format
    ws_income.cell(row=row, column=7).number_format = currency_format
    ws_income.cell(row=row, column=8).number_format = '0%'
    ws_income.cell(row=row, column=9).number_format = currency_format
    ws_income.cell(row=row, column=10).number_format = currency_format
ws_income.freeze_panes = 'A2'

dv_service = DataValidation(type="list", formula1='"Deep Cleaning,Regular Cleaning,Pest Control,Painting,Plumbing,Electrical,Carpentry,AC Service"', allow_blank=True)
dv_projtype = DataValidation(type="list", formula1='"Individual,Apartment Bulk,Commercial"', allow_blank=True)
dv_paymode = DataValidation(type="list", formula1='"Cash,UPI,Bank Transfer,Card"', allow_blank=True)
dv_paystatus = DataValidation(type="list", formula1='"Received,Pending"', allow_blank=True)
dv_gst = DataValidation(type="list", formula1='"5%,12%,18%"', allow_blank=True)
ws_income.add_data_validation(dv_service)
ws_income.add_data_validation(dv_projtype)
ws_income.add_data_validation(dv_paymode)
ws_income.add_data_validation(dv_paystatus)
ws_income.add_data_validation(dv_gst)
dv_service.add('D2:D1000')
dv_projtype.add('E2:E1000')
dv_paymode.add('K2:K1000')
dv_paystatus.add('L2:L1000')
dv_gst.add('H2:H1000')

# ============ EXPENSE TRACKER ============
ws_expense = wb.create_sheet("Expense Tracker")
expense_headers = ["Date", "Expense ID", "Category", "Description", "Vendor/Payee",
                   "Base Amount (₹)", "GST %", "GST Amount (₹)", "Total Amount (₹)", "Payment Mode", "Payment Status", "Notes"]
ws_expense.append(expense_headers)
style_header(ws_expense, 1, len(expense_headers))

vendors = ["Kumar Chemicals", "City Fuel Station", "Facebook Ads", "Airtel", "Office Depot", "Google Ads"]

for i in range(15):
    row_num = i + 2
    date = base_date + timedelta(days=random.randint(0, 28))
    exp_id = f"EXP-2025-{101+i}"
    cat = random.choice(["Vendor Payment", "Travel & Fuel", "Marketing", "Office Expenses", "Miscellaneous"])

    if cat == "Travel & Fuel":
        desc = "Petrol/Diesel"
        vendor = "City Fuel Station"
        base_amount = random.randint(500, 3000)
    elif cat == "Marketing":
        desc = random.choice(["Facebook Ads", "Google Ads", "Pamphlets"])
        vendor = desc
        base_amount = random.randint(2000, 15000)
    else:
        desc = random.choice(["Office Supplies", "Utility Bill", "Misc Purchase"])
        vendor = random.choice(vendors)
        base_amount = random.randint(500, 5000)

    gst_rate = 0.18
    mode = random.choice(payment_modes)
    status = random.choice(payment_status)

    ws_expense.cell(row=row_num, column=1).value = date
    ws_expense.cell(row=row_num, column=2).value = exp_id
    ws_expense.cell(row=row_num, column=3).value = cat
    ws_expense.cell(row=row_num, column=4).value = desc
    ws_expense.cell(row=row_num, column=5).value = vendor
    ws_expense.cell(row=row_num, column=6).value = base_amount
    ws_expense.cell(row=row_num, column=7).value = gst_rate
    ws_expense.cell(row=row_num, column=8).value = f"=F{row_num}*G{row_num}"
    ws_expense.cell(row=row_num, column=9).value = f"=F{row_num}+H{row_num}"
    ws_expense.cell(row=row_num, column=10).value = mode
    ws_expense.cell(row=row_num, column=11).value = status

style_data_rows(ws_expense, 2, 16, len(expense_headers))
auto_width(ws_expense, len(expense_headers))
ws_expense.column_dimensions['D'].width = 20
ws_expense.column_dimensions['G'].width = 8

for row in range(2, 17):
    ws_expense.cell(row=row, column=1).number_format = date_format
    ws_expense.cell(row=row, column=6).number_format = currency_format
    ws_expense.cell(row=row, column=7).number_format = '0%'
    ws_expense.cell(row=row, column=8).number_format = currency_format
    ws_expense.cell(row=row, column=9).number_format = currency_format
ws_expense.freeze_panes = 'A2'

dv_expcat = DataValidation(type="list", formula1='"Vendor Payment,Travel & Fuel,Marketing,Office Expenses,Miscellaneous"', allow_blank=True)
ws_expense.add_data_validation(dv_expcat)
ws_expense.add_data_validation(dv_paymode)
ws_expense.add_data_validation(dv_paystatus)
ws_expense.add_data_validation(dv_gst)
dv_expcat.add('C2:C1000')
dv_gst.add('G2:G1000')

# ============ EMPLOYEE SALARIES ============
ws_salary = wb.create_sheet("Employee Salaries")
salary_headers = ["Month", "Employee ID", "Employee Name", "Designation", "Department",
                  "Basic Salary (₹)", "HRA (₹)", "Other Allowances (₹)", "Gross Salary (₹)",
                  "PF Deduction (₹)", "ESI Deduction (₹)", "Other Deductions (₹)", "Total Deductions (₹)",
                  "Net Salary (₹)", "Payment Date", "Payment Mode", "Payment Status", "Notes"]
ws_salary.append(salary_headers)
style_header(ws_salary, 1, len(salary_headers))

employees = [
    ("EMP001", "Ramesh Kumar", "Driver", "Operations", 18000, 3600, 1500),
    ("EMP002", "Suresh Nair", "Technician", "Operations", 22000, 4400, 2000),
    ("EMP003", "Anil Sharma", "Helper", "Operations", 15000, 3000, 1000),
    ("EMP004", "Priya Menon", "Supervisor", "Operations", 28000, 5600, 2500),
    ("EMP005", "Karthik Reddy", "Technician", "Operations", 22000, 4400, 2000),
    ("EMP006", "Lakshmi Iyer", "Admin", "Office", 20000, 4000, 1500),
    ("EMP007", "Vijay Das", "Helper", "Operations", 15000, 3000, 1000),
    ("EMP008", "Meena Pillai", "Accounts", "Office", 25000, 5000, 2000),
]

for i, emp in enumerate(employees, start=2):
    ws_salary.cell(row=i, column=1).value = "January 2025"
    ws_salary.cell(row=i, column=2).value = emp[0]
    ws_salary.cell(row=i, column=3).value = emp[1]
    ws_salary.cell(row=i, column=4).value = emp[2]
    ws_salary.cell(row=i, column=5).value = emp[3]
    ws_salary.cell(row=i, column=6).value = emp[4]
    ws_salary.cell(row=i, column=7).value = emp[5]
    ws_salary.cell(row=i, column=8).value = emp[6]
    ws_salary.cell(row=i, column=9).value = f"=F{i}+G{i}+H{i}"
    ws_salary.cell(row=i, column=10).value = f"=ROUND(F{i}*0.12,0)"  # 12% PF
    ws_salary.cell(row=i, column=11).value = f"=IF(I{i}<=21000,ROUND(I{i}*0.0075,0),0)"  # 0.75% ESI if gross <= 21000
    ws_salary.cell(row=i, column=12).value = 0
    ws_salary.cell(row=i, column=13).value = f"=J{i}+K{i}+L{i}"
    ws_salary.cell(row=i, column=14).value = f"=I{i}-M{i}"
    ws_salary.cell(row=i, column=15).value = datetime(2025, 1, 31)
    ws_salary.cell(row=i, column=16).value = "Bank Transfer"
    ws_salary.cell(row=i, column=17).value = "Paid"

style_data_rows(ws_salary, 2, 9, len(salary_headers))
auto_width(ws_salary, len(salary_headers))
ws_salary.column_dimensions['C'].width = 18
ws_salary.column_dimensions['D'].width = 14

for row in range(2, 10):
    for col in [6, 7, 8, 9, 10, 11, 12, 13, 14]:
        ws_salary.cell(row=row, column=col).number_format = currency_format
    ws_salary.cell(row=row, column=15).number_format = date_format
ws_salary.freeze_panes = 'A2'

dv_month = DataValidation(type="list", formula1='"January 2025,February 2025,March 2025,April 2025,May 2025,June 2025,July 2025,August 2025,September 2025,October 2025,November 2025,December 2025"', allow_blank=True)
dv_salstatus = DataValidation(type="list", formula1='"Paid,Pending,On Hold"', allow_blank=True)
ws_salary.add_data_validation(dv_month)
ws_salary.add_data_validation(dv_salstatus)
ws_salary.add_data_validation(dv_paymode)
dv_month.add('A2:A1000')
dv_salstatus.add('Q2:Q1000')

# ============ CONTRACTOR PAYMENTS ============
ws_contractor = wb.create_sheet("Contractor Payments")
contractor_headers = ["Date", "Payment ID", "Contractor Name", "Contractor PAN", "Service Type",
                      "Project/Work Description", "Gross Amount (₹)", "TDS % (Sec 194C)", "TDS Amount (₹)",
                      "Net Payable (₹)", "Payment Date", "Payment Mode", "Payment Status", "TDS Deposited", "Notes"]
ws_contractor.append(contractor_headers)
style_header(ws_contractor, 1, len(contractor_headers))

contractors = [
    ("Sharma Cleaning Services", "ABCPS1234A", "Deep Cleaning", "Green Valley Phase 1", 45000, 0.02),
    ("Krishna Pest Control", "DEFPK5678B", "Pest Control", "Tech Park Building A", 35000, 0.02),
    ("Apex Painting Works", "GHIPA9012C", "Painting", "Residential Complex", 85000, 0.02),
    ("City Plumbers", "JKLPC3456D", "Plumbing", "Apartment Maintenance", 28000, 0.01),
    ("Royal AC Services", "MNOPR7890E", "AC Service", "Office Complex", 42000, 0.02),
    ("Metro Electrical", "PQRME1234F", "Electrical", "New Installation", 55000, 0.02),
    ("Clean Pro Services", "STUCL5678G", "Regular Cleaning", "Monthly Contract", 65000, 0.02),
    ("Pest Masters", "VWXPM9012H", "Pest Control", "Quarterly Treatment", 48000, 0.02),
]

for i, cont in enumerate(contractors, start=2):
    ws_contractor.cell(row=i, column=1).value = base_date + timedelta(days=random.randint(0, 28))
    ws_contractor.cell(row=i, column=2).value = f"CONT-2025-{301+i-2}"
    ws_contractor.cell(row=i, column=3).value = cont[0]
    ws_contractor.cell(row=i, column=4).value = cont[1]
    ws_contractor.cell(row=i, column=5).value = cont[2]
    ws_contractor.cell(row=i, column=6).value = cont[3]
    ws_contractor.cell(row=i, column=7).value = cont[4]
    ws_contractor.cell(row=i, column=8).value = cont[5]
    ws_contractor.cell(row=i, column=9).value = f"=G{i}*H{i}"
    ws_contractor.cell(row=i, column=10).value = f"=G{i}-I{i}"
    ws_contractor.cell(row=i, column=11).value = base_date + timedelta(days=random.randint(5, 30))
    ws_contractor.cell(row=i, column=12).value = random.choice(["Bank Transfer", "UPI", "Cheque"])
    ws_contractor.cell(row=i, column=13).value = random.choice(["Paid", "Pending"])
    ws_contractor.cell(row=i, column=14).value = random.choice(["Yes", "No"])

style_data_rows(ws_contractor, 2, 9, len(contractor_headers))
auto_width(ws_contractor, len(contractor_headers))
ws_contractor.column_dimensions['C'].width = 22
ws_contractor.column_dimensions['F'].width = 25
ws_contractor.column_dimensions['H'].width = 16

for row in range(2, 10):
    ws_contractor.cell(row=row, column=1).number_format = date_format
    ws_contractor.cell(row=row, column=7).number_format = currency_format
    ws_contractor.cell(row=row, column=8).number_format = '0.0%'
    ws_contractor.cell(row=row, column=9).number_format = currency_format
    ws_contractor.cell(row=row, column=10).number_format = currency_format
    ws_contractor.cell(row=row, column=11).number_format = date_format
ws_contractor.freeze_panes = 'A2'

dv_tds = DataValidation(type="list", formula1='"1%,2%,10%"', allow_blank=True)
dv_contstatus = DataValidation(type="list", formula1='"Paid,Pending,On Hold"', allow_blank=True)
dv_tdsdeposited = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws_contractor.add_data_validation(dv_tds)
ws_contractor.add_data_validation(dv_contstatus)
ws_contractor.add_data_validation(dv_tdsdeposited)
ws_contractor.add_data_validation(dv_service)
ws_contractor.add_data_validation(dv_paymode)
dv_tds.add('H2:H1000')
dv_contstatus.add('M2:M1000')
dv_tdsdeposited.add('N2:N1000')
dv_service.add('E2:E1000')

# ============ STOCK PURCHASES ============
ws_stock_purch = wb.create_sheet("Stock Purchases")
stock_purch_headers = ["Date", "Purchase ID", "Vendor Name", "Vendor GSTIN", "Invoice No", "Item Type",
                       "Item Name", "Quantity", "Unit", "Rate (₹)", "Base Amount (₹)", "GST %", "GST Amount (₹)",
                       "Total Amount (₹)", "Payment Status", "Notes"]
ws_stock_purch.append(stock_purch_headers)
style_header(ws_stock_purch, 1, len(stock_purch_headers))

stock_vendors = [
    ("Kumar Chemicals Pvt Ltd", "29AABCK1234A1ZV"),
    ("City Supplies", "29AABCS5678B2ZW"),
    ("Metro Cleaning Supplies", "29AABCM9012C3ZX"),
]

stock_items = [
    ("Chemical", "Floor Cleaner", "Liters", 120),
    ("Chemical", "Glass Cleaner", "Liters", 95),
    ("Accessory", "Microfiber Cloth", "Pieces", 80),
    ("Accessory", "Rubber Gloves", "Pairs", 45),
    ("Chemical", "Surface Disinfectant", "Liters", 180),
]

for i in range(10):
    row_num = i + 2
    date = base_date + timedelta(days=random.randint(0, 28))
    purch_id = f"PO-2025-{201+i}"
    vendor, gstin = random.choice(stock_vendors)
    inv_no = f"INV/{random.randint(1000, 9999)}"
    item_type, item_name, unit, rate = random.choice(stock_items)
    qty = random.randint(10, 100)
    gst_rate = 0.18

    ws_stock_purch.cell(row=row_num, column=1).value = date
    ws_stock_purch.cell(row=row_num, column=2).value = purch_id
    ws_stock_purch.cell(row=row_num, column=3).value = vendor
    ws_stock_purch.cell(row=row_num, column=4).value = gstin
    ws_stock_purch.cell(row=row_num, column=5).value = inv_no
    ws_stock_purch.cell(row=row_num, column=6).value = item_type
    ws_stock_purch.cell(row=row_num, column=7).value = item_name
    ws_stock_purch.cell(row=row_num, column=8).value = qty
    ws_stock_purch.cell(row=row_num, column=9).value = unit
    ws_stock_purch.cell(row=row_num, column=10).value = rate
    ws_stock_purch.cell(row=row_num, column=11).value = f"=H{row_num}*J{row_num}"
    ws_stock_purch.cell(row=row_num, column=12).value = gst_rate
    ws_stock_purch.cell(row=row_num, column=13).value = f"=K{row_num}*L{row_num}"
    ws_stock_purch.cell(row=row_num, column=14).value = f"=K{row_num}+M{row_num}"
    ws_stock_purch.cell(row=row_num, column=15).value = random.choice(["Paid", "Pending"])

style_data_rows(ws_stock_purch, 2, 11, len(stock_purch_headers))
auto_width(ws_stock_purch, len(stock_purch_headers))
ws_stock_purch.column_dimensions['C'].width = 22
ws_stock_purch.column_dimensions['G'].width = 20

for row in range(2, 12):
    ws_stock_purch.cell(row=row, column=1).number_format = date_format
    ws_stock_purch.cell(row=row, column=10).number_format = currency_format
    ws_stock_purch.cell(row=row, column=11).number_format = currency_format
    ws_stock_purch.cell(row=row, column=12).number_format = '0%'
    ws_stock_purch.cell(row=row, column=13).number_format = currency_format
    ws_stock_purch.cell(row=row, column=14).number_format = currency_format
ws_stock_purch.freeze_panes = 'A2'

dv_stockitem = DataValidation(type="list", formula1='"Chemical,Accessory,Machine,Spare Part"', allow_blank=True)
dv_purch_status = DataValidation(type="list", formula1='"Paid,Pending"', allow_blank=True)
ws_stock_purch.add_data_validation(dv_stockitem)
ws_stock_purch.add_data_validation(dv_purch_status)
ws_stock_purch.add_data_validation(dv_gst)
dv_stockitem.add('F2:F1000')
dv_purch_status.add('O2:O1000')
dv_gst.add('L2:L1000')

# ============ MACHINES & EQUIPMENT ============
ws_machines = wb.create_sheet("Machines & Equipment")
machine_headers = ["Machine ID", "Machine Name", "Model", "Category", "Purchase Date", "Base Cost (₹)",
                   "GST %", "GST Amount (₹)", "Total Cost (₹)", "Current Status", "Location", "Notes"]
ws_machines.append(machine_headers)
style_header(ws_machines, 1, len(machine_headers))

machine_data = [
    ["M001", "Floor Scrubber", "Karcher BD 50/50", "Cleaning Machine", datetime(2023, 3, 15), 72034, 0.18, "Active", "Team A"],
    ["M002", "Pressure Washer", "Bosch AQT 45-14X", "Cleaning Machine", datetime(2023, 5, 20), 38136, 0.18, "Active", "Team B"],
    ["M003", "Vacuum Cleaner", "Eureka Forbes Pro", "Cleaning Machine", datetime(2022, 8, 10), 23729, 0.18, "Active", "Team A"],
    ["M004", "Pest Control Sprayer", "Solo 475", "Spray Equipment", datetime(2023, 1, 5), 12712, 0.18, "Active", "Pest Team"],
    ["M005", "Paint Sprayer", "Graco Magnum X5", "Spray Equipment", datetime(2023, 7, 12), 55085, 0.18, "Under Repair", "Painting"],
    ["M006", "Drill Machine", "Bosch GSB 600", "Power Tools", datetime(2022, 4, 18), 7203, 0.18, "Active", "General"],
    ["M007", "Angle Grinder", "Makita GA4030", "Power Tools", datetime(2023, 2, 22), 5508, 0.18, "Active", "General"],
    ["M008", "Steam Cleaner", "Karcher SC3", "Cleaning Machine", datetime(2024, 1, 8), 29661, 0.18, "Active", "Team C"],
]

for i, row in enumerate(machine_data, start=2):
    ws_machines.cell(row=i, column=1).value = row[0]
    ws_machines.cell(row=i, column=2).value = row[1]
    ws_machines.cell(row=i, column=3).value = row[2]
    ws_machines.cell(row=i, column=4).value = row[3]
    ws_machines.cell(row=i, column=5).value = row[4]
    ws_machines.cell(row=i, column=6).value = row[5]
    ws_machines.cell(row=i, column=7).value = row[6]
    ws_machines.cell(row=i, column=8).value = f"=F{i}*G{i}"
    ws_machines.cell(row=i, column=9).value = f"=F{i}+H{i}"
    ws_machines.cell(row=i, column=10).value = row[7]
    ws_machines.cell(row=i, column=11).value = row[8]

style_data_rows(ws_machines, 2, 9, len(machine_headers))
auto_width(ws_machines, len(machine_headers))
ws_machines.column_dimensions['B'].width = 18
ws_machines.column_dimensions['C'].width = 18

for row in range(2, 10):
    ws_machines.cell(row=row, column=5).number_format = date_format
    ws_machines.cell(row=row, column=6).number_format = currency_format
    ws_machines.cell(row=row, column=7).number_format = '0%'
    ws_machines.cell(row=row, column=8).number_format = currency_format
    ws_machines.cell(row=row, column=9).number_format = currency_format
ws_machines.freeze_panes = 'A2'

dv_machcat = DataValidation(type="list", formula1='"Cleaning Machine,Spray Equipment,Power Tools,Safety Equipment"', allow_blank=True)
dv_machstatus = DataValidation(type="list", formula1='"Active,Under Repair,Retired"', allow_blank=True)
ws_machines.add_data_validation(dv_machcat)
ws_machines.add_data_validation(dv_machstatus)
ws_machines.add_data_validation(dv_gst)
dv_machcat.add('D2:D1000')
dv_machstatus.add('J2:J1000')
dv_gst.add('G2:G1000')

# ============ MACHINE MAINTENANCE ============
ws_maint = wb.create_sheet("Machine Maintenance")
maint_headers = ["Date", "Machine ID", "Machine Name", "Type", "Description", "Base Cost (₹)",
                 "GST %", "GST Amount (₹)", "Total Cost (₹)", "Done By", "Next Service", "Notes"]
ws_maint.append(maint_headers)
style_header(ws_maint, 1, len(maint_headers))

maint_data = [
    [datetime(2024, 8, 1), "M001", "Floor Scrubber", "Service", "Brush replacement", 3814, 0.18, "Karcher Service", datetime(2025, 2, 1)],
    [datetime(2024, 9, 5), "M002", "Pressure Washer", "Repair", "Nozzle fixed", 1525, 0.18, "Bosch Service", datetime(2025, 3, 5)],
    [datetime(2024, 10, 10), "M003", "Vacuum Cleaner", "Service", "Filter cleaning", 678, 0.18, "In-house", datetime(2025, 4, 10)],
    [datetime(2024, 11, 15), "M005", "Paint Sprayer", "Repair", "Motor rewinding", 7203, 0.18, "Graco Service", datetime(2025, 5, 15)],
    [datetime(2025, 1, 5), "M004", "Pest Sprayer", "Service", "Valve check", 424, 0.18, "In-house", datetime(2025, 7, 5)],
]

for i, row in enumerate(maint_data, start=2):
    ws_maint.cell(row=i, column=1).value = row[0]
    ws_maint.cell(row=i, column=2).value = row[1]
    ws_maint.cell(row=i, column=3).value = row[2]
    ws_maint.cell(row=i, column=4).value = row[3]
    ws_maint.cell(row=i, column=5).value = row[4]
    ws_maint.cell(row=i, column=6).value = row[5]
    ws_maint.cell(row=i, column=7).value = row[6]
    ws_maint.cell(row=i, column=8).value = f"=F{i}*G{i}"
    ws_maint.cell(row=i, column=9).value = f"=F{i}+H{i}"
    ws_maint.cell(row=i, column=10).value = row[7]
    ws_maint.cell(row=i, column=11).value = row[8]

style_data_rows(ws_maint, 2, 6, len(maint_headers))
auto_width(ws_maint, len(maint_headers))

for row in range(2, 7):
    ws_maint.cell(row=row, column=1).number_format = date_format
    ws_maint.cell(row=row, column=6).number_format = currency_format
    ws_maint.cell(row=row, column=7).number_format = '0%'
    ws_maint.cell(row=row, column=8).number_format = currency_format
    ws_maint.cell(row=row, column=9).number_format = currency_format
    ws_maint.cell(row=row, column=11).number_format = date_format
ws_maint.freeze_panes = 'A2'

dv_mainttype = DataValidation(type="list", formula1='"Repair,Service,Parts Replacement"', allow_blank=True)
dv_machid = DataValidation(type="list", formula1='"M001,M002,M003,M004,M005,M006,M007,M008"', allow_blank=True)
ws_maint.add_data_validation(dv_mainttype)
ws_maint.add_data_validation(dv_machid)
ws_maint.add_data_validation(dv_gst)
dv_mainttype.add('D2:D1000')
dv_machid.add('B2:B1000')
dv_gst.add('G2:G1000')

# ============ CHEMICALS STOCK ============
ws_chem = wb.create_sheet("Chemicals Stock")
chem_headers = ["Item ID", "Chemical Name", "Brand", "Category", "Unit", "Opening Stock", "Stock Added", "Stock Used", "Closing Stock", "Min Level", "Status", "Last Updated"]
ws_chem.append(chem_headers)
style_header(ws_chem, 1, len(chem_headers))

chem_data = [
    ["CH001", "Floor Cleaner", "Lizol", "Cleaning Agents", "Liters", 50, 100, 80, None, 20, None, datetime(2025, 1, 25)],
    ["CH002", "Glass Cleaner", "Colin", "Cleaning Agents", "Liters", 30, 50, 45, None, 15, None, datetime(2025, 1, 25)],
    ["CH003", "Toilet Cleaner", "Harpic", "Cleaning Agents", "Liters", 40, 80, 70, None, 20, None, datetime(2025, 1, 25)],
    ["CH004", "Pesticide Spray", "Baygon Pro", "Pest Control", "Cans", 25, 50, 40, None, 10, None, datetime(2025, 1, 25)],
    ["CH005", "Surface Disinfectant", "Lysol", "Sanitizers", "Liters", 25, 50, 40, None, 15, None, datetime(2025, 1, 25)],
]

for i, row in enumerate(chem_data, start=2):
    ws_chem.append(row[:8])
    ws_chem.cell(row=i, column=9).value = f"=F{i}+G{i}-H{i}"
    ws_chem.cell(row=i, column=10).value = row[9]
    ws_chem.cell(row=i, column=11).value = f'=IF(I{i}<J{i},"Low Stock","OK")'
    ws_chem.cell(row=i, column=12).value = row[11]

style_data_rows(ws_chem, 2, 6, len(chem_headers))
auto_width(ws_chem, len(chem_headers))
ws_chem.column_dimensions['B'].width = 20
for row in range(2, 7):
    ws_chem.cell(row=row, column=12).number_format = date_format
ws_chem.freeze_panes = 'A2'

low_stock_fill = PatternFill("solid", fgColor="FFCCCC")
ok_fill = PatternFill("solid", fgColor="CCFFCC")
ws_chem.conditional_formatting.add('K2:K1000', FormulaRule(formula=['K2="Low Stock"'], fill=low_stock_fill))
ws_chem.conditional_formatting.add('K2:K1000', FormulaRule(formula=['K2="OK"'], fill=ok_fill))

# ============ ACCESSORIES STOCK ============
ws_acc = wb.create_sheet("Accessories Stock")
acc_headers = ["Item ID", "Item Name", "Category", "Unit", "Opening Stock", "Stock Added", "Stock Used", "Closing Stock", "Min Level", "Status", "Last Updated"]
ws_acc.append(acc_headers)
style_header(ws_acc, 1, len(acc_headers))

acc_data = [
    ["AC001", "Rubber Hose (10m)", "Pipes & Hoses", "Pieces", 10, 15, 12, None, 5, None, datetime(2025, 1, 25)],
    ["AC002", "Floor Mop", "Brushes & Mops", "Pieces", 20, 30, 25, None, 10, None, datetime(2025, 1, 25)],
    ["AC003", "Microfiber Cloth", "Cloths & Wipes", "Pieces", 50, 100, 80, None, 25, None, datetime(2025, 1, 25)],
    ["AC004", "Rubber Gloves", "Gloves & Safety", "Pairs", 30, 50, 40, None, 15, None, datetime(2025, 1, 25)],
    ["AC005", "Spray Bottles", "Other", "Pieces", 20, 30, 22, None, 10, None, datetime(2025, 1, 25)],
]

for i, row in enumerate(acc_data, start=2):
    ws_acc.append(row[:7])
    ws_acc.cell(row=i, column=8).value = f"=E{i}+F{i}-G{i}"
    ws_acc.cell(row=i, column=9).value = row[8]
    ws_acc.cell(row=i, column=10).value = f'=IF(H{i}<I{i},"Low Stock","OK")'
    ws_acc.cell(row=i, column=11).value = row[10]

style_data_rows(ws_acc, 2, 6, len(acc_headers))
auto_width(ws_acc, len(acc_headers))
ws_acc.column_dimensions['B'].width = 18
for row in range(2, 7):
    ws_acc.cell(row=row, column=11).number_format = date_format
ws_acc.freeze_panes = 'A2'
ws_acc.conditional_formatting.add('J2:J1000', FormulaRule(formula=['J2="Low Stock"'], fill=low_stock_fill))
ws_acc.conditional_formatting.add('J2:J1000', FormulaRule(formula=['J2="OK"'], fill=ok_fill))

# ============ STOCK TRANSACTIONS ============
ws_trans = wb.create_sheet("Stock Transactions")
trans_headers = ["Date", "Transaction ID", "Item Type", "Item ID", "Item Name", "Transaction Type",
                 "Quantity", "Unit Rate (₹)", "GST %", "GST Amount (₹)", "Total Cost (₹)", "Project/Team", "Notes"]
ws_trans.append(trans_headers)
style_header(ws_trans, 1, len(trans_headers))

trans_data = [
    [datetime(2025, 1, 2), "TXN001", "Chemical", "CH001", "Floor Cleaner", "Purchase", 50, 120, 0.18, None, None, "Warehouse"],
    [datetime(2025, 1, 5), "TXN002", "Accessory", "AC003", "Microfiber Cloth", "Purchase", 50, 80, 0.18, None, None, "Warehouse"],
    [datetime(2025, 1, 6), "TXN003", "Chemical", "CH001", "Floor Cleaner", "Used", 20, 0, 0, None, None, "Team A"],
    [datetime(2025, 1, 8), "TXN004", "Accessory", "AC004", "Rubber Gloves", "Used", 10, 0, 0, None, None, "Team B"],
    [datetime(2025, 1, 10), "TXN005", "Chemical", "CH002", "Glass Cleaner", "Purchase", 25, 95, 0.18, None, None, "Warehouse"],
]

for i, row in enumerate(trans_data, start=2):
    ws_trans.cell(row=i, column=1).value = row[0]
    ws_trans.cell(row=i, column=2).value = row[1]
    ws_trans.cell(row=i, column=3).value = row[2]
    ws_trans.cell(row=i, column=4).value = row[3]
    ws_trans.cell(row=i, column=5).value = row[4]
    ws_trans.cell(row=i, column=6).value = row[5]
    ws_trans.cell(row=i, column=7).value = row[6]
    ws_trans.cell(row=i, column=8).value = row[7]
    ws_trans.cell(row=i, column=9).value = row[8]
    ws_trans.cell(row=i, column=10).value = f"=G{i}*H{i}*I{i}"
    ws_trans.cell(row=i, column=11).value = f"=G{i}*H{i}+J{i}"
    ws_trans.cell(row=i, column=12).value = row[11]

style_data_rows(ws_trans, 2, 6, len(trans_headers))
auto_width(ws_trans, len(trans_headers))
for row in range(2, 7):
    ws_trans.cell(row=row, column=1).number_format = date_format
    ws_trans.cell(row=row, column=8).number_format = currency_format
    ws_trans.cell(row=row, column=9).number_format = '0%'
    ws_trans.cell(row=row, column=10).number_format = currency_format
    ws_trans.cell(row=row, column=11).number_format = currency_format
ws_trans.freeze_panes = 'A2'

dv_transtype = DataValidation(type="list", formula1='"Purchase,Used,Returned,Damaged"', allow_blank=True)
ws_trans.add_data_validation(dv_stockitem)
ws_trans.add_data_validation(dv_transtype)
ws_trans.add_data_validation(dv_gst)
dv_stockitem.add('C2:C1000')
dv_transtype.add('F2:F1000')
dv_gst.add('I2:I1000')

# ============ INCOME SUMMARY ============
ws_inc_sum = wb.create_sheet("Income Summary")
ws_inc_sum['A1'] = "INCOME SUMMARY REPORT"
ws_inc_sum.merge_cells('A1:E1')
ws_inc_sum['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_inc_sum['A3'] = "SERVICE-WISE REVENUE"
ws_inc_sum['A3'].font = Font(bold=True, size=12)
ws_inc_sum.append(["Service Type", "Base Revenue (₹)", "GST Collected (₹)", "Total Revenue (₹)", "Orders"])
style_header(ws_inc_sum, 4, 5)

for i, svc in enumerate(services, start=5):
    ws_inc_sum.cell(row=i, column=1).value = svc
    ws_inc_sum.cell(row=i, column=2).value = f'=SUMIF(\'Income Tracker\'!D:D,A{i},\'Income Tracker\'!G:G)'
    ws_inc_sum.cell(row=i, column=3).value = f'=SUMIF(\'Income Tracker\'!D:D,A{i},\'Income Tracker\'!I:I)'
    ws_inc_sum.cell(row=i, column=4).value = f'=SUMIF(\'Income Tracker\'!D:D,A{i},\'Income Tracker\'!J:J)'
    ws_inc_sum.cell(row=i, column=5).value = f'=COUNTIF(\'Income Tracker\'!D:D,A{i})'
    for col in [2, 3, 4]:
        ws_inc_sum.cell(row=i, column=col).number_format = currency_format

ws_inc_sum.cell(row=13, column=1).value = "TOTAL"
ws_inc_sum.cell(row=13, column=1).font = Font(bold=True)
for col in range(2, 6):
    ws_inc_sum.cell(row=13, column=col).value = f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}12)"
    ws_inc_sum.cell(row=13, column=col).font = Font(bold=True)
    if col < 5:
        ws_inc_sum.cell(row=13, column=col).number_format = currency_format

style_data_rows(ws_inc_sum, 5, 13, 5)
auto_width(ws_inc_sum, 5)
ws_inc_sum.column_dimensions['A'].width = 18

# ============ EXPENSE SUMMARY ============
ws_exp_sum = wb.create_sheet("Expense Summary")
ws_exp_sum['A1'] = "EXPENSE SUMMARY REPORT"
ws_exp_sum.merge_cells('A1:E1')
ws_exp_sum['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_exp_sum['A3'] = "CATEGORY-WISE EXPENSES"
ws_exp_sum['A3'].font = Font(bold=True, size=12)
ws_exp_sum.append(["Category", "Base Amount (₹)", "GST Paid (₹)", "Total Amount (₹)", "Count"])
style_header(ws_exp_sum, 4, 5)

exp_cats_no_salary = ["Vendor Payment", "Travel & Fuel", "Marketing", "Office Expenses", "Miscellaneous"]
for i, cat in enumerate(exp_cats_no_salary, start=5):
    ws_exp_sum.cell(row=i, column=1).value = cat
    ws_exp_sum.cell(row=i, column=2).value = f'=SUMIF(\'Expense Tracker\'!C:C,A{i},\'Expense Tracker\'!F:F)'
    ws_exp_sum.cell(row=i, column=3).value = f'=SUMIF(\'Expense Tracker\'!C:C,A{i},\'Expense Tracker\'!H:H)'
    ws_exp_sum.cell(row=i, column=4).value = f'=SUMIF(\'Expense Tracker\'!C:C,A{i},\'Expense Tracker\'!I:I)'
    ws_exp_sum.cell(row=i, column=5).value = f'=COUNTIF(\'Expense Tracker\'!C:C,A{i})'
    for col in [2, 3, 4]:
        ws_exp_sum.cell(row=i, column=col).number_format = currency_format

ws_exp_sum.cell(row=10, column=1).value = "TOTAL"
ws_exp_sum.cell(row=10, column=1).font = Font(bold=True)
for col in range(2, 6):
    ws_exp_sum.cell(row=10, column=col).value = f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}9)"
    ws_exp_sum.cell(row=10, column=col).font = Font(bold=True)
    if col < 5:
        ws_exp_sum.cell(row=10, column=col).number_format = currency_format

style_data_rows(ws_exp_sum, 5, 10, 5)
auto_width(ws_exp_sum, 5)

# ============ GST SUMMARY ============
ws_gst = wb.create_sheet("GST Summary")
ws_gst['A1'] = "GST SUMMARY REPORT"
ws_gst.merge_cells('A1:C1')
ws_gst['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_gst['A3'] = "GST COLLECTED (OUTPUT TAX)"
ws_gst['A3'].font = Font(bold=True, size=12, color="008000")
ws_gst.append(["Source", "Base Amount (₹)", "GST Amount (₹)"])
style_header(ws_gst, 4, 3)
ws_gst['A5'] = "Service Revenue"
ws_gst['B5'] = "=SUM('Income Tracker'!G:G)"
ws_gst['C5'] = "=SUM('Income Tracker'!I:I)"
ws_gst['A6'] = "TOTAL GST COLLECTED"
ws_gst['A6'].font = Font(bold=True)
ws_gst['B6'] = "=B5"
ws_gst['C6'] = "=C5"
ws_gst['C6'].font = Font(bold=True, color="008000")
for row in [5, 6]:
    ws_gst.cell(row=row, column=2).number_format = currency_format
    ws_gst.cell(row=row, column=3).number_format = currency_format
style_data_rows(ws_gst, 5, 6, 3)

ws_gst['A9'] = "GST PAID (INPUT TAX)"
ws_gst['A9'].font = Font(bold=True, size=12, color="FF0000")
ws_gst.append(["Source", "Base Amount (₹)", "GST Amount (₹)"])
style_header(ws_gst, 10, 3)
gst_sources = [
    ("Expenses", "=SUM('Expense Tracker'!F:F)", "=SUM('Expense Tracker'!H:H)"),
    ("Stock Purchases", "=SUM('Stock Purchases'!K:K)", "=SUM('Stock Purchases'!M:M)"),
    ("Machine Purchases", "=SUM('Machines & Equipment'!F:F)", "=SUM('Machines & Equipment'!H:H)"),
    ("Maintenance", "=SUM('Machine Maintenance'!F:F)", "=SUM('Machine Maintenance'!H:H)"),
]
for i, (name, base, gst) in enumerate(gst_sources, start=11):
    ws_gst.cell(row=i, column=1).value = name
    ws_gst.cell(row=i, column=2).value = base
    ws_gst.cell(row=i, column=3).value = gst
    ws_gst.cell(row=i, column=2).number_format = currency_format
    ws_gst.cell(row=i, column=3).number_format = currency_format
ws_gst['A15'] = "TOTAL GST PAID"
ws_gst['A15'].font = Font(bold=True)
ws_gst['B15'] = "=SUM(B11:B14)"
ws_gst['C15'] = "=SUM(C11:C14)"
ws_gst['C15'].font = Font(bold=True, color="FF0000")
ws_gst['B15'].number_format = currency_format
ws_gst['C15'].number_format = currency_format
style_data_rows(ws_gst, 11, 15, 3)

ws_gst['A18'] = "NET GST PAYABLE"
ws_gst['A18'].font = Font(bold=True, size=14)
ws_gst['C18'] = "=C6-C15"
ws_gst['C18'].font = Font(bold=True, size=14, color="1F4E79")
ws_gst['C18'].number_format = currency_format
ws_gst['C18'].fill = PatternFill("solid", fgColor="FFFFCC")

auto_width(ws_gst, 3)
ws_gst.column_dimensions['A'].width = 20
ws_gst.column_dimensions['B'].width = 18
ws_gst.column_dimensions['C'].width = 18

# ============ TDS SUMMARY ============
ws_tds = wb.create_sheet("TDS Summary")
ws_tds['A1'] = "TDS SUMMARY REPORT"
ws_tds.merge_cells('A1:C1')
ws_tds['A1'].font = Font(bold=True, size=14, color="1F4E79")
ws_tds['A2'] = "(For TDS Filing - Section 194C)"
ws_tds['A2'].font = Font(italic=True, color="666666")

ws_tds['A4'] = "TDS DEDUCTED FROM CONTRACTORS"
ws_tds['A4'].font = Font(bold=True, size=12)
ws_tds.append(["Metric", "Amount (₹)"])
style_header(ws_tds, 5, 2)
ws_tds['A6'] = "Total Contractor Payments (Gross)"
ws_tds['B6'] = "=SUM('Contractor Payments'!G:G)"
ws_tds['A7'] = "Total TDS Deducted"
ws_tds['B7'] = "=SUM('Contractor Payments'!I:I)"
ws_tds['A8'] = "Total Net Paid to Contractors"
ws_tds['B8'] = "=SUM('Contractor Payments'!J:J)"
ws_tds['A9'] = "TDS Deposited"
ws_tds['B9'] = "=SUMIF('Contractor Payments'!N:N,\"Yes\",'Contractor Payments'!I:I)"
ws_tds['A10'] = "TDS Pending Deposit"
ws_tds['B10'] = "=B7-B9"
ws_tds['B10'].font = Font(bold=True, color="FF0000")

for row in range(6, 11):
    ws_tds.cell(row=row, column=2).number_format = currency_format
style_data_rows(ws_tds, 6, 10, 2)
auto_width(ws_tds, 2)
ws_tds.column_dimensions['A'].width = 35

# ============ PROFITABILITY REPORT ============
ws_profit = wb.create_sheet("Profitability Report")
ws_profit['A1'] = "PROFITABILITY ANALYSIS"
ws_profit.merge_cells('A1:C1')
ws_profit['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_profit['A3'] = "ACTUAL NET PROFIT CALCULATION"
ws_profit['A3'].font = Font(bold=True, size=12)
ws_profit.append(["Description", "Amount (₹)"])
style_header(ws_profit, 4, 2)

profit_items = [
    ("TOTAL REVENUE (Base, excl. GST)", "=SUM('Income Tracker'!G:G)", False, "008000"),
    ("", "", False, None),
    ("Less: Operating Expenses", "=SUM('Expense Tracker'!F:F)", False, "FF0000"),
    ("Less: Employee Salaries", "=SUM('Employee Salaries'!N:N)", False, "FF0000"),
    ("Less: Contractor Payments", "=SUM('Contractor Payments'!J:J)", False, "FF0000"),
    ("Less: Stock Purchases", "=SUM('Stock Purchases'!K:K)", False, "FF0000"),
    ("Less: Machine Maintenance", "=SUM('Machine Maintenance'!F:F)", False, "FF0000"),
    ("", "", False, None),
    ("TOTAL EXPENSES", "=B6+B7+B8+B9+B10", True, "FF0000"),
    ("", "", False, None),
    ("NET PROFIT / (LOSS)", "=B5-B12", True, "1F4E79"),
    ("", "", False, None),
    ("Net Profit Margin %", "=IF(B5>0,B14/B5,0)", False, "1F4E79"),
]

for i, (desc, formula, bold, color) in enumerate(profit_items, start=5):
    ws_profit.cell(row=i, column=1).value = desc
    if formula:
        ws_profit.cell(row=i, column=2).value = formula
        ws_profit.cell(row=i, column=2).number_format = currency_format if "%" not in desc else "0.0%"
    if bold:
        ws_profit.cell(row=i, column=1).font = Font(bold=True)
        ws_profit.cell(row=i, column=2).font = Font(bold=True, color=color if color else "000000")
    elif color:
        ws_profit.cell(row=i, column=2).font = Font(color=color)

ws_profit['B14'].fill = profit_fill
ws_profit['B14'].font = Font(bold=True, size=14, color="006600")
ws_profit['B14'].border = thick_border

style_data_rows(ws_profit, 5, 17, 2)
auto_width(ws_profit, 2)
ws_profit.column_dimensions['A'].width = 35
ws_profit.column_dimensions['B'].width = 20

# ============ STOCK REPORTS ============
ws_stock_rep = wb.create_sheet("Stock Reports")
ws_stock_rep['A1'] = "STOCK & INVENTORY REPORTS"
ws_stock_rep.merge_cells('A1:D1')
ws_stock_rep['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_stock_rep['A3'] = "ASSET VALUE SUMMARY"
ws_stock_rep['A3'].font = Font(bold=True, size=12)
ws_stock_rep.append(["Category", "Base Value (₹)", "GST (₹)", "Total (₹)"])
style_header(ws_stock_rep, 4, 4)

for i, cat in enumerate(machine_cats, start=5):
    ws_stock_rep.cell(row=i, column=1).value = cat
    ws_stock_rep.cell(row=i, column=2).value = f'=SUMIF(\'Machines & Equipment\'!D:D,A{i},\'Machines & Equipment\'!F:F)'
    ws_stock_rep.cell(row=i, column=3).value = f'=SUMIF(\'Machines & Equipment\'!D:D,A{i},\'Machines & Equipment\'!H:H)'
    ws_stock_rep.cell(row=i, column=4).value = f'=SUMIF(\'Machines & Equipment\'!D:D,A{i},\'Machines & Equipment\'!I:I)'
    for col in [2, 3, 4]:
        ws_stock_rep.cell(row=i, column=col).number_format = currency_format

ws_stock_rep.cell(row=9, column=1).value = "TOTAL"
ws_stock_rep.cell(row=9, column=1).font = Font(bold=True)
for col in range(2, 5):
    ws_stock_rep.cell(row=9, column=col).value = f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}8)"
    ws_stock_rep.cell(row=9, column=col).font = Font(bold=True)
    ws_stock_rep.cell(row=9, column=col).number_format = currency_format

style_data_rows(ws_stock_rep, 5, 9, 4)
auto_width(ws_stock_rep, 4)
ws_stock_rep.column_dimensions['A'].width = 20

# ============ DASHBOARD ============
ws_dash = wb.create_sheet("Dashboard")

# Title with background
ws_dash['A1'] = "MIS DASHBOARD"
ws_dash.merge_cells('A1:J1')
ws_dash['A1'].font = Font(bold=True, size=22, color="FFFFFF")
ws_dash['A1'].alignment = center_align
ws_dash['A1'].fill = PatternFill("solid", fgColor="1F4E79")

# KEY METRICS SECTION
ws_dash['A3'] = "📊 KEY FINANCIAL METRICS"
ws_dash['A3'].font = Font(bold=True, size=14, color="1F4E79")
ws_dash.merge_cells('A3:E3')

# Revenue Card
ws_dash['A5'] = "TOTAL REVENUE"
ws_dash['A5'].font = Font(bold=True, size=10, color="FFFFFF")
ws_dash['A5'].fill = PatternFill("solid", fgColor="28A745")
ws_dash['A5'].alignment = center_align
ws_dash.merge_cells('A5:B5')
ws_dash['A6'] = "=SUM('Income Tracker'!J:J)"
ws_dash['A6'].font = Font(bold=True, size=16)
ws_dash['A6'].number_format = currency_format
ws_dash['A6'].alignment = center_align
ws_dash.merge_cells('A6:B6')

# Expenses Card
ws_dash['C5'] = "TOTAL EXPENSES"
ws_dash['C5'].font = Font(bold=True, size=10, color="FFFFFF")
ws_dash['C5'].fill = PatternFill("solid", fgColor="DC3545")
ws_dash['C5'].alignment = center_align
ws_dash.merge_cells('C5:D5')
ws_dash['C6'] = "=SUM('Expense Tracker'!I:I)+SUM('Employee Salaries'!N:N)+SUM('Contractor Payments'!J:J)+SUM('Stock Purchases'!N:N)+SUM('Machine Maintenance'!I:I)"
ws_dash['C6'].font = Font(bold=True, size=16)
ws_dash['C6'].number_format = currency_format
ws_dash['C6'].alignment = center_align
ws_dash.merge_cells('C6:D6')

# Net Profit Card
ws_dash['E5'] = "NET PROFIT"
ws_dash['E5'].font = Font(bold=True, size=10, color="FFFFFF")
ws_dash['E5'].fill = PatternFill("solid", fgColor="007BFF")
ws_dash['E5'].alignment = center_align
ws_dash.merge_cells('E5:F5')
ws_dash['E6'] = "=SUM('Income Tracker'!G:G)-SUM('Expense Tracker'!F:F)-SUM('Employee Salaries'!N:N)-SUM('Contractor Payments'!J:J)-SUM('Stock Purchases'!K:K)-SUM('Machine Maintenance'!F:F)"
ws_dash['E6'].font = Font(bold=True, size=16, color="006600")
ws_dash['E6'].number_format = currency_format
ws_dash['E6'].alignment = center_align
ws_dash['E6'].fill = profit_fill
ws_dash.merge_cells('E6:F6')

# Profit Margin Card
ws_dash['G5'] = "PROFIT MARGIN"
ws_dash['G5'].font = Font(bold=True, size=10, color="FFFFFF")
ws_dash['G5'].fill = PatternFill("solid", fgColor="6F42C1")
ws_dash['G5'].alignment = center_align
ws_dash.merge_cells('G5:H5')
ws_dash['G6'] = "=IF(SUM('Income Tracker'!G:G)>0,E6/SUM('Income Tracker'!G:G),0)"
ws_dash['G6'].font = Font(bold=True, size=16)
ws_dash['G6'].number_format = "0.0%"
ws_dash['G6'].alignment = center_align
ws_dash.merge_cells('G6:H6')

# EXPENSE BREAKDOWN
ws_dash['A8'] = "EXPENSE BREAKDOWN"
ws_dash['A8'].font = Font(bold=True, size=14, color="1F4E79")
ws_dash.merge_cells('A8:B8')

expense_breakdown = [
    ("Operating Expenses", "=SUM('Expense Tracker'!I:I)"),
    ("Employee Salaries", "=SUM('Employee Salaries'!N:N)"),
    ("Contractor Payments", "=SUM('Contractor Payments'!J:J)"),
    ("Stock Purchases", "=SUM('Stock Purchases'!N:N)"),
    ("Machine Maintenance", "=SUM('Machine Maintenance'!I:I)"),
]

ws_dash['A9'] = "Expense Type"
ws_dash['B9'] = "Amount (₹)"
ws_dash['A9'].font = Font(bold=True, color="FFFFFF")
ws_dash['B9'].font = Font(bold=True, color="FFFFFF")
ws_dash['A9'].fill = header_fill
ws_dash['B9'].fill = header_fill

for i, (name, formula) in enumerate(expense_breakdown, start=10):
    ws_dash.cell(row=i, column=1).value = name
    ws_dash.cell(row=i, column=2).value = formula
    ws_dash.cell(row=i, column=2).number_format = currency_format
    ws_dash.cell(row=i, column=1).border = thin_border
    ws_dash.cell(row=i, column=2).border = thin_border
    if i % 2 == 0:
        ws_dash.cell(row=i, column=1).fill = alt_row_fill
        ws_dash.cell(row=i, column=2).fill = alt_row_fill

ws_dash['A15'] = "TOTAL"
ws_dash['B15'] = "=SUM(B10:B14)"
ws_dash['A15'].font = Font(bold=True)
ws_dash['B15'].font = Font(bold=True)
ws_dash['B15'].number_format = currency_format
ws_dash['A15'].fill = PatternFill("solid", fgColor="FFC107")
ws_dash['B15'].fill = PatternFill("solid", fgColor="FFC107")

# GST & TDS SECTION
ws_dash['D8'] = "TAX SUMMARY"
ws_dash['D8'].font = Font(bold=True, size=14, color="1F4E79")

ws_dash['D9'] = "Tax Type"
ws_dash['E9'] = "Amount (₹)"
ws_dash['D9'].font = Font(bold=True, color="FFFFFF")
ws_dash['E9'].font = Font(bold=True, color="FFFFFF")
ws_dash['D9'].fill = header_fill
ws_dash['E9'].fill = header_fill

tax_items = [
    ("GST Collected", "=SUM('Income Tracker'!I:I)"),
    ("GST Paid", "=SUM('Expense Tracker'!H:H)+SUM('Stock Purchases'!M:M)+SUM('Machine Maintenance'!H:H)"),
    ("Net GST Payable", "=E10-E11"),
    ("TDS Deducted", "=SUM('Contractor Payments'!I:I)"),
    ("TDS Pending Deposit", "=SUM('Contractor Payments'!I:I)-SUMIF('Contractor Payments'!N:N,\"Yes\",'Contractor Payments'!I:I)"),
]

for i, (name, formula) in enumerate(tax_items, start=10):
    ws_dash.cell(row=i, column=4).value = name
    ws_dash.cell(row=i, column=5).value = formula
    ws_dash.cell(row=i, column=5).number_format = currency_format
    ws_dash.cell(row=i, column=4).border = thin_border
    ws_dash.cell(row=i, column=5).border = thin_border
    if i % 2 == 0:
        ws_dash.cell(row=i, column=4).fill = alt_row_fill
        ws_dash.cell(row=i, column=5).fill = alt_row_fill

# PAYROLL SUMMARY
ws_dash['G8'] = "PAYROLL SUMMARY"
ws_dash['G8'].font = Font(bold=True, size=14, color="1F4E79")

ws_dash['G9'] = "Metric"
ws_dash['H9'] = "Amount (₹)"
ws_dash['G9'].font = Font(bold=True, color="FFFFFF")
ws_dash['H9'].font = Font(bold=True, color="FFFFFF")
ws_dash['G9'].fill = header_fill
ws_dash['H9'].fill = header_fill

payroll_items = [
    ("Total Gross Salary", "=SUM('Employee Salaries'!I:I)"),
    ("Total PF Deduction", "=SUM('Employee Salaries'!J:J)"),
    ("Total ESI Deduction", "=SUM('Employee Salaries'!K:K)"),
    ("Total Net Salary Paid", "=SUM('Employee Salaries'!N:N)"),
    ("Total Employees", "=COUNTA('Employee Salaries'!B:B)-1"),
]

for i, (name, formula) in enumerate(payroll_items, start=10):
    ws_dash.cell(row=i, column=7).value = name
    ws_dash.cell(row=i, column=8).value = formula
    if "Employees" not in name:
        ws_dash.cell(row=i, column=8).number_format = currency_format
    ws_dash.cell(row=i, column=7).border = thin_border
    ws_dash.cell(row=i, column=8).border = thin_border
    if i % 2 == 0:
        ws_dash.cell(row=i, column=7).fill = alt_row_fill
        ws_dash.cell(row=i, column=8).fill = alt_row_fill

# Chart Data (for service revenue)
ws_dash['A18'] = "Service"
ws_dash['B18'] = "Revenue"
for i, svc in enumerate(services, start=19):
    ws_dash.cell(row=i, column=1).value = svc
    ws_dash.cell(row=i, column=2).value = f'=SUMIF(\'Income Tracker\'!D:D,A{i},\'Income Tracker\'!J:J)'

# Create Revenue by Service Chart
chart = PieChart()
chart.title = "Revenue by Service"
data = Reference(ws_dash, min_col=2, min_row=18, max_row=26)
cats = Reference(ws_dash, min_col=1, min_row=19, max_row=26)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.dataLabels = DataLabelList()
chart.dataLabels.showPercent = True
chart.width = 14
chart.height = 10
ws_dash.add_chart(chart, "D18")

# Set column widths
for col in range(1, 11):
    ws_dash.column_dimensions[get_column_letter(col)].width = 16

ws_dash.freeze_panes = 'A3'

# Save
wb.save('/var/lib/freelancer/projects/40182876/MIS_System_v4.xlsx')
print("MIS System v4 created successfully!")
