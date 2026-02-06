from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import random

wb = Workbook()

# Styles
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill("solid", fgColor="2E75B6")
alt_row_fill = PatternFill("solid", fgColor="D6EAF8")
currency_format = '₹#,##0.00'
date_format = 'DD-MMM-YYYY'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

def style_data_rows(ws, start_row, end_row, cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = alt_row_fill

def auto_width(ws, cols):
    for col in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

# ============ HOW TO USE ============
ws_how = wb.active
ws_how.title = "HOW TO USE"
instructions = [
    ["", "MIS SYSTEM - USER GUIDE", ""],
    [""],
    ["SHEET", "PURPOSE", "HOW TO USE"],
    ["Income Tracker", "Record all service revenue", "Enter each payment received with service type and project details"],
    ["Expense Tracker", "Track all business expenses", "Log every expense with category and payment details"],
    ["Machines & Equipment", "Asset register for all machines", "Add each machine with unique ID (M001, M002...)"],
    ["Machine Maintenance", "Track repairs and servicing", "Log every repair/service against machine ID"],
    ["Chemicals Stock", "Manage chemical inventory", "Update stock added and used regularly"],
    ["Accessories Stock", "Track consumables inventory", "Update purchases and usage daily"],
    ["Stock Transactions", "Detailed stock movement log", "Record every purchase/usage transaction"],
    ["Income Summary", "Auto-calculated revenue reports", "View only - auto-updates from Income Tracker"],
    ["Expense Summary", "Auto-calculated expense reports", "View only - auto-updates from Expense Tracker"],
    ["Profitability Report", "Profit analysis", "View only - auto-calculates profit margins"],
    ["Stock Reports", "Inventory status and alerts", "View only - shows low stock and asset values"],
    ["Dashboard", "Visual overview with charts", "View only - key metrics at a glance"],
    [""],
    ["TIPS:", "", ""],
    ["1. Use dropdowns for consistency", "", ""],
    ["2. Enter data daily for accurate reports", "", ""],
    ["3. Check Dashboard weekly for insights", "", ""],
    ["4. Monitor Low Stock alerts regularly", "", ""],
]
for row_data in instructions:
    ws_how.append(row_data)
ws_how.merge_cells('A1:C1')
ws_how['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws_how['A1'].alignment = center_align
style_header(ws_how, 3, 3)
for row in range(4, 15):
    for col in range(1, 4):
        ws_how.cell(row=row, column=col).border = thin_border
ws_how.column_dimensions['A'].width = 25
ws_how.column_dimensions['B'].width = 35
ws_how.column_dimensions['C'].width = 50

# ============ INCOME TRACKER ============
ws_income = wb.create_sheet("Income Tracker")
income_headers = ["Date", "Invoice No", "Customer Name", "Service Type", "Project Type", "Project Name", "Amount (₹)", "Payment Mode", "Payment Status", "Notes"]
ws_income.append(income_headers)
style_header(ws_income, 1, len(income_headers))

services = ["Deep Cleaning", "Regular Cleaning", "Pest Control", "Painting", "Plumbing", "Electrical", "Carpentry", "AC Service"]
project_types = ["Individual", "Apartment Bulk", "Commercial"]
payment_modes = ["Cash", "UPI", "Bank Transfer", "Card"]
payment_status = ["Received", "Pending"]
customers = ["Rajesh Kumar", "Priya Sharma", "Amit Patel", "Sunita Reddy", "Vikram Singh", "Deepa Nair", "Karthik Iyer", "Meena Gupta", "Rahul Verma", "Anjali Menon", "Suresh Rao", "Lakshmi Pillai", "Arun Krishnan", "Kavitha Srinivasan", "Manoj Das"]

income_data = []
base_date = datetime(2025, 1, 1)
for i in range(15):
    date = base_date + timedelta(days=random.randint(0, 28))
    invoice = f"INV-2025-{1001+i}"
    customer = customers[i]
    service = random.choice(services)
    proj_type = random.choice(project_types)
    proj_name = f"{'Green Valley Apartments' if proj_type == 'Apartment Bulk' else 'Tech Park' if proj_type == 'Commercial' else ''}"
    amount = random.randint(2000, 25000) if proj_type == "Individual" else random.randint(50000, 200000)
    mode = random.choice(payment_modes)
    status = random.choice(payment_status)
    notes = ""
    income_data.append([date, invoice, customer, service, proj_type, proj_name, amount, mode, status, notes])

for row in income_data:
    ws_income.append(row)

style_data_rows(ws_income, 2, 16, len(income_headers))
ws_income.column_dimensions['A'].width = 12
ws_income.column_dimensions['B'].width = 15
ws_income.column_dimensions['C'].width = 18
ws_income.column_dimensions['D'].width = 16
ws_income.column_dimensions['E'].width = 14
ws_income.column_dimensions['F'].width = 22
ws_income.column_dimensions['G'].width = 14
ws_income.column_dimensions['G'].number_format = currency_format
for row in range(2, 17):
    ws_income.cell(row=row, column=7).number_format = currency_format
    ws_income.cell(row=row, column=1).number_format = date_format
ws_income.freeze_panes = 'A2'

dv_service = DataValidation(type="list", formula1='"Deep Cleaning,Regular Cleaning,Pest Control,Painting,Plumbing,Electrical,Carpentry,AC Service"', allow_blank=True)
dv_projtype = DataValidation(type="list", formula1='"Individual,Apartment Bulk,Commercial"', allow_blank=True)
dv_paymode = DataValidation(type="list", formula1='"Cash,UPI,Bank Transfer,Card"', allow_blank=True)
dv_paystatus = DataValidation(type="list", formula1='"Received,Pending"', allow_blank=True)
ws_income.add_data_validation(dv_service)
ws_income.add_data_validation(dv_projtype)
ws_income.add_data_validation(dv_paymode)
ws_income.add_data_validation(dv_paystatus)
dv_service.add('D2:D1000')
dv_projtype.add('E2:E1000')
dv_paymode.add('H2:H1000')
dv_paystatus.add('I2:I1000')

# ============ EXPENSE TRACKER ============
ws_expense = wb.create_sheet("Expense Tracker")
expense_headers = ["Date", "Expense ID", "Category", "Description", "Vendor/Payee", "Amount (₹)", "Payment Mode", "Payment Status", "Notes"]
ws_expense.append(expense_headers)
style_header(ws_expense, 1, len(expense_headers))

expense_cats = ["Staff Salary", "Vendor Payment", "Travel & Fuel", "Marketing", "Office Expenses", "Miscellaneous"]
vendors = ["Ramesh (Driver)", "Kumar Chemicals", "City Fuel Station", "Facebook Ads", "Airtel", "Office Depot", "Suresh (Technician)", "Local Transport", "Google Ads", "Electricity Board", "Water Supply", "Security Agency"]
expense_data = []
for i in range(20):
    date = base_date + timedelta(days=random.randint(0, 28))
    exp_id = f"EXP-{2025}-{101+i}"
    cat = random.choice(expense_cats)
    if cat == "Staff Salary":
        desc = f"Salary - {random.choice(['Driver', 'Technician', 'Helper', 'Supervisor'])}"
        vendor = random.choice(["Ramesh", "Kumar", "Suresh", "Anil"])
        amount = random.randint(15000, 35000)
    elif cat == "Travel & Fuel":
        desc = "Petrol/Diesel"
        vendor = "City Fuel Station"
        amount = random.randint(500, 3000)
    elif cat == "Marketing":
        desc = random.choice(["Facebook Ads", "Google Ads", "Pamphlets"])
        vendor = desc
        amount = random.randint(2000, 15000)
    else:
        desc = random.choice(["Office Supplies", "Utility Bill", "Misc Purchase"])
        vendor = random.choice(vendors)
        amount = random.randint(500, 5000)
    mode = random.choice(payment_modes)
    status = random.choice(payment_status)
    expense_data.append([date, exp_id, cat, desc, vendor, amount, mode, status, ""])

for row in expense_data:
    ws_expense.append(row)

style_data_rows(ws_expense, 2, 21, len(expense_headers))
auto_width(ws_expense, len(expense_headers))
ws_expense.column_dimensions['D'].width = 25
for row in range(2, 22):
    ws_expense.cell(row=row, column=6).number_format = currency_format
    ws_expense.cell(row=row, column=1).number_format = date_format
ws_expense.freeze_panes = 'A2'

dv_expcat = DataValidation(type="list", formula1='"Staff Salary,Vendor Payment,Travel & Fuel,Marketing,Office Expenses,Miscellaneous"', allow_blank=True)
ws_expense.add_data_validation(dv_expcat)
ws_expense.add_data_validation(dv_paymode)
ws_expense.add_data_validation(dv_paystatus)
dv_expcat.add('C2:C1000')

# ============ MACHINES & EQUIPMENT ============
ws_machines = wb.create_sheet("Machines & Equipment")
machine_headers = ["Machine ID", "Machine Name", "Model", "Category", "Purchase Date", "Purchase Cost (₹)", "Current Status", "Location/Assigned To", "Notes"]
ws_machines.append(machine_headers)
style_header(ws_machines, 1, len(machine_headers))

machine_cats = ["Cleaning Machine", "Spray Equipment", "Power Tools", "Safety Equipment"]
machine_data = [
    ["M001", "Floor Scrubber", "Karcher BD 50/50", "Cleaning Machine", datetime(2023, 3, 15), 85000, "Active", "Team A", ""],
    ["M002", "Pressure Washer", "Bosch AQT 45-14X", "Cleaning Machine", datetime(2023, 5, 20), 45000, "Active", "Team B", ""],
    ["M003", "Vacuum Cleaner", "Eureka Forbes Pro", "Cleaning Machine", datetime(2022, 8, 10), 28000, "Active", "Team A", ""],
    ["M004", "Pest Control Sprayer", "Solo 475", "Spray Equipment", datetime(2023, 1, 5), 15000, "Active", "Pest Team", ""],
    ["M005", "Paint Sprayer", "Graco Magnum X5", "Spray Equipment", datetime(2023, 7, 12), 65000, "Under Repair", "Painting Team", "Motor issue"],
    ["M006", "Drill Machine", "Bosch GSB 600", "Power Tools", datetime(2022, 4, 18), 8500, "Active", "General", ""],
    ["M007", "Angle Grinder", "Makita GA4030", "Power Tools", datetime(2023, 2, 22), 6500, "Active", "General", ""],
    ["M008", "Ladder (12ft)", "Aluminium", "Safety Equipment", datetime(2022, 1, 10), 12000, "Active", "Warehouse", ""],
    ["M009", "Steam Cleaner", "Karcher SC3", "Cleaning Machine", datetime(2024, 1, 8), 35000, "Active", "Team C", ""],
    ["M010", "Industrial Blower", "Dewalt DCE100", "Cleaning Machine", datetime(2023, 11, 5), 18000, "Retired", "Storage", "Replaced"],
]

for row in machine_data:
    ws_machines.append(row)

style_data_rows(ws_machines, 2, 11, len(machine_headers))
auto_width(ws_machines, len(machine_headers))
ws_machines.column_dimensions['B'].width = 20
ws_machines.column_dimensions['C'].width = 20
for row in range(2, 12):
    ws_machines.cell(row=row, column=6).number_format = currency_format
    ws_machines.cell(row=row, column=5).number_format = date_format
ws_machines.freeze_panes = 'A2'

dv_machcat = DataValidation(type="list", formula1='"Cleaning Machine,Spray Equipment,Power Tools,Safety Equipment"', allow_blank=True)
dv_machstatus = DataValidation(type="list", formula1='"Active,Under Repair,Retired"', allow_blank=True)
ws_machines.add_data_validation(dv_machcat)
ws_machines.add_data_validation(dv_machstatus)
dv_machcat.add('D2:D1000')
dv_machstatus.add('G2:G1000')

# ============ MACHINE MAINTENANCE LOG ============
ws_maint = wb.create_sheet("Machine Maintenance")
maint_headers = ["Date", "Machine ID", "Machine Name", "Maintenance Type", "Description", "Cost (₹)", "Done By", "Next Service Due", "Notes"]
ws_maint.append(maint_headers)
style_header(ws_maint, 1, len(maint_headers))

maint_types = ["Repair", "Service", "Parts Replacement"]
maint_data = [
    [datetime(2024, 2, 10), "M001", "Floor Scrubber", "Service", "Regular servicing", 2500, "Karcher Service", datetime(2024, 8, 10), ""],
    [datetime(2024, 3, 15), "M002", "Pressure Washer", "Parts Replacement", "Nozzle replaced", 1200, "Local Mechanic", datetime(2024, 9, 15), ""],
    [datetime(2024, 4, 5), "M003", "Vacuum Cleaner", "Service", "Filter cleaning", 800, "In-house", datetime(2024, 10, 5), ""],
    [datetime(2024, 5, 20), "M005", "Paint Sprayer", "Repair", "Motor rewinding", 8500, "Graco Service", datetime(2024, 11, 20), "Under warranty"],
    [datetime(2024, 6, 8), "M004", "Pest Control Sprayer", "Service", "Valve check", 500, "In-house", datetime(2024, 12, 8), ""],
    [datetime(2024, 7, 12), "M006", "Drill Machine", "Parts Replacement", "Chuck replaced", 1500, "Bosch Service", datetime(2025, 1, 12), ""],
    [datetime(2024, 8, 1), "M001", "Floor Scrubber", "Service", "Brush replacement", 4500, "Karcher Service", datetime(2025, 2, 1), ""],
    [datetime(2024, 8, 18), "M007", "Angle Grinder", "Repair", "Switch repair", 600, "Local Mechanic", datetime(2025, 2, 18), ""],
    [datetime(2024, 9, 5), "M009", "Steam Cleaner", "Service", "Descaling", 1200, "Karcher Service", datetime(2025, 3, 5), ""],
    [datetime(2024, 10, 10), "M002", "Pressure Washer", "Service", "Regular maintenance", 1800, "Bosch Service", datetime(2025, 4, 10), ""],
    [datetime(2024, 11, 15), "M003", "Vacuum Cleaner", "Parts Replacement", "Motor replaced", 5500, "Eureka Service", datetime(2025, 5, 15), ""],
    [datetime(2024, 12, 1), "M005", "Paint Sprayer", "Repair", "Hose replacement", 2200, "Graco Service", datetime(2025, 6, 1), "Still under repair"],
    [datetime(2025, 1, 5), "M004", "Pest Control Sprayer", "Service", "Annual service", 1000, "Solo Service", datetime(2025, 7, 5), ""],
    [datetime(2025, 1, 15), "M006", "Drill Machine", "Service", "Oiling & cleaning", 300, "In-house", datetime(2025, 7, 15), ""],
    [datetime(2025, 1, 20), "M009", "Steam Cleaner", "Parts Replacement", "Gasket replaced", 800, "Karcher Service", datetime(2025, 7, 20), ""],
]

for row in maint_data:
    ws_maint.append(row)

style_data_rows(ws_maint, 2, 16, len(maint_headers))
auto_width(ws_maint, len(maint_headers))
ws_maint.column_dimensions['E'].width = 25
for row in range(2, 17):
    ws_maint.cell(row=row, column=6).number_format = currency_format
    ws_maint.cell(row=row, column=1).number_format = date_format
    ws_maint.cell(row=row, column=8).number_format = date_format
ws_maint.freeze_panes = 'A2'

dv_mainttype = DataValidation(type="list", formula1='"Repair,Service,Parts Replacement"', allow_blank=True)
dv_machid = DataValidation(type="list", formula1='"M001,M002,M003,M004,M005,M006,M007,M008,M009,M010"', allow_blank=True)
ws_maint.add_data_validation(dv_mainttype)
ws_maint.add_data_validation(dv_machid)
dv_mainttype.add('D2:D1000')
dv_machid.add('B2:B1000')

# ============ CHEMICALS STOCK ============
ws_chem = wb.create_sheet("Chemicals Stock")
chem_headers = ["Item ID", "Chemical Name", "Brand", "Category", "Unit", "Opening Stock", "Stock Added", "Stock Used", "Closing Stock", "Min Level", "Status", "Last Updated"]
ws_chem.append(chem_headers)
style_header(ws_chem, 1, len(chem_headers))

chem_cats = ["Cleaning Agents", "Pest Control", "Sanitizers", "Specialty"]
chem_units = ["Liters", "KG", "Bottles", "Cans"]
chem_data = [
    ["CH001", "Floor Cleaner", "Lizol", "Cleaning Agents", "Liters", 50, 100, 80, None, 20, None, datetime(2025, 1, 25)],
    ["CH002", "Glass Cleaner", "Colin", "Cleaning Agents", "Liters", 30, 50, 45, None, 15, None, datetime(2025, 1, 25)],
    ["CH003", "Toilet Cleaner", "Harpic", "Cleaning Agents", "Liters", 40, 80, 70, None, 20, None, datetime(2025, 1, 25)],
    ["CH004", "Pesticide Spray", "Baygon Pro", "Pest Control", "Cans", 25, 50, 55, None, 10, None, datetime(2025, 1, 25)],
    ["CH005", "Termite Solution", "Premise", "Pest Control", "Liters", 20, 40, 35, None, 10, None, datetime(2025, 1, 25)],
    ["CH006", "Hand Sanitizer", "Dettol", "Sanitizers", "Liters", 15, 30, 28, None, 10, None, datetime(2025, 1, 25)],
    ["CH007", "Surface Disinfectant", "Lysol", "Sanitizers", "Liters", 25, 50, 40, None, 15, None, datetime(2025, 1, 25)],
    ["CH008", "Degreaser", "WD-40", "Specialty", "Cans", 10, 20, 18, None, 5, None, datetime(2025, 1, 25)],
    ["CH009", "Stain Remover", "Vanish", "Specialty", "KG", 8, 15, 12, None, 5, None, datetime(2025, 1, 25)],
    ["CH010", "Air Freshener", "Odonil", "Specialty", "Cans", 30, 60, 65, None, 15, None, datetime(2025, 1, 25)],
    ["CH011", "Cockroach Gel", "Advion", "Pest Control", "Bottles", 20, 30, 25, None, 10, None, datetime(2025, 1, 25)],
    ["CH012", "Fabric Softener", "Comfort", "Cleaning Agents", "Liters", 15, 25, 22, None, 10, None, datetime(2025, 1, 25)],
]

for i, row in enumerate(chem_data, start=2):
    ws_chem.append(row[:8])
    ws_chem.cell(row=i, column=9).value = f"=F{i}+G{i}-H{i}"
    ws_chem.cell(row=i, column=10).value = row[9]
    ws_chem.cell(row=i, column=11).value = f'=IF(I{i}<J{i},"Low Stock","OK")'
    ws_chem.cell(row=i, column=12).value = row[11]

style_data_rows(ws_chem, 2, 13, len(chem_headers))
auto_width(ws_chem, len(chem_headers))
ws_chem.column_dimensions['B'].width = 22
for row in range(2, 14):
    ws_chem.cell(row=row, column=12).number_format = date_format
ws_chem.freeze_panes = 'A2'

dv_chemcat = DataValidation(type="list", formula1='"Cleaning Agents,Pest Control,Sanitizers,Specialty"', allow_blank=True)
dv_chemunit = DataValidation(type="list", formula1='"Liters,KG,Bottles,Cans"', allow_blank=True)
ws_chem.add_data_validation(dv_chemcat)
ws_chem.add_data_validation(dv_chemunit)
dv_chemcat.add('D2:D1000')
dv_chemunit.add('E2:E1000')

low_stock_fill = PatternFill("solid", fgColor="FFCCCC")
ok_fill = PatternFill("solid", fgColor="CCFFCC")
ws_chem.conditional_formatting.add('K2:K1000', FormulaRule(formula=['K2="Low Stock"'], fill=low_stock_fill))
ws_chem.conditional_formatting.add('K2:K1000', FormulaRule(formula=['K2="OK"'], fill=ok_fill))

# ============ ACCESSORIES STOCK ============
ws_acc = wb.create_sheet("Accessories Stock")
acc_headers = ["Item ID", "Item Name", "Category", "Unit", "Opening Stock", "Stock Added", "Stock Used", "Closing Stock", "Min Level", "Status", "Last Updated"]
ws_acc.append(acc_headers)
style_header(ws_acc, 1, len(acc_headers))

acc_cats = ["Pipes & Hoses", "Brushes & Mops", "Cloths & Wipes", "Gloves & Safety", "Tools", "Other"]
acc_data = [
    ["AC001", "Rubber Hose (10m)", "Pipes & Hoses", "Pieces", 10, 15, 12, None, 5, None, datetime(2025, 1, 25)],
    ["AC002", "Floor Mop", "Brushes & Mops", "Pieces", 20, 30, 28, None, 10, None, datetime(2025, 1, 25)],
    ["AC003", "Scrub Brush", "Brushes & Mops", "Pieces", 25, 40, 35, None, 10, None, datetime(2025, 1, 25)],
    ["AC004", "Microfiber Cloth", "Cloths & Wipes", "Pieces", 50, 100, 90, None, 25, None, datetime(2025, 1, 25)],
    ["AC005", "Rubber Gloves", "Gloves & Safety", "Pairs", 30, 50, 48, None, 15, None, datetime(2025, 1, 25)],
    ["AC006", "Safety Goggles", "Gloves & Safety", "Pieces", 10, 15, 8, None, 5, None, datetime(2025, 1, 25)],
    ["AC007", "Spray Bottles", "Other", "Pieces", 20, 30, 22, None, 10, None, datetime(2025, 1, 25)],
    ["AC008", "Dust Pan Set", "Brushes & Mops", "Sets", 15, 20, 18, None, 8, None, datetime(2025, 1, 25)],
    ["AC009", "Extension Cord", "Tools", "Pieces", 8, 10, 7, None, 4, None, datetime(2025, 1, 25)],
    ["AC010", "Bucket (20L)", "Other", "Pieces", 25, 30, 28, None, 10, None, datetime(2025, 1, 25)],
    ["AC011", "Face Masks", "Gloves & Safety", "Boxes", 10, 20, 18, None, 5, None, datetime(2025, 1, 25)],
    ["AC012", "Sponge Pack", "Cloths & Wipes", "Packs", 20, 35, 30, None, 10, None, datetime(2025, 1, 25)],
    ["AC013", "Squeegee", "Brushes & Mops", "Pieces", 12, 18, 15, None, 6, None, datetime(2025, 1, 25)],
    ["AC014", "Wire Brush", "Brushes & Mops", "Pieces", 15, 20, 22, None, 8, None, datetime(2025, 1, 25)],
    ["AC015", "Tarpaulin Sheet", "Other", "Pieces", 8, 12, 10, None, 4, None, datetime(2025, 1, 25)],
]

for i, row in enumerate(acc_data, start=2):
    ws_acc.append(row[:7])
    ws_acc.cell(row=i, column=8).value = f"=E{i}+F{i}-G{i}"
    ws_acc.cell(row=i, column=9).value = row[8]
    ws_acc.cell(row=i, column=10).value = f'=IF(H{i}<I{i},"Low Stock","OK")'
    ws_acc.cell(row=i, column=11).value = row[10]

style_data_rows(ws_acc, 2, 16, len(acc_headers))
auto_width(ws_acc, len(acc_headers))
ws_acc.column_dimensions['B'].width = 20
for row in range(2, 17):
    ws_acc.cell(row=row, column=11).number_format = date_format
ws_acc.freeze_panes = 'A2'

dv_acccat = DataValidation(type="list", formula1='"Pipes & Hoses,Brushes & Mops,Cloths & Wipes,Gloves & Safety,Tools,Other"', allow_blank=True)
ws_acc.add_data_validation(dv_acccat)
dv_acccat.add('C2:C1000')
ws_acc.conditional_formatting.add('J2:J1000', FormulaRule(formula=['J2="Low Stock"'], fill=low_stock_fill))
ws_acc.conditional_formatting.add('J2:J1000', FormulaRule(formula=['J2="OK"'], fill=ok_fill))

# ============ STOCK TRANSACTIONS ============
ws_trans = wb.create_sheet("Stock Transactions")
trans_headers = ["Date", "Transaction ID", "Item Type", "Item ID", "Item Name", "Transaction Type", "Quantity", "Unit Cost (₹)", "Total Cost (₹)", "Project/Team", "Notes"]
ws_trans.append(trans_headers)
style_header(ws_trans, 1, len(trans_headers))

item_types = ["Chemical", "Accessory", "Machine"]
trans_types = ["Purchase", "Used", "Returned", "Damaged"]
trans_data = [
    [datetime(2025, 1, 2), "TXN001", "Chemical", "CH001", "Floor Cleaner", "Purchase", 50, 120, None, "Warehouse", "Monthly stock"],
    [datetime(2025, 1, 3), "TXN002", "Chemical", "CH004", "Pesticide Spray", "Purchase", 25, 350, None, "Warehouse", ""],
    [datetime(2025, 1, 5), "TXN003", "Accessory", "AC004", "Microfiber Cloth", "Purchase", 50, 80, None, "Warehouse", "Bulk order"],
    [datetime(2025, 1, 6), "TXN004", "Chemical", "CH001", "Floor Cleaner", "Used", 20, 0, None, "Team A", "Green Valley project"],
    [datetime(2025, 1, 7), "TXN005", "Accessory", "AC005", "Rubber Gloves", "Used", 10, 0, None, "Team B", ""],
    [datetime(2025, 1, 8), "TXN006", "Chemical", "CH002", "Glass Cleaner", "Purchase", 25, 95, None, "Warehouse", ""],
    [datetime(2025, 1, 10), "TXN007", "Accessory", "AC002", "Floor Mop", "Damaged", 2, 0, None, "Team A", "Worn out"],
    [datetime(2025, 1, 11), "TXN008", "Chemical", "CH007", "Surface Disinfectant", "Used", 15, 0, None, "Team C", "Hospital contract"],
    [datetime(2025, 1, 12), "TXN009", "Accessory", "AC003", "Scrub Brush", "Purchase", 20, 45, None, "Warehouse", ""],
    [datetime(2025, 1, 14), "TXN010", "Chemical", "CH005", "Termite Solution", "Purchase", 20, 850, None, "Warehouse", ""],
    [datetime(2025, 1, 15), "TXN011", "Accessory", "AC001", "Rubber Hose (10m)", "Used", 3, 0, None, "Pest Team", ""],
    [datetime(2025, 1, 16), "TXN012", "Chemical", "CH003", "Toilet Cleaner", "Used", 25, 0, None, "Team B", ""],
    [datetime(2025, 1, 18), "TXN013", "Accessory", "AC011", "Face Masks", "Purchase", 10, 250, None, "Warehouse", ""],
    [datetime(2025, 1, 19), "TXN014", "Chemical", "CH006", "Hand Sanitizer", "Used", 8, 0, None, "All Teams", ""],
    [datetime(2025, 1, 20), "TXN015", "Accessory", "AC007", "Spray Bottles", "Returned", 3, 0, None, "Team A", "Extra stock"],
    [datetime(2025, 1, 21), "TXN016", "Chemical", "CH010", "Air Freshener", "Purchase", 30, 180, None, "Warehouse", ""],
    [datetime(2025, 1, 22), "TXN017", "Accessory", "AC006", "Safety Goggles", "Used", 3, 0, None, "Painting Team", ""],
    [datetime(2025, 1, 23), "TXN018", "Chemical", "CH008", "Degreaser", "Used", 5, 0, None, "Team C", "Kitchen cleaning"],
    [datetime(2025, 1, 24), "TXN019", "Accessory", "AC012", "Sponge Pack", "Purchase", 15, 120, None, "Warehouse", ""],
    [datetime(2025, 1, 25), "TXN020", "Chemical", "CH011", "Cockroach Gel", "Used", 8, 0, None, "Pest Team", ""],
]

for i, row in enumerate(trans_data, start=2):
    ws_trans.append(row[:8])
    ws_trans.cell(row=i, column=9).value = f"=G{i}*H{i}"
    ws_trans.cell(row=i, column=10).value = row[9]
    ws_trans.cell(row=i, column=11).value = row[10]

style_data_rows(ws_trans, 2, 21, len(trans_headers))
auto_width(ws_trans, len(trans_headers))
for row in range(2, 22):
    ws_trans.cell(row=row, column=8).number_format = currency_format
    ws_trans.cell(row=row, column=9).number_format = currency_format
    ws_trans.cell(row=row, column=1).number_format = date_format
ws_trans.freeze_panes = 'A2'

dv_itemtype = DataValidation(type="list", formula1='"Chemical,Accessory,Machine"', allow_blank=True)
dv_transtype = DataValidation(type="list", formula1='"Purchase,Used,Returned,Damaged"', allow_blank=True)
ws_trans.add_data_validation(dv_itemtype)
ws_trans.add_data_validation(dv_transtype)
dv_itemtype.add('C2:C1000')
dv_transtype.add('F2:F1000')

# ============ INCOME SUMMARY ============
ws_inc_sum = wb.create_sheet("Income Summary")

ws_inc_sum['A1'] = "INCOME SUMMARY REPORT"
ws_inc_sum.merge_cells('A1:D1')
ws_inc_sum['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_inc_sum['A3'] = "SERVICE-WISE REVENUE"
ws_inc_sum['A3'].font = Font(bold=True, size=12)
ws_inc_sum.append(["Service Type", "Total Revenue (₹)", "Order Count", "Avg Order Value"])
style_header(ws_inc_sum, 4, 4)

for i, svc in enumerate(services, start=5):
    ws_inc_sum.cell(row=i, column=1).value = svc
    ws_inc_sum.cell(row=i, column=2).value = f'=SUMIF(\'Income Tracker\'!D:D,A{i},\'Income Tracker\'!G:G)'
    ws_inc_sum.cell(row=i, column=3).value = f'=COUNTIF(\'Income Tracker\'!D:D,A{i})'
    ws_inc_sum.cell(row=i, column=4).value = f'=IF(C{i}>0,B{i}/C{i},0)'
    ws_inc_sum.cell(row=i, column=2).number_format = currency_format
    ws_inc_sum.cell(row=i, column=4).number_format = currency_format

ws_inc_sum.cell(row=13, column=1).value = "TOTAL"
ws_inc_sum.cell(row=13, column=1).font = Font(bold=True)
ws_inc_sum.cell(row=13, column=2).value = "=SUM(B5:B12)"
ws_inc_sum.cell(row=13, column=2).font = Font(bold=True)
ws_inc_sum.cell(row=13, column=2).number_format = currency_format
ws_inc_sum.cell(row=13, column=3).value = "=SUM(C5:C12)"
ws_inc_sum.cell(row=13, column=3).font = Font(bold=True)

style_data_rows(ws_inc_sum, 5, 13, 4)

ws_inc_sum['A16'] = "PROJECT TYPE REVENUE"
ws_inc_sum['A16'].font = Font(bold=True, size=12)
ws_inc_sum.append(["Project Type", "Total Revenue (₹)", "Order Count"])
style_header(ws_inc_sum, 17, 3)

for i, pt in enumerate(project_types, start=18):
    ws_inc_sum.cell(row=i, column=1).value = pt
    ws_inc_sum.cell(row=i, column=2).value = f'=SUMIF(\'Income Tracker\'!E:E,A{i},\'Income Tracker\'!G:G)'
    ws_inc_sum.cell(row=i, column=3).value = f'=COUNTIF(\'Income Tracker\'!E:E,A{i})'
    ws_inc_sum.cell(row=i, column=2).number_format = currency_format

style_data_rows(ws_inc_sum, 18, 20, 3)

ws_inc_sum['A23'] = "PAYMENT STATUS"
ws_inc_sum['A23'].font = Font(bold=True, size=12)
ws_inc_sum.append(["Status", "Amount (₹)", "Count"])
style_header(ws_inc_sum, 24, 3)
ws_inc_sum['A25'] = "Received"
ws_inc_sum['B25'] = "=SUMIF('Income Tracker'!I:I,\"Received\",'Income Tracker'!G:G)"
ws_inc_sum['C25'] = "=COUNTIF('Income Tracker'!I:I,\"Received\")"
ws_inc_sum['A26'] = "Pending"
ws_inc_sum['B26'] = "=SUMIF('Income Tracker'!I:I,\"Pending\",'Income Tracker'!G:G)"
ws_inc_sum['C26'] = "=COUNTIF('Income Tracker'!I:I,\"Pending\")"
for row in [25, 26]:
    ws_inc_sum.cell(row=row, column=2).number_format = currency_format

style_data_rows(ws_inc_sum, 25, 26, 3)
auto_width(ws_inc_sum, 4)
ws_inc_sum.column_dimensions['A'].width = 22
ws_inc_sum.column_dimensions['B'].width = 18

# ============ EXPENSE SUMMARY ============
ws_exp_sum = wb.create_sheet("Expense Summary")

ws_exp_sum['A1'] = "EXPENSE SUMMARY REPORT"
ws_exp_sum.merge_cells('A1:D1')
ws_exp_sum['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_exp_sum['A3'] = "CATEGORY-WISE EXPENSES"
ws_exp_sum['A3'].font = Font(bold=True, size=12)
ws_exp_sum.append(["Category", "Total Amount (₹)", "Transaction Count", "Avg Transaction"])
style_header(ws_exp_sum, 4, 4)

for i, cat in enumerate(expense_cats, start=5):
    ws_exp_sum.cell(row=i, column=1).value = cat
    ws_exp_sum.cell(row=i, column=2).value = f'=SUMIF(\'Expense Tracker\'!C:C,A{i},\'Expense Tracker\'!F:F)'
    ws_exp_sum.cell(row=i, column=3).value = f'=COUNTIF(\'Expense Tracker\'!C:C,A{i})'
    ws_exp_sum.cell(row=i, column=4).value = f'=IF(C{i}>0,B{i}/C{i},0)'
    ws_exp_sum.cell(row=i, column=2).number_format = currency_format
    ws_exp_sum.cell(row=i, column=4).number_format = currency_format

ws_exp_sum.cell(row=11, column=1).value = "TOTAL"
ws_exp_sum.cell(row=11, column=1).font = Font(bold=True)
ws_exp_sum.cell(row=11, column=2).value = "=SUM(B5:B10)"
ws_exp_sum.cell(row=11, column=2).font = Font(bold=True)
ws_exp_sum.cell(row=11, column=2).number_format = currency_format
ws_exp_sum.cell(row=11, column=3).value = "=SUM(C5:C10)"
ws_exp_sum.cell(row=11, column=3).font = Font(bold=True)

style_data_rows(ws_exp_sum, 5, 11, 4)

ws_exp_sum['A14'] = "PAYMENT STATUS"
ws_exp_sum['A14'].font = Font(bold=True, size=12)
ws_exp_sum.append(["Status", "Amount (₹)", "Count"])
style_header(ws_exp_sum, 15, 3)
ws_exp_sum['A16'] = "Received"
ws_exp_sum['B16'] = "=SUMIF('Expense Tracker'!H:H,\"Received\",'Expense Tracker'!F:F)"
ws_exp_sum['C16'] = "=COUNTIF('Expense Tracker'!H:H,\"Received\")"
ws_exp_sum['A17'] = "Pending"
ws_exp_sum['B17'] = "=SUMIF('Expense Tracker'!H:H,\"Pending\",'Expense Tracker'!F:F)"
ws_exp_sum['C17'] = "=COUNTIF('Expense Tracker'!H:H,\"Pending\")"
for row in [16, 17]:
    ws_exp_sum.cell(row=row, column=2).number_format = currency_format

style_data_rows(ws_exp_sum, 16, 17, 3)
auto_width(ws_exp_sum, 4)
ws_exp_sum.column_dimensions['A'].width = 18

# ============ PROFITABILITY REPORT ============
ws_profit = wb.create_sheet("Profitability Report")

ws_profit['A1'] = "PROFITABILITY ANALYSIS"
ws_profit.merge_cells('A1:D1')
ws_profit['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_profit['A3'] = "OVERALL PROFITABILITY"
ws_profit['A3'].font = Font(bold=True, size=12)
ws_profit.append(["Metric", "Amount (₹)"])
style_header(ws_profit, 4, 2)

ws_profit['A5'] = "Total Revenue"
ws_profit['B5'] = "=SUM('Income Tracker'!G:G)"
ws_profit['A6'] = "Total Expenses"
ws_profit['B6'] = "=SUM('Expense Tracker'!F:F)"
ws_profit['A7'] = "Gross Profit"
ws_profit['B7'] = "=B5-B6"
ws_profit['A8'] = "Profit Margin %"
ws_profit['B8'] = "=IF(B5>0,B7/B5,0)"
ws_profit.cell(row=8, column=2).number_format = '0.0%'

for row in range(5, 8):
    ws_profit.cell(row=row, column=2).number_format = currency_format
ws_profit['B7'].font = Font(bold=True, color="008000")

style_data_rows(ws_profit, 5, 8, 2)

ws_profit['A11'] = "SERVICE-WISE PROFITABILITY"
ws_profit['A11'].font = Font(bold=True, size=12)
ws_profit.append(["Service", "Revenue (₹)", "% of Total"])
style_header(ws_profit, 12, 3)

for i, svc in enumerate(services, start=13):
    ws_profit.cell(row=i, column=1).value = svc
    ws_profit.cell(row=i, column=2).value = f'=SUMIF(\'Income Tracker\'!D:D,A{i},\'Income Tracker\'!G:G)'
    ws_profit.cell(row=i, column=3).value = f'=IF($B$5>0,B{i}/$B$5,0)'
    ws_profit.cell(row=i, column=2).number_format = currency_format
    ws_profit.cell(row=i, column=3).number_format = '0.0%'

style_data_rows(ws_profit, 13, 20, 3)
auto_width(ws_profit, 3)
ws_profit.column_dimensions['A'].width = 20
ws_profit.column_dimensions['B'].width = 16

# ============ STOCK REPORTS ============
ws_stock_rep = wb.create_sheet("Stock Reports")

ws_stock_rep['A1'] = "STOCK & INVENTORY REPORTS"
ws_stock_rep.merge_cells('A1:E1')
ws_stock_rep['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_stock_rep['A3'] = "LOW STOCK ALERTS"
ws_stock_rep['A3'].font = Font(bold=True, size=12, color="FF0000")
ws_stock_rep.append(["Item Type", "Item ID", "Item Name", "Current Stock", "Min Level"])
style_header(ws_stock_rep, 4, 5)

ws_stock_rep['A5'] = "Chemicals with Low Stock:"
ws_stock_rep['A5'].font = Font(italic=True)
ws_stock_rep.merge_cells('A5:E5')
ws_stock_rep['A6'] = "(Check Chemicals Stock sheet - Status column shows 'Low Stock' items)"
ws_stock_rep['A6'].font = Font(italic=True, color="666666")
ws_stock_rep.merge_cells('A6:E6')

ws_stock_rep['A8'] = "Accessories with Low Stock:"
ws_stock_rep['A8'].font = Font(italic=True)
ws_stock_rep.merge_cells('A8:E8')
ws_stock_rep['A9'] = "(Check Accessories Stock sheet - Status column shows 'Low Stock' items)"
ws_stock_rep['A9'].font = Font(italic=True, color="666666")
ws_stock_rep.merge_cells('A9:E9')

ws_stock_rep['A12'] = "ASSET VALUE SUMMARY"
ws_stock_rep['A12'].font = Font(bold=True, size=12)
ws_stock_rep.append(["Category", "Total Value (₹)", "Active Count", "Under Repair", "Retired"])
style_header(ws_stock_rep, 13, 5)

for i, cat in enumerate(machine_cats, start=14):
    ws_stock_rep.cell(row=i, column=1).value = cat
    ws_stock_rep.cell(row=i, column=2).value = f'=SUMIF(\'Machines & Equipment\'!D:D,A{i},\'Machines & Equipment\'!F:F)'
    ws_stock_rep.cell(row=i, column=3).value = f'=COUNTIFS(\'Machines & Equipment\'!D:D,A{i},\'Machines & Equipment\'!G:G,"Active")'
    ws_stock_rep.cell(row=i, column=4).value = f'=COUNTIFS(\'Machines & Equipment\'!D:D,A{i},\'Machines & Equipment\'!G:G,"Under Repair")'
    ws_stock_rep.cell(row=i, column=5).value = f'=COUNTIFS(\'Machines & Equipment\'!D:D,A{i},\'Machines & Equipment\'!G:G,"Retired")'
    ws_stock_rep.cell(row=i, column=2).number_format = currency_format

ws_stock_rep.cell(row=18, column=1).value = "TOTAL ASSET VALUE"
ws_stock_rep.cell(row=18, column=1).font = Font(bold=True)
ws_stock_rep.cell(row=18, column=2).value = "=SUM('Machines & Equipment'!F:F)"
ws_stock_rep.cell(row=18, column=2).font = Font(bold=True)
ws_stock_rep.cell(row=18, column=2).number_format = currency_format
ws_stock_rep.cell(row=18, column=3).value = "=COUNTIF('Machines & Equipment'!G:G,\"Active\")"
ws_stock_rep.cell(row=18, column=4).value = "=COUNTIF('Machines & Equipment'!G:G,\"Under Repair\")"
ws_stock_rep.cell(row=18, column=5).value = "=COUNTIF('Machines & Equipment'!G:G,\"Retired\")"

style_data_rows(ws_stock_rep, 14, 18, 5)

ws_stock_rep['A21'] = "MAINTENANCE COST SUMMARY"
ws_stock_rep['A21'].font = Font(bold=True, size=12)
ws_stock_rep.append(["Maintenance Type", "Total Cost (₹)", "Count"])
style_header(ws_stock_rep, 22, 3)

for i, mt in enumerate(maint_types, start=23):
    ws_stock_rep.cell(row=i, column=1).value = mt
    ws_stock_rep.cell(row=i, column=2).value = f'=SUMIF(\'Machine Maintenance\'!D:D,A{i},\'Machine Maintenance\'!F:F)'
    ws_stock_rep.cell(row=i, column=3).value = f'=COUNTIF(\'Machine Maintenance\'!D:D,A{i})'
    ws_stock_rep.cell(row=i, column=2).number_format = currency_format

ws_stock_rep.cell(row=26, column=1).value = "TOTAL MAINTENANCE"
ws_stock_rep.cell(row=26, column=1).font = Font(bold=True)
ws_stock_rep.cell(row=26, column=2).value = "=SUM('Machine Maintenance'!F:F)"
ws_stock_rep.cell(row=26, column=2).font = Font(bold=True)
ws_stock_rep.cell(row=26, column=2).number_format = currency_format

style_data_rows(ws_stock_rep, 23, 26, 3)
auto_width(ws_stock_rep, 5)
ws_stock_rep.column_dimensions['C'].width = 18

# ============ DASHBOARD ============
ws_dash = wb.create_sheet("Dashboard")

ws_dash['A1'] = "MIS DASHBOARD"
ws_dash.merge_cells('A1:F1')
ws_dash['A1'].font = Font(bold=True, size=18, color="1F4E79")
ws_dash['A1'].alignment = center_align

dash_header_fill = PatternFill("solid", fgColor="2E75B6")
ws_dash['A3'] = "KEY METRICS"
ws_dash['A3'].font = Font(bold=True, size=12, color="FFFFFF")
ws_dash['A3'].fill = dash_header_fill
ws_dash.merge_cells('A3:F3')

metrics = [
    ["Total Revenue", "=SUM('Income Tracker'!G:G)", "Total Expenses", "=SUM('Expense Tracker'!F:F)"],
    ["Net Profit", "=B4-D4", "Profit Margin", "=IF(B4>0,B5/B4,0)"],
    ["Active Machines", "=COUNTIF('Machines & Equipment'!G:G,\"Active\")", "Machines Under Repair", "=COUNTIF('Machines & Equipment'!G:G,\"Under Repair\")"],
    ["Pending Payments", "=SUMIF('Income Tracker'!I:I,\"Pending\",'Income Tracker'!G:G)", "Pending Expenses", "=SUMIF('Expense Tracker'!H:H,\"Pending\",'Expense Tracker'!F:F)"],
]

for i, row in enumerate(metrics, start=4):
    ws_dash.cell(row=i, column=1).value = row[0]
    ws_dash.cell(row=i, column=2).value = row[1]
    ws_dash.cell(row=i, column=3).value = row[2]
    ws_dash.cell(row=i, column=4).value = row[3]
    ws_dash.cell(row=i, column=1).font = Font(bold=True)
    ws_dash.cell(row=i, column=3).font = Font(bold=True)
    ws_dash.cell(row=i, column=2).number_format = currency_format
    ws_dash.cell(row=i, column=4).number_format = currency_format if i != 5 else '0.0%'

ws_dash['B5'].font = Font(bold=True, color="008000", size=12)

ws_dash['A9'] = "SERVICE PERFORMANCE"
ws_dash['A9'].font = Font(bold=True, size=12, color="FFFFFF")
ws_dash['A9'].fill = dash_header_fill
ws_dash.merge_cells('A9:C9')

ws_dash.append(["Service", "Revenue", "Orders"])
style_header(ws_dash, 10, 3)
for i, svc in enumerate(services, start=11):
    ws_dash.cell(row=i, column=1).value = svc
    ws_dash.cell(row=i, column=2).value = f'=SUMIF(\'Income Tracker\'!D:D,A{i},\'Income Tracker\'!G:G)'
    ws_dash.cell(row=i, column=3).value = f'=COUNTIF(\'Income Tracker\'!D:D,A{i})'
    ws_dash.cell(row=i, column=2).number_format = currency_format

style_data_rows(ws_dash, 11, 18, 3)

ws_dash['E9'] = "EXPENSE BREAKDOWN"
ws_dash['E9'].font = Font(bold=True, size=12, color="FFFFFF")
ws_dash['E9'].fill = dash_header_fill
ws_dash.merge_cells('E9:G9')

ws_dash['E10'] = "Category"
ws_dash['F10'] = "Amount"
ws_dash['G10'] = "Count"
for col in ['E', 'F', 'G']:
    ws_dash[f'{col}10'].fill = header_fill
    ws_dash[f'{col}10'].font = header_font
    ws_dash[f'{col}10'].alignment = center_align

for i, cat in enumerate(expense_cats, start=11):
    ws_dash.cell(row=i, column=5).value = cat
    ws_dash.cell(row=i, column=6).value = f'=SUMIF(\'Expense Tracker\'!C:C,E{i},\'Expense Tracker\'!F:F)'
    ws_dash.cell(row=i, column=7).value = f'=COUNTIF(\'Expense Tracker\'!C:C,E{i})'
    ws_dash.cell(row=i, column=6).number_format = currency_format

ws_dash['A21'] = "STOCK STATUS"
ws_dash['A21'].font = Font(bold=True, size=12, color="FFFFFF")
ws_dash['A21'].fill = dash_header_fill
ws_dash.merge_cells('A21:D21')

ws_dash['A22'] = "Total Asset Value (Machines)"
ws_dash['B22'] = "=SUM('Machines & Equipment'!F:F)"
ws_dash['B22'].number_format = currency_format
ws_dash['A23'] = "Total Maintenance Cost"
ws_dash['B23'] = "=SUM('Machine Maintenance'!F:F)"
ws_dash['B23'].number_format = currency_format
ws_dash['A24'] = "Stock Transactions This Month"
ws_dash['B24'] = "=COUNTA('Stock Transactions'!A:A)-1"

for col in range(1, 8):
    ws_dash.column_dimensions[get_column_letter(col)].width = 18

ws_dash.freeze_panes = 'A3'

# Save the workbook
wb.save('/var/lib/freelancer/projects/40182876/MIS_System_Complete.xlsx')
print("MIS System created successfully!")
