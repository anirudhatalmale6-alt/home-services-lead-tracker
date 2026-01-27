import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, Reference
from copy import copy

wb = Workbook()

# ============================================================
# SHEET 0 — HOW TO USE (Instructions Sheet - First Tab)
# ============================================================
ws0 = wb.active
ws0.title = "HOW TO USE"

# Title
ws0.merge_cells('A1:E1')
ws0['A1'].value = "HOW TO OPEN & USE THIS FILE"
ws0['A1'].font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
ws0['A1'].fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type='solid')
ws0['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws0.row_dimensions[1].height = 50

ws0.merge_cells('A2:E2')
ws0['A2'].value = "Please read these steps carefully before using the sheet"
ws0['A2'].font = Font(name='Calibri', size=11, italic=True, color="FFFFFF")
ws0['A2'].fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type='solid')
ws0['A2'].alignment = Alignment(horizontal='center', vertical='center')

# Step-by-step instructions
steps = [
    "",
    "STEP 1: OPEN THIS FILE IN GOOGLE SHEETS",
    "  - Go to drive.google.com",
    "  - Click '+ New' button (top left) → 'File Upload'",
    "  - Select this .xlsx file from your computer",
    "  - Once uploaded, double-click the file in Google Drive",
    "  - It will open automatically in Google Sheets!",
    "",
    "STEP 2: ENABLE FILTERS (Very Important!)",
    "  - Go to the 'Lead Tracker' tab (bottom of screen)",
    "  - Click anywhere in Row 3 (the header row with blue background)",
    "  - Go to menu: Data → Create a filter",
    "  - Now every column header will show a small dropdown arrow (▼)",
    "  - Click any ▼ arrow to filter by: Confirmed, Completed, Pending, etc.",
    "  - This is how you search/filter your data!",
    "",
    "STEP 3: UNDERSTAND THE COLOR CODES",
    "  - GREEN text on green = CONFIRMED (order confirmed by customer)",
    "  - DARK GREEN with white text = COMPLETED (service done)",
    "  - YELLOW = PENDING (customer will confirm later)",
    "  - RED = CANCELLED (customer not interested)",
    "  - BLUE = SCHEDULED (date fixed, vendor assigned)",
    "  - Each entire row also changes color based on status!",
    "",
    "STEP 4: HOW TO USE THE LEAD TRACKER",
    "  - When Google Form is submitted → data auto-fills in columns B-M",
    "  - Order ID (ST-0001) auto-generates in column C",
    "  - YOU manually update: Order Status, Vendor, Payment details",
    "  - Use the dropdown menus (click any cell to see the dropdown arrow)",
    "",
    "STEP 4B: CALENDAR & TIME PICKERS",
    "  - Date columns (Preferred Date, Scheduled Date, Completed Date) have DATE format",
    "  - In Google Sheets, click on any date cell → a calendar picker appears!",
    "  - Scheduled Time column has a dropdown with time slots (Morning, Afternoon, Evening)",
    "  - This makes scheduling look professional and easy to use",
    "",
    "STEP 5: DASHBOARD TAB",
    "  - Go to 'Dashboard' tab to see all your stats",
    "  - Everything is auto-calculated — no manual work needed!",
    "  - Shows: Total orders, status counts, payment summary",
    "  - Shows: Service-wise & City-wise breakdown",
    "  - Shows: Charts and graphs",
    "",
    "STEP 6: LINK GOOGLE FORM",
    "  - See the 'Form Fields Reference' tab for the complete setup guide",
    "  - Create a Google Form with those fields",
    "  - Link form responses to the 'Lead Tracker' sheet",
    "",
]

for i, step in enumerate(steps):
    row = 4 + i
    ws0.merge_cells(f'A{row}:E{row}')
    cell = ws0[f'A{row}']
    cell.value = step
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    if step and not step.startswith("  "):
        cell.font = Font(name='Calibri', size=12, bold=True, color="1B2A4A")
        cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type='solid')
    else:
        cell.font = Font(name='Calibri', size=11)

ws0.column_dimensions['A'].width = 15
ws0.column_dimensions['B'].width = 15
ws0.column_dimensions['C'].width = 15
ws0.column_dimensions['D'].width = 15
ws0.column_dimensions['E'].width = 15

# ============================================================
# SHEET 1 — LEAD TRACKER
# ============================================================
ws1 = wb.create_sheet("Lead Tracker")

# --- Define colors ---
DARK_BLUE = "1B2A4A"
MEDIUM_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
GREEN_BG = "C6EFCE"
GREEN_FONT = "006100"
YELLOW_BG = "FFEB9C"
YELLOW_FONT = "9C6500"
RED_BG = "FFC7CE"
RED_FONT = "9C0006"
BLUE_BG = "B4C6E7"
BLUE_FONT = "1F4E79"
PURPLE_BG = "E2D0F8"
PURPLE_FONT = "5B2C8E"
COMPLETED_BG = "548235"
COMPLETED_FONT = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
BORDER_COLOR = "B4B4B4"

thin_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR)
)

# --- Column headers ---
headers = [
    "S.No",                    # A
    "Timestamp",               # B
    "Order ID",                # C
    "Customer Name",           # D
    "Phone Number",            # E
    "Alternate Phone",         # F
    "Email",                   # G
    "City",                    # H
    "Full Address",            # I
    "Service Required",        # J
    "Add-on / Changed Service",# K
    "Preferred Date",          # L  (date format - calendar in Google Sheets)
    "Slot Time",               # M
    "Order Status",            # N
    "Reason / Notes",          # O
    "Vendor Name",             # P
    "Vendor Contact",          # Q
    "Vendor Alternate No.",    # R
    "Order Scheduled Date",    # S  (date format - calendar in Google Sheets)
    "Scheduled Time",          # T  (time slot dropdown)
    "Order Completed Date",    # U  (date format - calendar in Google Sheets)
    "Advance Amount",          # V
    "Advance Status",          # W
    "Payment Value (Quoted)",  # X
    "Payment Status",          # Y
    "Payment Mode",            # Z
    "Invoice Number",          # AA
    "Transaction Ref. No.",    # AB
]

# --- Title Row ---
ws1.merge_cells('A1:AB1')
title_cell = ws1['A1']
title_cell.value = "HOME SERVICES — LEAD TRACKER & ORDER MANAGEMENT"
title_cell.font = Font(name='Calibri', size=16, bold=True, color=WHITE)
title_cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 45

# --- Subtitle Row ---
ws1.merge_cells('A2:AB2')
subtitle_cell = ws1['A2']
subtitle_cell.value = "Auto-linked with Google Form  |  Order ID Format: ST-XXXX  |  Use Filters for Quick Search"
subtitle_cell.font = Font(name='Calibri', size=10, italic=True, color=WHITE)
subtitle_cell.fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 25

# --- Header Row (Row 3) ---
header_font = Font(name='Calibri', size=10, bold=True, color=WHITE)
header_fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col_idx, header in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws1.row_dimensions[3].height = 40

# --- Column widths ---
col_widths = {
    'A': 6,   # S.No
    'B': 18,  # Timestamp
    'C': 12,  # Order ID
    'D': 20,  # Customer Name
    'E': 16,  # Phone
    'F': 16,  # Alt Phone
    'G': 22,  # Email
    'H': 14,  # City
    'I': 30,  # Full Address
    'J': 22,  # Service Required
    'K': 22,  # Add-on Service
    'L': 16,  # Preferred Date (calendar)
    'M': 22,  # Slot Time
    'N': 16,  # Order Status
    'O': 25,  # Reason/Notes
    'P': 18,  # Vendor Name
    'Q': 16,  # Vendor Contact
    'R': 18,  # Vendor Alt No
    'S': 18,  # Scheduled Date (calendar)
    'T': 22,  # Scheduled Time
    'U': 18,  # Completed Date (calendar)
    'V': 16,  # Advance Amount
    'W': 16,  # Advance Status
    'X': 18,  # Payment Value
    'Y': 16,  # Payment Status
    'Z': 16,  # Payment Mode
    'AA': 16, # Invoice Number
    'AB': 20, # Transaction Ref
}
for col_letter, width in col_widths.items():
    ws1.column_dimensions[col_letter].width = width

# --- Data Validation (Dropdowns) ---
# Service Required
services = '"Bathroom Cleaning,Kitchen Cleaning,Full Home Cleaning,Rental Property Cleaning,Ready to Move In Cleaning,Painting,Pest Control,Plumbing,Electrician,Other"'
dv_service = DataValidation(type="list", formula1=services, allow_blank=True)
dv_service.error = "Please select a valid service"
dv_service.errorTitle = "Invalid Service"
dv_service.prompt = "Select the service"
dv_service.promptTitle = "Service Required"
ws1.add_data_validation(dv_service)
dv_service.add('J4:J1000')

# Add-on Service (same list)
dv_addon = DataValidation(type="list", formula1=services, allow_blank=True)
dv_addon.prompt = "Select add-on or changed service (if any)"
ws1.add_data_validation(dv_addon)
dv_addon.add('K4:K1000')

# Order Status
statuses = '"Confirmed,Pending,Cancelled,Scheduled,Completed"'
dv_status = DataValidation(type="list", formula1=statuses, allow_blank=True)
dv_status.error = "Please select a valid status"
dv_status.errorTitle = "Invalid Status"
dv_status.prompt = "Select order status"
dv_status.promptTitle = "Order Status"
ws1.add_data_validation(dv_status)
dv_status.add('N4:N1000')

# Advance Status (Column W)
adv_statuses = '"Received,NIL,Cleared"'
dv_advance = DataValidation(type="list", formula1=adv_statuses, allow_blank=True)
dv_advance.prompt = "Select advance status"
ws1.add_data_validation(dv_advance)
dv_advance.add('W4:W1000')

# Payment Status (Column Y)
pay_statuses = '"Received,Pending"'
dv_payment = DataValidation(type="list", formula1=pay_statuses, allow_blank=True)
dv_payment.prompt = "Select payment status"
ws1.add_data_validation(dv_payment)
dv_payment.add('Y4:Y1000')

# Payment Mode (Column Z)
pay_modes = '"Cash,UPI,Debit Card,Payment Gateway,Bank Transfer,Other"'
dv_paymode = DataValidation(type="list", formula1=pay_modes, allow_blank=True)
dv_paymode.prompt = "Select payment mode"
ws1.add_data_validation(dv_paymode)
dv_paymode.add('Z4:Z1000')

# City
cities = '"Bangalore,Mumbai,Delhi,Hyderabad,Chennai,Pune,Other"'
dv_city = DataValidation(type="list", formula1=cities, allow_blank=True)
dv_city.prompt = "Select city"
ws1.add_data_validation(dv_city)
dv_city.add('H4:H1000')

# Slot Time (Column M - customer preferred)
slots = '"Morning (8AM-10AM),Morning (10AM-12PM),Afternoon (12PM-2PM),Afternoon (2PM-4PM),Evening (4PM-6PM),Evening (6PM-8PM)"'
dv_slot = DataValidation(type="list", formula1=slots, allow_blank=True)
dv_slot.prompt = "Select time slot"
ws1.add_data_validation(dv_slot)
dv_slot.add('M4:M1000')

# Scheduled Time (Column T - actual scheduled time)
dv_sched_time = DataValidation(type="list", formula1=slots, allow_blank=True)
dv_sched_time.prompt = "Select scheduled time slot"
dv_sched_time.promptTitle = "Scheduled Time"
ws1.add_data_validation(dv_sched_time)
dv_sched_time.add('T4:T1000')

# Date validation for date columns (L, S, U) — format as date
# In Google Sheets, date-formatted cells automatically show calendar picker
dv_date_pref = DataValidation(type="date", allow_blank=True)
dv_date_pref.prompt = "Enter date (click for calendar in Google Sheets)"
dv_date_pref.promptTitle = "Select Date"
ws1.add_data_validation(dv_date_pref)
dv_date_pref.add('L4:L1000')

dv_date_sched = DataValidation(type="date", allow_blank=True)
dv_date_sched.prompt = "Enter scheduled date (click for calendar in Google Sheets)"
dv_date_sched.promptTitle = "Scheduled Date"
ws1.add_data_validation(dv_date_sched)
dv_date_sched.add('S4:S1000')

dv_date_comp = DataValidation(type="date", allow_blank=True)
dv_date_comp.prompt = "Enter completion date (click for calendar in Google Sheets)"
dv_date_comp.promptTitle = "Completed Date"
ws1.add_data_validation(dv_date_comp)
dv_date_comp.add('U4:U1000')

# --- Conditional Formatting for Order Status (Column N) ---
# Green for Confirmed
ws1.conditional_formatting.add('N4:N1000',
    CellIsRule(operator='equal', formula=['"Confirmed"'],
              fill=PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type='solid'),
              font=Font(color=GREEN_FONT, bold=True)))

# Yellow for Pending
ws1.conditional_formatting.add('N4:N1000',
    CellIsRule(operator='equal', formula=['"Pending"'],
              fill=PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type='solid'),
              font=Font(color=YELLOW_FONT, bold=True)))

# Red for Cancelled
ws1.conditional_formatting.add('N4:N1000',
    CellIsRule(operator='equal', formula=['"Cancelled"'],
              fill=PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type='solid'),
              font=Font(color=RED_FONT, bold=True)))

# Blue for Scheduled
ws1.conditional_formatting.add('N4:N1000',
    CellIsRule(operator='equal', formula=['"Scheduled"'],
              fill=PatternFill(start_color=BLUE_BG, end_color=BLUE_BG, fill_type='solid'),
              font=Font(color=BLUE_FONT, bold=True)))

# Dark Green/Purple for Completed (distinctly different from Confirmed)
ws1.conditional_formatting.add('N4:N1000',
    CellIsRule(operator='equal', formula=['"Completed"'],
              fill=PatternFill(start_color=COMPLETED_BG, end_color=COMPLETED_BG, fill_type='solid'),
              font=Font(color=COMPLETED_FONT, bold=True)))

# --- Conditional Formatting for Payment Status (Column X) ---
ws1.conditional_formatting.add('Y4:Y1000',
    CellIsRule(operator='equal', formula=['"Received"'],
              fill=PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type='solid'),
              font=Font(color=GREEN_FONT, bold=True)))

ws1.conditional_formatting.add('Y4:Y1000',
    CellIsRule(operator='equal', formula=['"Pending"'],
              fill=PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type='solid'),
              font=Font(color=RED_FONT, bold=True)))

# --- Conditional Formatting for Advance Status (Column V) ---
ws1.conditional_formatting.add('W4:W1000',
    CellIsRule(operator='equal', formula=['"Received"'],
              fill=PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type='solid'),
              font=Font(color=GREEN_FONT, bold=True)))

ws1.conditional_formatting.add('W4:W1000',
    CellIsRule(operator='equal', formula=['"NIL"'],
              fill=PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type='solid'),
              font=Font(color=YELLOW_FONT, bold=True)))

ws1.conditional_formatting.add('W4:W1000',
    CellIsRule(operator='equal', formula=['"Cleared"'],
              fill=PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type='solid'),
              font=Font(color="375623", bold=True)))

# --- Row-level conditional formatting (color entire row based on status) ---
# Light green row for Confirmed
ws1.conditional_formatting.add('A4:AB1000',
    FormulaRule(formula=['$N4="Confirmed"'],
               fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type='solid')))

# Light yellow row for Pending
ws1.conditional_formatting.add('A4:AB1000',
    FormulaRule(formula=['$N4="Pending"'],
               fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type='solid')))

# Light red row for Cancelled
ws1.conditional_formatting.add('A4:AB1000',
    FormulaRule(formula=['$N4="Cancelled"'],
               fill=PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type='solid')))

# Light blue row for Scheduled
ws1.conditional_formatting.add('A4:AB1000',
    FormulaRule(formula=['$N4="Scheduled"'],
               fill=PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type='solid')))

# Light purple row for Completed (distinct from Confirmed)
ws1.conditional_formatting.add('A4:AB1000',
    FormulaRule(formula=['$N4="Completed"'],
               fill=PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type='solid')))

# --- Auto-generate formulas for rows 4-1000 ---
data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_font = Font(name='Calibri', size=10)

for row in range(4, 101):  # Pre-format rows 4-100
    # S.No formula
    ws1.cell(row=row, column=1).value = f'=IF(B{row}<>"",ROW()-3,"")'
    ws1.cell(row=row, column=1).alignment = data_align
    ws1.cell(row=row, column=1).font = data_font
    ws1.cell(row=row, column=1).border = thin_border

    # Order ID formula: ST-0001 format
    ws1.cell(row=row, column=3).value = f'=IF(B{row}<>"","ST-"&TEXT(ROW()-3,"0000"),"")'
    ws1.cell(row=row, column=3).alignment = data_align
    ws1.cell(row=row, column=3).font = Font(name='Calibri', size=10, bold=True, color=MEDIUM_BLUE)
    ws1.cell(row=row, column=3).border = thin_border

    # Format other cells (28 columns now: A=1 to AB=28)
    for col in range(2, 29):
        cell = ws1.cell(row=row, column=col)
        if cell.value is None:
            cell.value = None
        cell.alignment = data_align
        cell.font = data_font
        cell.border = thin_border

    # Date format for: Timestamp (B=2), Preferred Date (L=12), Scheduled Date (S=19), Completed Date (U=21)
    for date_col in [2, 12, 19, 21]:
        ws1.cell(row=row, column=date_col).number_format = 'DD-MMM-YYYY'
    # Timestamp with time
    ws1.cell(row=row, column=2).number_format = 'DD-MMM-YYYY HH:MM'

    # Currency format for Advance Amount (V=22) and Payment Value (X=24)
    ws1.cell(row=row, column=22).number_format = '₹#,##0.00'
    ws1.cell(row=row, column=24).number_format = '₹#,##0.00'

# --- Freeze panes ---
ws1.freeze_panes = 'D4'

# --- Auto filter ---
ws1.auto_filter.ref = 'A3:AB1000'

# --- Add sample data for demonstration ---
# Sample data: columns A-AB (28 cols)
# A=S.No, B=Timestamp, C=OrderID, D=Name, E=Phone, F=AltPhone, G=Email,
# H=City, I=Address, J=Service, K=Addon, L=PrefDate, M=SlotTime,
# N=Status, O=Reason, P=VendorName, Q=VendorContact, R=VendorAlt,
# S=SchedDate, T=SchedTime, U=CompDate, V=AdvAmt, W=AdvStatus,
# X=PayValue, Y=PayStatus, Z=PayMode, AA=InvoiceNo, AB=TxnRef
sample_data = [
    [None, "15-Jan-2026 10:30", None, "Rajesh Kumar", "9876543210", "9876543211", "rajesh@email.com",
     "Bangalore", "123, MG Road, Indiranagar, Bangalore - 560038", "Bathroom Cleaning", "",
     "18-Jan-2026", "Morning (10AM-12PM)", "Confirmed", "Customer confirmed on call",
     "Suresh", "9988776655", "9988776600", "18-Jan-2026", "Morning (10AM-12PM)", "",
     500, "Received", 2000, "Pending", "", "INV-001", ""],
    [None, "15-Jan-2026 11:15", None, "Priya Sharma", "8765432109", "", "priya.s@email.com",
     "Bangalore", "45, HSR Layout, Sector 2, Bangalore - 560102", "Pest Control", "",
     "20-Jan-2026", "Afternoon (2PM-4PM)", "Pending", "Will confirm by evening",
     "", "", "", "", "", "",
     0, "NIL", 3500, "Pending", "", "", ""],
    [None, "16-Jan-2026 09:00", None, "Amit Patel", "7654321098", "7654321099", "amit.p@email.com",
     "Bangalore", "78, Whitefield Main Rd, Bangalore - 560066", "Painting", "Full Home Cleaning",
     "25-Jan-2026", "Morning (8AM-10AM)", "Cancelled", "Price too high, found cheaper option",
     "", "", "", "", "", "",
     0, "NIL", 15000, "Pending", "", "", ""],
    [None, "16-Jan-2026 14:20", None, "Meera Reddy", "6543210987", "", "meera.r@email.com",
     "Bangalore", "22, Koramangala 5th Block, Bangalore - 560095", "Full Home Cleaning", "",
     "22-Jan-2026", "Morning (10AM-12PM)", "Scheduled", "Scheduled with vendor Ramesh",
     "Ramesh", "9090909090", "9090909091", "22-Jan-2026", "Morning (10AM-12PM)", "",
     1000, "Received", 4500, "Pending", "", "", ""],
    [None, "17-Jan-2026 08:45", None, "Karthik Nair", "5432109876", "5432109800", "karthik@email.com",
     "Bangalore", "99, JP Nagar 6th Phase, Bangalore - 560078", "Plumbing", "",
     "17-Jan-2026", "Evening (4PM-6PM)", "Completed", "Service completed successfully",
     "Vijay", "8080808080", "8080808081", "17-Jan-2026", "Evening (4PM-6PM)", "17-Jan-2026",
     0, "NIL", 1500, "Received", "Cash", "INV-002", ""],
]

for i, data in enumerate(sample_data):
    row = 4 + i
    for col_idx, value in enumerate(data, 1):
        if value is not None and value != "":
            cell = ws1.cell(row=row, column=col_idx, value=value)

# ============================================================
# SHEET 2 — DASHBOARD
# ============================================================
ws2 = wb.create_sheet("Dashboard")

# --- Title ---
ws2.merge_cells('A1:L1')
ws2['A1'].value = "DASHBOARD — BUSINESS ANALYTICS & STATS"
ws2['A1'].font = Font(name='Calibri', size=16, bold=True, color=WHITE)
ws2['A1'].fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 45

ws2.merge_cells('A2:L2')
ws2['A2'].value = "All data auto-calculated from Lead Tracker  |  Use Date Filters for Custom Reports"
ws2['A2'].font = Font(name='Calibri', size=10, italic=True, color=WHITE)
ws2['A2'].fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
ws2['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[2].height = 25

# --- COLOR LEGEND ---
ws2.merge_cells('A4:E4')
ws2['A4'].value = "STATUS COLOR LEGEND"
ws2['A4'].font = Font(name='Calibri', size=11, bold=True, color=WHITE)
ws2['A4'].fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
ws2['A4'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[4].height = 25

legend_items = [
    ("Confirmed", GREEN_BG, GREEN_FONT),
    ("Pending", YELLOW_BG, YELLOW_FONT),
    ("Cancelled", RED_BG, RED_FONT),
    ("Scheduled", BLUE_BG, BLUE_FONT),
    ("Completed", COMPLETED_BG, COMPLETED_FONT),
]
for i, (label, bg, fc) in enumerate(legend_items):
    cell = ws2.cell(row=5, column=i+1, value=label)
    cell.font = Font(name='Calibri', size=10, bold=True, color=fc)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
ws2.row_dimensions[5].height = 30

# --- Section 1: ORDER STATISTICS ---
R_OS = 7  # Row for Order Statistics section header
ws2.merge_cells(f'A{R_OS}:F{R_OS}')
ws2[f'A{R_OS}'].value = "ORDER STATISTICS"
ws2[f'A{R_OS}'].font = Font(name='Calibri', size=13, bold=True, color=WHITE)
ws2[f'A{R_OS}'].fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
ws2[f'A{R_OS}'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[R_OS].height = 30

# Stats headers
stat_headers = ["Total Orders", "Confirmed", "Pending", "Cancelled", "Scheduled", "Completed"]
stat_formulas = [
    '=COUNTA(\'Lead Tracker\'!C4:C1000)-COUNTBLANK(\'Lead Tracker\'!C4:C1000)',
    '=COUNTIF(\'Lead Tracker\'!N4:N1000,"Confirmed")',
    '=COUNTIF(\'Lead Tracker\'!N4:N1000,"Pending")',
    '=COUNTIF(\'Lead Tracker\'!N4:N1000,"Cancelled")',
    '=COUNTIF(\'Lead Tracker\'!N4:N1000,"Scheduled")',
    '=COUNTIF(\'Lead Tracker\'!N4:N1000,"Completed")',
]
stat_colors = ["1B2A4A", "006100", "9C6500", "9C0006", "1F4E79", COMPLETED_FONT]
stat_bg_colors = ["D6E4F0", GREEN_BG, YELLOW_BG, RED_BG, BLUE_BG, COMPLETED_BG]

for i, (header, formula, font_color, bg_color) in enumerate(zip(stat_headers, stat_formulas, stat_colors, stat_bg_colors)):
    col = i + 1
    cell_h = ws2.cell(row=R_OS+1, column=col, value=header)
    cell_h.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    cell_h.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    cell_h.alignment = Alignment(horizontal='center', vertical='center')
    cell_h.border = thin_border

    cell_v = ws2.cell(row=R_OS+2, column=col, value=formula)
    cell_v.font = Font(name='Calibri', size=22, bold=True, color=font_color)
    cell_v.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    cell_v.alignment = Alignment(horizontal='center', vertical='center')
    cell_v.border = thin_border

ws2.row_dimensions[R_OS+1].height = 28
ws2.row_dimensions[R_OS+2].height = 50

# --- Section 2: PAYMENT SUMMARY ---
R_PS = R_OS + 4  # Row 11
ws2.merge_cells(f'A{R_PS}:F{R_PS}')
ws2[f'A{R_PS}'].value = "PAYMENT SUMMARY"
ws2[f'A{R_PS}'].font = Font(name='Calibri', size=13, bold=True, color=WHITE)
ws2[f'A{R_PS}'].fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
ws2[f'A{R_PS}'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[R_PS].height = 30

pay_headers = ["Total Quoted", "Total Received", "Total Pending", "Cash Received", "UPI Received", "Card/Gateway"]
pay_formulas = [
    '=SUMPRODUCT((\'Lead Tracker\'!X4:X1000)*1)',
    '=SUMPRODUCT((\'Lead Tracker\'!Y4:Y1000="Received")*(\'Lead Tracker\'!X4:X1000))',
    '=SUMPRODUCT((\'Lead Tracker\'!Y4:Y1000="Pending")*(\'Lead Tracker\'!X4:X1000))',
    '=SUMPRODUCT((\'Lead Tracker\'!Y4:Y1000="Received")*(\'Lead Tracker\'!Z4:Z1000="Cash")*(\'Lead Tracker\'!X4:X1000))',
    '=SUMPRODUCT((\'Lead Tracker\'!Y4:Y1000="Received")*(\'Lead Tracker\'!Z4:Z1000="UPI")*(\'Lead Tracker\'!X4:X1000))',
    '=SUMPRODUCT((\'Lead Tracker\'!Y4:Y1000="Received")*((\'Lead Tracker\'!Z4:Z1000="Debit Card")+(\'Lead Tracker\'!Z4:Z1000="Payment Gateway"))*(\'Lead Tracker\'!X4:X1000))',
]

for i, (header, formula) in enumerate(zip(pay_headers, pay_formulas)):
    col = i + 1
    cell_h = ws2.cell(row=R_PS+1, column=col, value=header)
    cell_h.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    cell_h.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    cell_h.alignment = Alignment(horizontal='center', vertical='center')
    cell_h.border = thin_border

    cell_v = ws2.cell(row=R_PS+2, column=col, value=formula)
    cell_v.font = Font(name='Calibri', size=16, bold=True, color=DARK_BLUE)
    cell_v.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
    cell_v.alignment = Alignment(horizontal='center', vertical='center')
    cell_v.number_format = '₹#,##0'
    cell_v.border = thin_border

ws2.row_dimensions[R_PS+1].height = 28
ws2.row_dimensions[R_PS+2].height = 50

# --- Section 3: SERVICE-WISE BREAKDOWN ---
R_SB = R_PS + 4  # Row 15
ws2.merge_cells(f'A{R_SB}:H{R_SB}')
ws2[f'A{R_SB}'].value = "SERVICE-WISE BREAKDOWN"
ws2[f'A{R_SB}'].font = Font(name='Calibri', size=13, bold=True, color=WHITE)
ws2[f'A{R_SB}'].fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
ws2[f'A{R_SB}'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[R_SB].height = 30

service_list = [
    "Bathroom Cleaning", "Kitchen Cleaning", "Full Home Cleaning",
    "Rental Property Cleaning", "Ready to Move In Cleaning",
    "Painting", "Pest Control", "Plumbing", "Electrician", "Other"
]

# Headers for service table (Confirmed & Completed as SEPARATE columns)
R_SB_HDR = R_SB + 1  # header row for service table
R_SB_DATA = R_SB + 2  # first data row

svc_table_headers = ["Service", "Total Orders", "Confirmed", "Scheduled", "Completed", "Pending", "Cancelled", "Revenue"]
for i, h in enumerate(svc_table_headers):
    cell = ws2.cell(row=R_SB_HDR, column=i+1, value=h)
    cell.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

for idx, svc in enumerate(service_list):
    row = R_SB_DATA + idx
    alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid') if idx % 2 == 0 else None

    # Service name
    cell_s = ws2.cell(row=row, column=1, value=svc)
    cell_s.font = Font(name='Calibri', size=10)
    cell_s.alignment = Alignment(horizontal='left', vertical='center')
    cell_s.border = thin_border

    # Total orders
    ws2.cell(row=row, column=2, value=f'=COUNTIF(\'Lead Tracker\'!J4:J1000,A{row})').border = thin_border
    ws2.cell(row=row, column=2).font = Font(name='Calibri', size=10, bold=True)
    ws2.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')

    # Confirmed only
    ws2.cell(row=row, column=3, value=f'=COUNTIFS(\'Lead Tracker\'!J4:J1000,A{row},\'Lead Tracker\'!N4:N1000,"Confirmed")').border = thin_border
    ws2.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')

    # Scheduled only
    ws2.cell(row=row, column=4, value=f'=COUNTIFS(\'Lead Tracker\'!J4:J1000,A{row},\'Lead Tracker\'!N4:N1000,"Scheduled")').border = thin_border
    ws2.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')

    # Completed only
    ws2.cell(row=row, column=5, value=f'=COUNTIFS(\'Lead Tracker\'!J4:J1000,A{row},\'Lead Tracker\'!N4:N1000,"Completed")').border = thin_border
    ws2.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center')

    # Pending only
    ws2.cell(row=row, column=6, value=f'=COUNTIFS(\'Lead Tracker\'!J4:J1000,A{row},\'Lead Tracker\'!N4:N1000,"Pending")').border = thin_border
    ws2.cell(row=row, column=6).alignment = Alignment(horizontal='center', vertical='center')

    # Cancelled only
    ws2.cell(row=row, column=7, value=f'=COUNTIFS(\'Lead Tracker\'!J4:J1000,A{row},\'Lead Tracker\'!N4:N1000,"Cancelled")').border = thin_border
    ws2.cell(row=row, column=7).alignment = Alignment(horizontal='center', vertical='center')

    # Revenue
    ws2.cell(row=row, column=8, value=f'=SUMPRODUCT((\'Lead Tracker\'!J4:J1000=A{row})*(\'Lead Tracker\'!Y4:Y1000="Received")*(\'Lead Tracker\'!X4:X1000))').border = thin_border
    ws2.cell(row=row, column=8).font = Font(name='Calibri', size=10, bold=True)
    ws2.cell(row=row, column=8).alignment = Alignment(horizontal='center', vertical='center')
    ws2.cell(row=row, column=8).number_format = '₹#,##0'

    # Apply alternating row color
    if alt_fill:
        for c in range(1, 9):
            ws2.cell(row=row, column=c).fill = alt_fill

# Total row for services
total_row = R_SB_DATA + len(service_list)
for c in range(1, 9):
    cell = ws2.cell(row=total_row, column=c)
    cell.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

ws2.cell(row=total_row, column=1, value="TOTAL")
ws2.cell(row=total_row, column=2, value=f'=SUM(B{R_SB_DATA}:B{total_row-1})')
ws2.cell(row=total_row, column=3, value=f'=SUM(C{R_SB_DATA}:C{total_row-1})')
ws2.cell(row=total_row, column=4, value=f'=SUM(D{R_SB_DATA}:D{total_row-1})')
ws2.cell(row=total_row, column=5, value=f'=SUM(E{R_SB_DATA}:E{total_row-1})')
ws2.cell(row=total_row, column=6, value=f'=SUM(F{R_SB_DATA}:F{total_row-1})')
ws2.cell(row=total_row, column=7, value=f'=SUM(G{R_SB_DATA}:G{total_row-1})')
ws2.cell(row=total_row, column=8, value=f'=SUM(H{R_SB_DATA}:H{total_row-1})')
ws2.cell(row=total_row, column=8).number_format = '₹#,##0'

# --- Section 4: AREA-WISE BREAKDOWN ---
area_start = total_row + 2
ws2.merge_cells(f'A{area_start}:F{area_start}')
ws2[f'A{area_start}'].value = "AREA-WISE ORDER DISTRIBUTION"
ws2[f'A{area_start}'].font = Font(name='Calibri', size=13, bold=True, color=WHITE)
ws2[f'A{area_start}'].fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
ws2[f'A{area_start}'].alignment = Alignment(horizontal='center', vertical='center')

area_headers = ["City", "Total Orders", "Confirmed", "Completed", "Pending", "Revenue"]
for i, h in enumerate(area_headers):
    r = area_start + 1
    cell = ws2.cell(row=r, column=i+1, value=h)
    cell.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

cities_list = ["Bangalore", "Mumbai", "Delhi", "Hyderabad", "Chennai", "Pune", "Other"]
for idx, city in enumerate(cities_list):
    row = area_start + 2 + idx
    ws2.cell(row=row, column=1, value=city).border = thin_border
    ws2.cell(row=row, column=1).font = Font(name='Calibri', size=10)
    ws2.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')

    ws2.cell(row=row, column=2, value=f'=COUNTIF(\'Lead Tracker\'!H4:H1000,A{row})').border = thin_border
    ws2.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')

    # Confirmed only
    ws2.cell(row=row, column=3, value=f'=COUNTIFS(\'Lead Tracker\'!H4:H1000,A{row},\'Lead Tracker\'!N4:N1000,"Confirmed")').border = thin_border
    ws2.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')

    # Completed only
    ws2.cell(row=row, column=4, value=f'=COUNTIFS(\'Lead Tracker\'!H4:H1000,A{row},\'Lead Tracker\'!N4:N1000,"Completed")').border = thin_border
    ws2.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')

    # Pending only
    ws2.cell(row=row, column=5, value=f'=COUNTIFS(\'Lead Tracker\'!H4:H1000,A{row},\'Lead Tracker\'!N4:N1000,"Pending")').border = thin_border
    ws2.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center')

    ws2.cell(row=row, column=6, value=f'=SUMPRODUCT((\'Lead Tracker\'!H4:H1000=A{row})*(\'Lead Tracker\'!Y4:Y1000="Received")*(\'Lead Tracker\'!X4:X1000))').border = thin_border
    ws2.cell(row=row, column=6).number_format = '₹#,##0'
    ws2.cell(row=row, column=6).alignment = Alignment(horizontal='center', vertical='center')

    if idx % 2 == 0:
        for c in range(1, 7):
            ws2.cell(row=row, column=c).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

# --- Section 5: DATE-WISE FILTER HELPER ---
date_start = area_start + 2 + len(cities_list) + 2
ws2.merge_cells(f'A{date_start}:F{date_start}')
ws2[f'A{date_start}'].value = "DATE-WISE REPORT (Enter dates below to filter)"
ws2[f'A{date_start}'].font = Font(name='Calibri', size=13, bold=True, color=WHITE)
ws2[f'A{date_start}'].fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
ws2[f'A{date_start}'].alignment = Alignment(horizontal='center', vertical='center')

ws2.cell(row=date_start+1, column=1, value="From Date:").font = Font(name='Calibri', size=12, bold=True)
ws2.cell(row=date_start+1, column=1).alignment = Alignment(horizontal='right', vertical='center')
ws2.cell(row=date_start+1, column=2).border = Border(
    left=Side(style='medium', color='2E5090'), right=Side(style='medium', color='2E5090'),
    top=Side(style='medium', color='2E5090'), bottom=Side(style='medium', color='2E5090'))
ws2.cell(row=date_start+1, column=2).number_format = 'DD-MMM-YYYY'
ws2.cell(row=date_start+1, column=2).fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type='solid')
ws2.cell(row=date_start+1, column=2).font = Font(name='Calibri', size=12, bold=True)
ws2.cell(row=date_start+1, column=2).alignment = Alignment(horizontal='center', vertical='center')

ws2.cell(row=date_start+1, column=3, value="To Date:").font = Font(name='Calibri', size=12, bold=True)
ws2.cell(row=date_start+1, column=3).alignment = Alignment(horizontal='right', vertical='center')
ws2.cell(row=date_start+1, column=4).border = Border(
    left=Side(style='medium', color='2E5090'), right=Side(style='medium', color='2E5090'),
    top=Side(style='medium', color='2E5090'), bottom=Side(style='medium', color='2E5090'))
ws2.cell(row=date_start+1, column=4).number_format = 'DD-MMM-YYYY'
ws2.cell(row=date_start+1, column=4).fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type='solid')
ws2.cell(row=date_start+1, column=4).font = Font(name='Calibri', size=12, bold=True)
ws2.cell(row=date_start+1, column=4).alignment = Alignment(horizontal='center', vertical='center')

ws2.row_dimensions[date_start+1].height = 35

# Add date validation to From/To date cells so Google Sheets shows calendar picker
dv_dash_from = DataValidation(type="date", allow_blank=True)
dv_dash_from.prompt = "Click to open calendar and select FROM date"
dv_dash_from.promptTitle = "Select From Date"
dv_dash_from.showInputMessage = True
ws2.add_data_validation(dv_dash_from)
dv_dash_from.add(ws2.cell(row=date_start+1, column=2))

dv_dash_to = DataValidation(type="date", allow_blank=True)
dv_dash_to.prompt = "Click to open calendar and select TO date"
dv_dash_to.promptTitle = "Select To Date"
dv_dash_to.showInputMessage = True
ws2.add_data_validation(dv_dash_to)
dv_dash_to.add(ws2.cell(row=date_start+1, column=4))

# Add helper text below the date fields
ws2.merge_cells(f'A{date_start+1}:A{date_start+1}')
ws2.cell(row=date_start+1, column=5, value="← Click the yellow cells, a calendar will appear in Google Sheets")
ws2.cell(row=date_start+1, column=5).font = Font(name='Calibri', size=9, italic=True, color="666666")
ws2.cell(row=date_start+1, column=5).alignment = Alignment(horizontal='left', vertical='center')

dr = date_start + 2
date_report_headers = ["Orders in Range", "Confirmed", "Completed", "Pending", "Cancelled", "Revenue in Range", "Payments Pending"]
for i, h in enumerate(date_report_headers):
    cell = ws2.cell(row=dr, column=i+1, value=h)
    cell.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

fdr = date_start + 1  # row with from/to dates
dr_val = dr + 1
# Date range formulas
ws2.cell(row=dr_val, column=1, value=f'=IF(AND(B{fdr}<>"",D{fdr}<>""),COUNTIFS(\'Lead Tracker\'!B4:B1000,">="&B{fdr},\'Lead Tracker\'!B4:B1000,"<="&D{fdr}),0)')
ws2.cell(row=dr_val, column=2, value=f'=IF(AND(B{fdr}<>"",D{fdr}<>""),COUNTIFS(\'Lead Tracker\'!B4:B1000,">="&B{fdr},\'Lead Tracker\'!B4:B1000,"<="&D{fdr},\'Lead Tracker\'!N4:N1000,"Confirmed"),0)')
ws2.cell(row=dr_val, column=3, value=f'=IF(AND(B{fdr}<>"",D{fdr}<>""),COUNTIFS(\'Lead Tracker\'!B4:B1000,">="&B{fdr},\'Lead Tracker\'!B4:B1000,"<="&D{fdr},\'Lead Tracker\'!N4:N1000,"Completed"),0)')
ws2.cell(row=dr_val, column=4, value=f'=IF(AND(B{fdr}<>"",D{fdr}<>""),COUNTIFS(\'Lead Tracker\'!B4:B1000,">="&B{fdr},\'Lead Tracker\'!B4:B1000,"<="&D{fdr},\'Lead Tracker\'!N4:N1000,"Pending"),0)')
ws2.cell(row=dr_val, column=5, value=f'=IF(AND(B{fdr}<>"",D{fdr}<>""),COUNTIFS(\'Lead Tracker\'!B4:B1000,">="&B{fdr},\'Lead Tracker\'!B4:B1000,"<="&D{fdr},\'Lead Tracker\'!N4:N1000,"Cancelled"),0)')
ws2.cell(row=dr_val, column=6, value=f'=IF(AND(B{fdr}<>"",D{fdr}<>""),SUMPRODUCT((\'Lead Tracker\'!B4:B1000>=B{fdr})*(\'Lead Tracker\'!B4:B1000<=D{fdr})*(\'Lead Tracker\'!Y4:Y1000="Received")*(\'Lead Tracker\'!X4:X1000)),0)')
ws2.cell(row=dr_val, column=7, value=f'=IF(AND(B{fdr}<>"",D{fdr}<>""),SUMPRODUCT((\'Lead Tracker\'!B4:B1000>=B{fdr})*(\'Lead Tracker\'!B4:B1000<=D{fdr})*(\'Lead Tracker\'!Y4:Y1000="Pending")*(\'Lead Tracker\'!X4:X1000)),0)')

for c in range(1, 8):
    cell = ws2.cell(row=dr_val, column=c)
    cell.font = Font(name='Calibri', size=16, bold=True, color=DARK_BLUE)
    cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    if c >= 6:
        cell.number_format = '₹#,##0'

ws2.row_dimensions[dr_val].height = 45

# --- Charts ---
# Service-wise Bar Chart
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Service-Wise Orders"
chart1.style = 10
chart1.y_axis.title = "Number of Orders"
chart1.x_axis.title = "Service"
data1 = Reference(ws2, min_col=2, min_row=R_SB_HDR, max_row=total_row-1, max_col=2)
cats1 = Reference(ws2, min_col=1, min_row=R_SB_DATA, max_row=total_row-1)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.shape = 4
chart1.width = 20
chart1.height = 12

# Area-wise Pie Chart
chart2 = PieChart()
chart2.title = "Area-Wise Order Distribution"
chart2.style = 10
data2 = Reference(ws2, min_col=2, min_row=area_start+1, max_row=area_start+1+len(cities_list), max_col=2)
cats2 = Reference(ws2, min_col=1, min_row=area_start+2, max_row=area_start+1+len(cities_list))
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width = 16
chart2.height = 12

# Place charts to the right (shifted to accommodate wider tables)
ws2.add_chart(chart1, f"J{R_SB}")
ws2.add_chart(chart2, f"J{area_start}")

# --- Column widths for Dashboard ---
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws2.column_dimensions[col].width = 18

# --- Freeze ---
ws2.freeze_panes = 'A4'

# ============================================================
# SHEET 3 — Google Form Fields Reference
# ============================================================
ws3 = wb.create_sheet("Form Fields Reference")

ws3.merge_cells('A1:C1')
ws3['A1'].value = "GOOGLE FORM — FIELD SETUP REFERENCE"
ws3['A1'].font = Font(name='Calibri', size=14, bold=True, color=WHITE)
ws3['A1'].fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 40

form_headers = ["Field Name", "Field Type", "Options / Notes"]
for i, h in enumerate(form_headers):
    cell = ws3.cell(row=3, column=i+1, value=h)
    cell.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

form_fields = [
    ["Customer Name", "Short Answer", "Required"],
    ["Phone Number", "Short Answer", "Required, Number validation"],
    ["Alternate Phone", "Short Answer", "Optional"],
    ["Email", "Short Answer", "Email validation"],
    ["City", "Dropdown", "Bangalore, Mumbai, Delhi, Hyderabad, Chennai, Pune, Other"],
    ["Full Address", "Paragraph", "Required"],
    ["Service Required", "Dropdown", "Bathroom Cleaning, Kitchen Cleaning, Full Home Cleaning, Rental Property Cleaning, Ready to Move In Cleaning, Painting, Pest Control, Plumbing, Electrician, Other"],
    ["Preferred Date", "Date", "Date picker"],
    ["Preferred Time Slot", "Dropdown", "Morning (8AM-10AM), Morning (10AM-12PM), Afternoon (12PM-2PM), Afternoon (2PM-4PM), Evening (4PM-6PM), Evening (6PM-8PM)"],
    ["Any Special Instructions", "Paragraph", "Optional"],
]

for i, (name, ftype, notes) in enumerate(form_fields):
    row = 4 + i
    ws3.cell(row=row, column=1, value=name).border = thin_border
    ws3.cell(row=row, column=1).font = Font(name='Calibri', size=10, bold=True)
    ws3.cell(row=row, column=2, value=ftype).border = thin_border
    ws3.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws3.cell(row=row, column=3, value=notes).border = thin_border
    if i % 2 == 0:
        for c in range(1, 4):
            ws3.cell(row=row, column=c).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 60

# Instructions
inst_row = 4 + len(form_fields) + 2
ws3.merge_cells(f'A{inst_row}:C{inst_row}')
ws3[f'A{inst_row}'].value = "HOW TO LINK GOOGLE FORM TO THIS SHEET"
ws3[f'A{inst_row}'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
ws3[f'A{inst_row}'].fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type='solid')
ws3[f'A{inst_row}'].alignment = Alignment(horizontal='center', vertical='center')

instructions = [
    "1. Open Google Forms (forms.google.com) and create a new form",
    "2. Add all the fields listed above with the specified field types and options",
    "3. Click on 'Responses' tab in Google Forms",
    "4. Click the green Sheets icon to 'Create Spreadsheet'",
    "5. Select 'Select existing spreadsheet' and choose THIS spreadsheet",
    "6. Map the form responses to go into the 'Lead Tracker' sheet starting from Row 4",
    "7. The Order ID will auto-generate once Timestamp is populated",
    "8. All other columns (status, vendor, payment) are filled manually by your team",
    "",
    "TIP: After uploading this file to Google Sheets, go to File → Import and select this .xlsx file",
    "All formatting, formulas, dropdowns, and conditional formatting will be preserved!",
]

for i, inst in enumerate(instructions):
    row = inst_row + 1 + i
    ws3.merge_cells(f'A{row}:C{row}')
    ws3[f'A{row}'].value = inst
    ws3[f'A{row}'].font = Font(name='Calibri', size=10)
    ws3[f'A{row}'].alignment = Alignment(vertical='center', wrap_text=True)

# ============================================================
# SAVE
# ============================================================
filepath = "/var/lib/freelancer/projects/40182876/Home_Services_Lead_Tracker.xlsx"
wb.save(filepath)
print(f"File saved to: {filepath}")
print("Done!")
