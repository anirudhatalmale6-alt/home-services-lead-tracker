from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, PieChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import random

wb = Workbook()

# Styles
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_row_fill = PatternFill("solid", fgColor="D6EAF8")
currency_format = '₹#,##0.00'
percent_format = '0.00%'
date_format = 'DD-MMM-YYYY'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

def style_data_rows(ws, start_row, end_row, cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = alt_row_fill

def auto_width(ws, cols):
    for col in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

# ============ HOW TO USE ============
ws_how = wb.active
ws_how.title = "HOW TO USE"
instructions = [
    ["", "MIS SYSTEM - USER GUIDE", ""],
    [""],
    ["SHEET", "PURPOSE", "HOW TO USE"],
    ["Income Tracker", "Record all service revenue with GST", "Enter each payment with service type, GST rate auto-calculates"],
    ["Expense Tracker", "Track all expenses with GST", "Log every expense with category and GST details"],
    ["Stock Purchases", "Track all stock/inventory purchases", "Record vendor purchases with invoice and GST"],
    ["Machines & Equipment", "Asset register for all machines", "Add each machine with unique ID (M001, M002...)"],
    ["Machine Maintenance", "Track repairs and servicing", "Log every repair/service against machine ID"],
    ["Chemicals Stock", "Manage chemical inventory", "Update stock added and used regularly"],
    ["Accessories Stock", "Track consumables inventory", "Update purchases and usage daily"],
    ["Stock Transactions", "Detailed stock movement log", "Record every purchase/usage transaction"],
    ["Income Summary", "Auto-calculated revenue reports", "View only - auto-updates from Income Tracker"],
    ["Expense Summary", "Auto-calculated expense reports", "View only - auto-updates from Expense Tracker"],
    ["GST Summary", "GST collected vs paid report", "View only - for GST filing reference"],
    ["Profitability Report", "Profit analysis", "View only - auto-calculates profit margins"],
    ["Stock Reports", "Inventory status and alerts", "View only - shows low stock and asset values"],
    ["Dashboard", "Visual overview with charts", "View only - key metrics at a glance"],
    [""],
    ["GST RATES:", "", ""],
    ["5% - Basic services", "", ""],
    ["12% - Standard services", "", ""],
    ["18% - Most services (default)", "", ""],
]
for row_data in instructions:
    ws_how.append(row_data)
ws_how.merge_cells('A1:C1')
ws_how['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws_how['A1'].alignment = center_align
style_header(ws_how, 3, 3)
for row in range(4, 18):
    for col in range(1, 4):
        ws_how.cell(row=row, column=col).border = thin_border
ws_how.column_dimensions['A'].width = 25
ws_how.column_dimensions['B'].width = 35
ws_how.column_dimensions['C'].width = 50

# ============ INCOME TRACKER WITH GST ============
ws_income = wb.create_sheet("Income Tracker")
income_headers = ["Date", "Invoice No", "Customer Name", "Service Type", "Project Type", "Project Name",
                  "Base Amount (₹)", "GST %", "GST Amount (₹)", "Total Amount (₹)", "Payment Mode", "Payment Status", "Notes"]
ws_income.append(income_headers)
style_header(ws_income, 1, len(income_headers))

services = ["Deep Cleaning", "Regular Cleaning", "Pest Control", "Painting", "Plumbing", "Electrical", "Carpentry", "AC Service"]
project_types = ["Individual", "Apartment Bulk", "Commercial"]
payment_modes = ["Cash", "UPI", "Bank Transfer", "Card"]
payment_status = ["Received", "Pending"]
customers = ["Rajesh Kumar", "Priya Sharma", "Amit Patel", "Sunita Reddy", "Vikram Singh", "Deepa Nair", "Karthik Iyer", "Meena Gupta", "Rahul Verma", "Anjali Menon", "Suresh Rao", "Lakshmi Pillai", "Arun Krishnan", "Kavitha Srinivasan", "Manoj Das"]

base_date = datetime(2025, 1, 1)
for i in range(15):
    row_num = i + 2
    date = base_date + timedelta(days=random.randint(0, 28))
    invoice = f"INV-2025-{1001+i}"
    customer = customers[i]
    service = random.choice(services)
    proj_type = random.choice(project_types)
    proj_name = "Green Valley Apartments" if proj_type == "Apartment Bulk" else ("Tech Park" if proj_type == "Commercial" else "")
    base_amount = random.randint(2000, 25000) if proj_type == "Individual" else random.randint(50000, 200000)
    gst_rate = 0.18
    mode = random.choice(payment_modes)
    status = random.choice(payment_status)

    ws_income.cell(row=row_num, column=1).value = date
    ws_income.cell(row=row_num, column=2).value = invoice
    ws_income.cell(row=row_num, column=3).value = customer
    ws_income.cell(row=row_num, column=4).value = service
    ws_income.cell(row=row_num, column=5).value = proj_type
    ws_income.cell(row=row_num, column=6).value = proj_name
    ws_income.cell(row=row_num, column=7).value = base_amount
    ws_income.cell(row=row_num, column=8).value = gst_rate
    ws_income.cell(row=row_num, column=9).value = f"=G{row_num}*H{row_num}"
    ws_income.cell(row=row_num, column=10).value = f"=G{row_num}+I{row_num}"
    ws_income.cell(row=row_num, column=11).value = mode
    ws_income.cell(row=row_num, column=12).value = status

style_data_rows(ws_income, 2, 16, len(income_headers))
for col in ['A']: ws_income.column_dimensions[col].width = 12
for col in ['B']: ws_income.column_dimensions[col].width = 15
for col in ['C']: ws_income.column_dimensions[col].width = 18
for col in ['D', 'E']: ws_income.column_dimensions[col].width = 16
for col in ['F']: ws_income.column_dimensions[col].width = 22
for col in ['G', 'I', 'J']: ws_income.column_dimensions[col].width = 14
for col in ['H']: ws_income.column_dimensions[col].width = 8

for row in range(2, 17):
    ws_income.cell(row=row, column=1).number_format = date_format
    ws_income.cell(row=row, column=7).number_format = currency_format
    ws_income.cell(row=row, column=8).number_format = '0%'
    ws_income.cell(row=row, column=9).number_format = currency_format
    ws_income.cell(row=row, column=10).number_format = currency_format
ws_income.freeze_panes = 'A2'

dv_service = DataValidation(type="list", formula1='"Deep Cleaning,Regular Cleaning,Pest Control,Painting,Plumbing,Electrical,Carpentry,AC Service"', allow_blank=True)
dv_projtype = DataValidation(type="list", formula1='"Individual,Apartment Bulk,Commercial"', allow_blank=True)
dv_paymode = DataValidation(type="list", formula1='"Cash,UPI,Bank Transfer,Card"', allow_blank=True)
dv_paystatus = DataValidation(type="list", formula1='"Received,Pending"', allow_blank=True)
dv_gst = DataValidation(type="list", formula1='"5%,12%,18%"', allow_blank=True)
ws_income.add_data_validation(dv_service)
ws_income.add_data_validation(dv_projtype)
ws_income.add_data_validation(dv_paymode)
ws_income.add_data_validation(dv_paystatus)
ws_income.add_data_validation(dv_gst)
dv_service.add('D2:D1000')
dv_projtype.add('E2:E1000')
dv_paymode.add('K2:K1000')
dv_paystatus.add('L2:L1000')
dv_gst.add('H2:H1000')

# ============ EXPENSE TRACKER WITH GST ============
ws_expense = wb.create_sheet("Expense Tracker")
expense_headers = ["Date", "Expense ID", "Category", "Description", "Vendor/Payee",
                   "Base Amount (₹)", "GST %", "GST Amount (₹)", "Total Amount (₹)", "Payment Mode", "Payment Status", "Notes"]
ws_expense.append(expense_headers)
style_header(ws_expense, 1, len(expense_headers))

expense_cats = ["Staff Salary", "Vendor Payment", "Travel & Fuel", "Marketing", "Office Expenses", "Miscellaneous"]
vendors = ["Ramesh (Driver)", "Kumar Chemicals", "City Fuel Station", "Facebook Ads", "Airtel", "Office Depot"]

for i in range(20):
    row_num = i + 2
    date = base_date + timedelta(days=random.randint(0, 28))
    exp_id = f"EXP-2025-{101+i}"
    cat = random.choice(expense_cats)

    if cat == "Staff Salary":
        desc = f"Salary - {random.choice(['Driver', 'Technician', 'Helper', 'Supervisor'])}"
        vendor = random.choice(["Ramesh", "Kumar", "Suresh", "Anil"])
        base_amount = random.randint(15000, 35000)
        gst_rate = 0
    elif cat == "Travel & Fuel":
        desc = "Petrol/Diesel"
        vendor = "City Fuel Station"
        base_amount = random.randint(500, 3000)
        gst_rate = 0.18
    elif cat == "Marketing":
        desc = random.choice(["Facebook Ads", "Google Ads", "Pamphlets"])
        vendor = desc
        base_amount = random.randint(2000, 15000)
        gst_rate = 0.18
    else:
        desc = random.choice(["Office Supplies", "Utility Bill", "Misc Purchase"])
        vendor = random.choice(vendors)
        base_amount = random.randint(500, 5000)
        gst_rate = 0.18

    mode = random.choice(payment_modes)
    status = random.choice(payment_status)

    ws_expense.cell(row=row_num, column=1).value = date
    ws_expense.cell(row=row_num, column=2).value = exp_id
    ws_expense.cell(row=row_num, column=3).value = cat
    ws_expense.cell(row=row_num, column=4).value = desc
    ws_expense.cell(row=row_num, column=5).value = vendor
    ws_expense.cell(row=row_num, column=6).value = base_amount
    ws_expense.cell(row=row_num, column=7).value = gst_rate
    ws_expense.cell(row=row_num, column=8).value = f"=F{row_num}*G{row_num}"
    ws_expense.cell(row=row_num, column=9).value = f"=F{row_num}+H{row_num}"
    ws_expense.cell(row=row_num, column=10).value = mode
    ws_expense.cell(row=row_num, column=11).value = status

style_data_rows(ws_expense, 2, 21, len(expense_headers))
auto_width(ws_expense, len(expense_headers))
ws_expense.column_dimensions['D'].width = 25
ws_expense.column_dimensions['G'].width = 8

for row in range(2, 22):
    ws_expense.cell(row=row, column=1).number_format = date_format
    ws_expense.cell(row=row, column=6).number_format = currency_format
    ws_expense.cell(row=row, column=7).number_format = '0%'
    ws_expense.cell(row=row, column=8).number_format = currency_format
    ws_expense.cell(row=row, column=9).number_format = currency_format
ws_expense.freeze_panes = 'A2'

dv_expcat = DataValidation(type="list", formula1='"Staff Salary,Vendor Payment,Travel & Fuel,Marketing,Office Expenses,Miscellaneous"', allow_blank=True)
ws_expense.add_data_validation(dv_expcat)
ws_expense.add_data_validation(dv_paymode)
ws_expense.add_data_validation(dv_paystatus)
ws_expense.add_data_validation(dv_gst)
dv_expcat.add('C2:C1000')
dv_gst.add('G2:G1000')

# ============ STOCK PURCHASES ============
ws_stock_purch = wb.create_sheet("Stock Purchases")
stock_purch_headers = ["Date", "Purchase ID", "Vendor Name", "Vendor GSTIN", "Invoice No", "Item Type",
                       "Item Name", "Quantity", "Unit", "Rate (₹)", "Base Amount (₹)", "GST %", "GST Amount (₹)",
                       "Total Amount (₹)", "Payment Status", "Notes"]
ws_stock_purch.append(stock_purch_headers)
style_header(ws_stock_purch, 1, len(stock_purch_headers))

stock_vendors = [
    ("Kumar Chemicals Pvt Ltd", "29AABCK1234A1ZV"),
    ("City Supplies", "29AABCS5678B2ZW"),
    ("Metro Cleaning Supplies", "29AABCM9012C3ZX"),
    ("SafeGuard Equipment", "29AABCG3456D4ZY"),
]

stock_items = [
    ("Chemical", "Floor Cleaner", "Liters", 120),
    ("Chemical", "Glass Cleaner", "Liters", 95),
    ("Chemical", "Pesticide Spray", "Cans", 350),
    ("Accessory", "Microfiber Cloth", "Pieces", 80),
    ("Accessory", "Rubber Gloves", "Pairs", 45),
    ("Accessory", "Floor Mop", "Pieces", 250),
    ("Chemical", "Surface Disinfectant", "Liters", 180),
    ("Accessory", "Scrub Brush", "Pieces", 45),
]

for i in range(15):
    row_num = i + 2
    date = base_date + timedelta(days=random.randint(0, 28))
    purch_id = f"PO-2025-{201+i}"
    vendor, gstin = random.choice(stock_vendors)
    inv_no = f"INV/{random.randint(1000, 9999)}"
    item_type, item_name, unit, rate = random.choice(stock_items)
    qty = random.randint(10, 100)
    gst_rate = 0.18

    ws_stock_purch.cell(row=row_num, column=1).value = date
    ws_stock_purch.cell(row=row_num, column=2).value = purch_id
    ws_stock_purch.cell(row=row_num, column=3).value = vendor
    ws_stock_purch.cell(row=row_num, column=4).value = gstin
    ws_stock_purch.cell(row=row_num, column=5).value = inv_no
    ws_stock_purch.cell(row=row_num, column=6).value = item_type
    ws_stock_purch.cell(row=row_num, column=7).value = item_name
    ws_stock_purch.cell(row=row_num, column=8).value = qty
    ws_stock_purch.cell(row=row_num, column=9).value = unit
    ws_stock_purch.cell(row=row_num, column=10).value = rate
    ws_stock_purch.cell(row=row_num, column=11).value = f"=H{row_num}*J{row_num}"
    ws_stock_purch.cell(row=row_num, column=12).value = gst_rate
    ws_stock_purch.cell(row=row_num, column=13).value = f"=K{row_num}*L{row_num}"
    ws_stock_purch.cell(row=row_num, column=14).value = f"=K{row_num}+M{row_num}"
    ws_stock_purch.cell(row=row_num, column=15).value = random.choice(["Paid", "Pending"])

style_data_rows(ws_stock_purch, 2, 16, len(stock_purch_headers))
auto_width(ws_stock_purch, len(stock_purch_headers))
ws_stock_purch.column_dimensions['C'].width = 25
ws_stock_purch.column_dimensions['D'].width = 18
ws_stock_purch.column_dimensions['G'].width = 20

for row in range(2, 17):
    ws_stock_purch.cell(row=row, column=1).number_format = date_format
    ws_stock_purch.cell(row=row, column=10).number_format = currency_format
    ws_stock_purch.cell(row=row, column=11).number_format = currency_format
    ws_stock_purch.cell(row=row, column=12).number_format = '0%'
    ws_stock_purch.cell(row=row, column=13).number_format = currency_format
    ws_stock_purch.cell(row=row, column=14).number_format = currency_format
ws_stock_purch.freeze_panes = 'A2'

dv_itemtype = DataValidation(type="list", formula1='"Chemical,Accessory,Machine,Spare Part"', allow_blank=True)
dv_purch_status = DataValidation(type="list", formula1='"Paid,Pending"', allow_blank=True)
ws_stock_purch.add_data_validation(dv_itemtype)
ws_stock_purch.add_data_validation(dv_purch_status)
ws_stock_purch.add_data_validation(dv_gst)
dv_itemtype.add('F2:F1000')
dv_purch_status.add('O2:O1000')
dv_gst.add('L2:L1000')

# ============ MACHINES & EQUIPMENT ============
ws_machines = wb.create_sheet("Machines & Equipment")
machine_headers = ["Machine ID", "Machine Name", "Model", "Category", "Purchase Date", "Base Cost (₹)",
                   "GST %", "GST Amount (₹)", "Total Cost (₹)", "Current Status", "Location/Assigned To", "Notes"]
ws_machines.append(machine_headers)
style_header(ws_machines, 1, len(machine_headers))

machine_cats = ["Cleaning Machine", "Spray Equipment", "Power Tools", "Safety Equipment"]
machine_data = [
    ["M001", "Floor Scrubber", "Karcher BD 50/50", "Cleaning Machine", datetime(2023, 3, 15), 72034, 0.18, "Active", "Team A"],
    ["M002", "Pressure Washer", "Bosch AQT 45-14X", "Cleaning Machine", datetime(2023, 5, 20), 38136, 0.18, "Active", "Team B"],
    ["M003", "Vacuum Cleaner", "Eureka Forbes Pro", "Cleaning Machine", datetime(2022, 8, 10), 23729, 0.18, "Active", "Team A"],
    ["M004", "Pest Control Sprayer", "Solo 475", "Spray Equipment", datetime(2023, 1, 5), 12712, 0.18, "Active", "Pest Team"],
    ["M005", "Paint Sprayer", "Graco Magnum X5", "Spray Equipment", datetime(2023, 7, 12), 55085, 0.18, "Under Repair", "Painting Team"],
    ["M006", "Drill Machine", "Bosch GSB 600", "Power Tools", datetime(2022, 4, 18), 7203, 0.18, "Active", "General"],
    ["M007", "Angle Grinder", "Makita GA4030", "Power Tools", datetime(2023, 2, 22), 5508, 0.18, "Active", "General"],
    ["M008", "Ladder (12ft)", "Aluminium", "Safety Equipment", datetime(2022, 1, 10), 10169, 0.18, "Active", "Warehouse"],
    ["M009", "Steam Cleaner", "Karcher SC3", "Cleaning Machine", datetime(2024, 1, 8), 29661, 0.18, "Active", "Team C"],
    ["M010", "Industrial Blower", "Dewalt DCE100", "Cleaning Machine", datetime(2023, 11, 5), 15254, 0.18, "Retired", "Storage"],
]

for i, row in enumerate(machine_data, start=2):
    ws_machines.cell(row=i, column=1).value = row[0]
    ws_machines.cell(row=i, column=2).value = row[1]
    ws_machines.cell(row=i, column=3).value = row[2]
    ws_machines.cell(row=i, column=4).value = row[3]
    ws_machines.cell(row=i, column=5).value = row[4]
    ws_machines.cell(row=i, column=6).value = row[5]
    ws_machines.cell(row=i, column=7).value = row[6]
    ws_machines.cell(row=i, column=8).value = f"=F{i}*G{i}"
    ws_machines.cell(row=i, column=9).value = f"=F{i}+H{i}"
    ws_machines.cell(row=i, column=10).value = row[7]
    ws_machines.cell(row=i, column=11).value = row[8]

style_data_rows(ws_machines, 2, 11, len(machine_headers))
auto_width(ws_machines, len(machine_headers))
ws_machines.column_dimensions['B'].width = 20
ws_machines.column_dimensions['C'].width = 20

for row in range(2, 12):
    ws_machines.cell(row=row, column=5).number_format = date_format
    ws_machines.cell(row=row, column=6).number_format = currency_format
    ws_machines.cell(row=row, column=7).number_format = '0%'
    ws_machines.cell(row=row, column=8).number_format = currency_format
    ws_machines.cell(row=row, column=9).number_format = currency_format
ws_machines.freeze_panes = 'A2'

dv_machcat = DataValidation(type="list", formula1='"Cleaning Machine,Spray Equipment,Power Tools,Safety Equipment"', allow_blank=True)
dv_machstatus = DataValidation(type="list", formula1='"Active,Under Repair,Retired"', allow_blank=True)
ws_machines.add_data_validation(dv_machcat)
ws_machines.add_data_validation(dv_machstatus)
ws_machines.add_data_validation(dv_gst)
dv_machcat.add('D2:D1000')
dv_machstatus.add('J2:J1000')
dv_gst.add('G2:G1000')

# ============ MACHINE MAINTENANCE LOG ============
ws_maint = wb.create_sheet("Machine Maintenance")
maint_headers = ["Date", "Machine ID", "Machine Name", "Maintenance Type", "Description", "Base Cost (₹)",
                 "GST %", "GST Amount (₹)", "Total Cost (₹)", "Done By", "Next Service Due", "Notes"]
ws_maint.append(maint_headers)
style_header(ws_maint, 1, len(maint_headers))

maint_types = ["Repair", "Service", "Parts Replacement"]
maint_data = [
    [datetime(2024, 2, 10), "M001", "Floor Scrubber", "Service", "Regular servicing", 2119, 0.18, "Karcher Service", datetime(2024, 8, 10)],
    [datetime(2024, 3, 15), "M002", "Pressure Washer", "Parts Replacement", "Nozzle replaced", 1017, 0.18, "Local Mechanic", datetime(2024, 9, 15)],
    [datetime(2024, 4, 5), "M003", "Vacuum Cleaner", "Service", "Filter cleaning", 678, 0.18, "In-house", datetime(2024, 10, 5)],
    [datetime(2024, 5, 20), "M005", "Paint Sprayer", "Repair", "Motor rewinding", 7203, 0.18, "Graco Service", datetime(2024, 11, 20)],
    [datetime(2024, 6, 8), "M004", "Pest Control Sprayer", "Service", "Valve check", 424, 0.18, "In-house", datetime(2024, 12, 8)],
    [datetime(2024, 7, 12), "M006", "Drill Machine", "Parts Replacement", "Chuck replaced", 1271, 0.18, "Bosch Service", datetime(2025, 1, 12)],
    [datetime(2024, 8, 1), "M001", "Floor Scrubber", "Service", "Brush replacement", 3814, 0.18, "Karcher Service", datetime(2025, 2, 1)],
    [datetime(2024, 8, 18), "M007", "Angle Grinder", "Repair", "Switch repair", 508, 0.18, "Local Mechanic", datetime(2025, 2, 18)],
    [datetime(2024, 9, 5), "M009", "Steam Cleaner", "Service", "Descaling", 1017, 0.18, "Karcher Service", datetime(2025, 3, 5)],
    [datetime(2024, 10, 10), "M002", "Pressure Washer", "Service", "Regular maintenance", 1525, 0.18, "Bosch Service", datetime(2025, 4, 10)],
]

for i, row in enumerate(maint_data, start=2):
    ws_maint.cell(row=i, column=1).value = row[0]
    ws_maint.cell(row=i, column=2).value = row[1]
    ws_maint.cell(row=i, column=3).value = row[2]
    ws_maint.cell(row=i, column=4).value = row[3]
    ws_maint.cell(row=i, column=5).value = row[4]
    ws_maint.cell(row=i, column=6).value = row[5]
    ws_maint.cell(row=i, column=7).value = row[6]
    ws_maint.cell(row=i, column=8).value = f"=F{i}*G{i}"
    ws_maint.cell(row=i, column=9).value = f"=F{i}+H{i}"
    ws_maint.cell(row=i, column=10).value = row[7]
    ws_maint.cell(row=i, column=11).value = row[8]

style_data_rows(ws_maint, 2, 11, len(maint_headers))
auto_width(ws_maint, len(maint_headers))
ws_maint.column_dimensions['E'].width = 25

for row in range(2, 12):
    ws_maint.cell(row=row, column=1).number_format = date_format
    ws_maint.cell(row=row, column=6).number_format = currency_format
    ws_maint.cell(row=row, column=7).number_format = '0%'
    ws_maint.cell(row=row, column=8).number_format = currency_format
    ws_maint.cell(row=row, column=9).number_format = currency_format
    ws_maint.cell(row=row, column=11).number_format = date_format
ws_maint.freeze_panes = 'A2'

dv_mainttype = DataValidation(type="list", formula1='"Repair,Service,Parts Replacement"', allow_blank=True)
dv_machid = DataValidation(type="list", formula1='"M001,M002,M003,M004,M005,M006,M007,M008,M009,M010"', allow_blank=True)
ws_maint.add_data_validation(dv_mainttype)
ws_maint.add_data_validation(dv_machid)
ws_maint.add_data_validation(dv_gst)
dv_mainttype.add('D2:D1000')
dv_machid.add('B2:B1000')
dv_gst.add('G2:G1000')

# ============ CHEMICALS STOCK ============
ws_chem = wb.create_sheet("Chemicals Stock")
chem_headers = ["Item ID", "Chemical Name", "Brand", "Category", "Unit", "Opening Stock", "Stock Added", "Stock Used", "Closing Stock", "Min Level", "Status", "Last Updated"]
ws_chem.append(chem_headers)
style_header(ws_chem, 1, len(chem_headers))

chem_data = [
    ["CH001", "Floor Cleaner", "Lizol", "Cleaning Agents", "Liters", 50, 100, 80, None, 20, None, datetime(2025, 1, 25)],
    ["CH002", "Glass Cleaner", "Colin", "Cleaning Agents", "Liters", 30, 50, 45, None, 15, None, datetime(2025, 1, 25)],
    ["CH003", "Toilet Cleaner", "Harpic", "Cleaning Agents", "Liters", 40, 80, 70, None, 20, None, datetime(2025, 1, 25)],
    ["CH004", "Pesticide Spray", "Baygon Pro", "Pest Control", "Cans", 25, 50, 55, None, 10, None, datetime(2025, 1, 25)],
    ["CH005", "Termite Solution", "Premise", "Pest Control", "Liters", 20, 40, 35, None, 10, None, datetime(2025, 1, 25)],
    ["CH006", "Hand Sanitizer", "Dettol", "Sanitizers", "Liters", 15, 30, 28, None, 10, None, datetime(2025, 1, 25)],
    ["CH007", "Surface Disinfectant", "Lysol", "Sanitizers", "Liters", 25, 50, 40, None, 15, None, datetime(2025, 1, 25)],
    ["CH008", "Degreaser", "WD-40", "Specialty", "Cans", 10, 20, 18, None, 5, None, datetime(2025, 1, 25)],
    ["CH009", "Stain Remover", "Vanish", "Specialty", "KG", 8, 15, 12, None, 5, None, datetime(2025, 1, 25)],
    ["CH010", "Air Freshener", "Odonil", "Specialty", "Cans", 30, 60, 65, None, 15, None, datetime(2025, 1, 25)],
    ["CH011", "Cockroach Gel", "Advion", "Pest Control", "Bottles", 20, 30, 25, None, 10, None, datetime(2025, 1, 25)],
    ["CH012", "Fabric Softener", "Comfort", "Cleaning Agents", "Liters", 15, 25, 22, None, 10, None, datetime(2025, 1, 25)],
]

for i, row in enumerate(chem_data, start=2):
    ws_chem.append(row[:8])
    ws_chem.cell(row=i, column=9).value = f"=F{i}+G{i}-H{i}"
    ws_chem.cell(row=i, column=10).value = row[9]
    ws_chem.cell(row=i, column=11).value = f'=IF(I{i}<J{i},"Low Stock","OK")'
    ws_chem.cell(row=i, column=12).value = row[11]

style_data_rows(ws_chem, 2, 13, len(chem_headers))
auto_width(ws_chem, len(chem_headers))
ws_chem.column_dimensions['B'].width = 22
for row in range(2, 14):
    ws_chem.cell(row=row, column=12).number_format = date_format
ws_chem.freeze_panes = 'A2'

dv_chemcat = DataValidation(type="list", formula1='"Cleaning Agents,Pest Control,Sanitizers,Specialty"', allow_blank=True)
dv_chemunit = DataValidation(type="list", formula1='"Liters,KG,Bottles,Cans"', allow_blank=True)
ws_chem.add_data_validation(dv_chemcat)
ws_chem.add_data_validation(dv_chemunit)
dv_chemcat.add('D2:D1000')
dv_chemunit.add('E2:E1000')

low_stock_fill = PatternFill("solid", fgColor="FFCCCC")
ok_fill = PatternFill("solid", fgColor="CCFFCC")
ws_chem.conditional_formatting.add('K2:K1000', FormulaRule(formula=['K2="Low Stock"'], fill=low_stock_fill))
ws_chem.conditional_formatting.add('K2:K1000', FormulaRule(formula=['K2="OK"'], fill=ok_fill))

# ============ ACCESSORIES STOCK ============
ws_acc = wb.create_sheet("Accessories Stock")
acc_headers = ["Item ID", "Item Name", "Category", "Unit", "Opening Stock", "Stock Added", "Stock Used", "Closing Stock", "Min Level", "Status", "Last Updated"]
ws_acc.append(acc_headers)
style_header(ws_acc, 1, len(acc_headers))

acc_data = [
    ["AC001", "Rubber Hose (10m)", "Pipes & Hoses", "Pieces", 10, 15, 12, None, 5, None, datetime(2025, 1, 25)],
    ["AC002", "Floor Mop", "Brushes & Mops", "Pieces", 20, 30, 28, None, 10, None, datetime(2025, 1, 25)],
    ["AC003", "Scrub Brush", "Brushes & Mops", "Pieces", 25, 40, 35, None, 10, None, datetime(2025, 1, 25)],
    ["AC004", "Microfiber Cloth", "Cloths & Wipes", "Pieces", 50, 100, 90, None, 25, None, datetime(2025, 1, 25)],
    ["AC005", "Rubber Gloves", "Gloves & Safety", "Pairs", 30, 50, 48, None, 15, None, datetime(2025, 1, 25)],
    ["AC006", "Safety Goggles", "Gloves & Safety", "Pieces", 10, 15, 8, None, 5, None, datetime(2025, 1, 25)],
    ["AC007", "Spray Bottles", "Other", "Pieces", 20, 30, 22, None, 10, None, datetime(2025, 1, 25)],
    ["AC008", "Dust Pan Set", "Brushes & Mops", "Sets", 15, 20, 18, None, 8, None, datetime(2025, 1, 25)],
    ["AC009", "Extension Cord", "Tools", "Pieces", 8, 10, 7, None, 4, None, datetime(2025, 1, 25)],
    ["AC010", "Bucket (20L)", "Other", "Pieces", 25, 30, 28, None, 10, None, datetime(2025, 1, 25)],
    ["AC011", "Face Masks", "Gloves & Safety", "Boxes", 10, 20, 18, None, 5, None, datetime(2025, 1, 25)],
    ["AC012", "Sponge Pack", "Cloths & Wipes", "Packs", 20, 35, 30, None, 10, None, datetime(2025, 1, 25)],
]

for i, row in enumerate(acc_data, start=2):
    ws_acc.append(row[:7])
    ws_acc.cell(row=i, column=8).value = f"=E{i}+F{i}-G{i}"
    ws_acc.cell(row=i, column=9).value = row[8]
    ws_acc.cell(row=i, column=10).value = f'=IF(H{i}<I{i},"Low Stock","OK")'
    ws_acc.cell(row=i, column=11).value = row[10]

style_data_rows(ws_acc, 2, 13, len(acc_headers))
auto_width(ws_acc, len(acc_headers))
ws_acc.column_dimensions['B'].width = 20
for row in range(2, 14):
    ws_acc.cell(row=row, column=11).number_format = date_format
ws_acc.freeze_panes = 'A2'

dv_acccat = DataValidation(type="list", formula1='"Pipes & Hoses,Brushes & Mops,Cloths & Wipes,Gloves & Safety,Tools,Other"', allow_blank=True)
ws_acc.add_data_validation(dv_acccat)
dv_acccat.add('C2:C1000')
ws_acc.conditional_formatting.add('J2:J1000', FormulaRule(formula=['J2="Low Stock"'], fill=low_stock_fill))
ws_acc.conditional_formatting.add('J2:J1000', FormulaRule(formula=['J2="OK"'], fill=ok_fill))

# ============ STOCK TRANSACTIONS ============
ws_trans = wb.create_sheet("Stock Transactions")
trans_headers = ["Date", "Transaction ID", "Item Type", "Item ID", "Item Name", "Transaction Type",
                 "Quantity", "Unit Rate (₹)", "GST %", "GST Amount (₹)", "Total Cost (₹)", "Project/Team", "Notes"]
ws_trans.append(trans_headers)
style_header(ws_trans, 1, len(trans_headers))

trans_data = [
    [datetime(2025, 1, 2), "TXN001", "Chemical", "CH001", "Floor Cleaner", "Purchase", 50, 120, 0.18, None, None, "Warehouse", "Monthly stock"],
    [datetime(2025, 1, 3), "TXN002", "Chemical", "CH004", "Pesticide Spray", "Purchase", 25, 350, 0.18, None, None, "Warehouse", ""],
    [datetime(2025, 1, 5), "TXN003", "Accessory", "AC004", "Microfiber Cloth", "Purchase", 50, 80, 0.18, None, None, "Warehouse", "Bulk order"],
    [datetime(2025, 1, 6), "TXN004", "Chemical", "CH001", "Floor Cleaner", "Used", 20, 0, 0, None, None, "Team A", "Green Valley project"],
    [datetime(2025, 1, 7), "TXN005", "Accessory", "AC005", "Rubber Gloves", "Used", 10, 0, 0, None, None, "Team B", ""],
    [datetime(2025, 1, 8), "TXN006", "Chemical", "CH002", "Glass Cleaner", "Purchase", 25, 95, 0.18, None, None, "Warehouse", ""],
    [datetime(2025, 1, 10), "TXN007", "Accessory", "AC002", "Floor Mop", "Damaged", 2, 0, 0, None, None, "Team A", "Worn out"],
    [datetime(2025, 1, 11), "TXN008", "Chemical", "CH007", "Surface Disinfectant", "Used", 15, 0, 0, None, None, "Team C", "Hospital contract"],
    [datetime(2025, 1, 12), "TXN009", "Accessory", "AC003", "Scrub Brush", "Purchase", 20, 45, 0.18, None, None, "Warehouse", ""],
    [datetime(2025, 1, 14), "TXN010", "Chemical", "CH005", "Termite Solution", "Purchase", 20, 850, 0.18, None, None, "Warehouse", ""],
]

for i, row in enumerate(trans_data, start=2):
    ws_trans.cell(row=i, column=1).value = row[0]
    ws_trans.cell(row=i, column=2).value = row[1]
    ws_trans.cell(row=i, column=3).value = row[2]
    ws_trans.cell(row=i, column=4).value = row[3]
    ws_trans.cell(row=i, column=5).value = row[4]
    ws_trans.cell(row=i, column=6).value = row[5]
    ws_trans.cell(row=i, column=7).value = row[6]
    ws_trans.cell(row=i, column=8).value = row[7]
    ws_trans.cell(row=i, column=9).value = row[8]
    ws_trans.cell(row=i, column=10).value = f"=G{i}*H{i}*I{i}"
    ws_trans.cell(row=i, column=11).value = f"=G{i}*H{i}+J{i}"
    ws_trans.cell(row=i, column=12).value = row[11]
    ws_trans.cell(row=i, column=13).value = row[12]

style_data_rows(ws_trans, 2, 11, len(trans_headers))
auto_width(ws_trans, len(trans_headers))
for row in range(2, 12):
    ws_trans.cell(row=row, column=1).number_format = date_format
    ws_trans.cell(row=row, column=8).number_format = currency_format
    ws_trans.cell(row=row, column=9).number_format = '0%'
    ws_trans.cell(row=row, column=10).number_format = currency_format
    ws_trans.cell(row=row, column=11).number_format = currency_format
ws_trans.freeze_panes = 'A2'

dv_transtype = DataValidation(type="list", formula1='"Purchase,Used,Returned,Damaged"', allow_blank=True)
ws_trans.add_data_validation(dv_itemtype)
ws_trans.add_data_validation(dv_transtype)
ws_trans.add_data_validation(dv_gst)
dv_itemtype.add('C2:C1000')
dv_transtype.add('F2:F1000')
dv_gst.add('I2:I1000')

# ============ INCOME SUMMARY ============
ws_inc_sum = wb.create_sheet("Income Summary")

ws_inc_sum['A1'] = "INCOME SUMMARY REPORT"
ws_inc_sum.merge_cells('A1:E1')
ws_inc_sum['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_inc_sum['A3'] = "SERVICE-WISE REVENUE"
ws_inc_sum['A3'].font = Font(bold=True, size=12)
ws_inc_sum.append(["Service Type", "Base Revenue (₹)", "GST Collected (₹)", "Total Revenue (₹)", "Order Count"])
style_header(ws_inc_sum, 4, 5)

for i, svc in enumerate(services, start=5):
    ws_inc_sum.cell(row=i, column=1).value = svc
    ws_inc_sum.cell(row=i, column=2).value = f'=SUMIF(\'Income Tracker\'!D:D,A{i},\'Income Tracker\'!G:G)'
    ws_inc_sum.cell(row=i, column=3).value = f'=SUMIF(\'Income Tracker\'!D:D,A{i},\'Income Tracker\'!I:I)'
    ws_inc_sum.cell(row=i, column=4).value = f'=SUMIF(\'Income Tracker\'!D:D,A{i},\'Income Tracker\'!J:J)'
    ws_inc_sum.cell(row=i, column=5).value = f'=COUNTIF(\'Income Tracker\'!D:D,A{i})'
    ws_inc_sum.cell(row=i, column=2).number_format = currency_format
    ws_inc_sum.cell(row=i, column=3).number_format = currency_format
    ws_inc_sum.cell(row=i, column=4).number_format = currency_format

ws_inc_sum.cell(row=13, column=1).value = "TOTAL"
ws_inc_sum.cell(row=13, column=1).font = Font(bold=True)
ws_inc_sum.cell(row=13, column=2).value = "=SUM(B5:B12)"
ws_inc_sum.cell(row=13, column=3).value = "=SUM(C5:C12)"
ws_inc_sum.cell(row=13, column=4).value = "=SUM(D5:D12)"
ws_inc_sum.cell(row=13, column=5).value = "=SUM(E5:E12)"
for col in [2, 3, 4]:
    ws_inc_sum.cell(row=13, column=col).font = Font(bold=True)
    ws_inc_sum.cell(row=13, column=col).number_format = currency_format

style_data_rows(ws_inc_sum, 5, 13, 5)
auto_width(ws_inc_sum, 5)
ws_inc_sum.column_dimensions['A'].width = 20
ws_inc_sum.column_dimensions['B'].width = 18
ws_inc_sum.column_dimensions['C'].width = 18
ws_inc_sum.column_dimensions['D'].width = 18

# ============ EXPENSE SUMMARY ============
ws_exp_sum = wb.create_sheet("Expense Summary")

ws_exp_sum['A1'] = "EXPENSE SUMMARY REPORT"
ws_exp_sum.merge_cells('A1:E1')
ws_exp_sum['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_exp_sum['A3'] = "CATEGORY-WISE EXPENSES"
ws_exp_sum['A3'].font = Font(bold=True, size=12)
ws_exp_sum.append(["Category", "Base Amount (₹)", "GST Paid (₹)", "Total Amount (₹)", "Transaction Count"])
style_header(ws_exp_sum, 4, 5)

for i, cat in enumerate(expense_cats, start=5):
    ws_exp_sum.cell(row=i, column=1).value = cat
    ws_exp_sum.cell(row=i, column=2).value = f'=SUMIF(\'Expense Tracker\'!C:C,A{i},\'Expense Tracker\'!F:F)'
    ws_exp_sum.cell(row=i, column=3).value = f'=SUMIF(\'Expense Tracker\'!C:C,A{i},\'Expense Tracker\'!H:H)'
    ws_exp_sum.cell(row=i, column=4).value = f'=SUMIF(\'Expense Tracker\'!C:C,A{i},\'Expense Tracker\'!I:I)'
    ws_exp_sum.cell(row=i, column=5).value = f'=COUNTIF(\'Expense Tracker\'!C:C,A{i})'
    ws_exp_sum.cell(row=i, column=2).number_format = currency_format
    ws_exp_sum.cell(row=i, column=3).number_format = currency_format
    ws_exp_sum.cell(row=i, column=4).number_format = currency_format

ws_exp_sum.cell(row=11, column=1).value = "TOTAL"
ws_exp_sum.cell(row=11, column=1).font = Font(bold=True)
ws_exp_sum.cell(row=11, column=2).value = "=SUM(B5:B10)"
ws_exp_sum.cell(row=11, column=3).value = "=SUM(C5:C10)"
ws_exp_sum.cell(row=11, column=4).value = "=SUM(D5:D10)"
ws_exp_sum.cell(row=11, column=5).value = "=SUM(E5:E10)"
for col in [2, 3, 4]:
    ws_exp_sum.cell(row=11, column=col).font = Font(bold=True)
    ws_exp_sum.cell(row=11, column=col).number_format = currency_format

style_data_rows(ws_exp_sum, 5, 11, 5)
auto_width(ws_exp_sum, 5)
ws_exp_sum.column_dimensions['A'].width = 18

# ============ GST SUMMARY ============
ws_gst = wb.create_sheet("GST Summary")

ws_gst['A1'] = "GST SUMMARY REPORT"
ws_gst.merge_cells('A1:D1')
ws_gst['A1'].font = Font(bold=True, size=14, color="1F4E79")
ws_gst['A2'] = "(For GST Filing Reference)"
ws_gst['A2'].font = Font(italic=True, color="666666")

ws_gst['A4'] = "GST COLLECTED (OUTPUT TAX)"
ws_gst['A4'].font = Font(bold=True, size=12, color="008000")
ws_gst.append(["Source", "Base Amount (₹)", "GST Amount (₹)"])
style_header(ws_gst, 5, 3)

ws_gst['A6'] = "Income from Services"
ws_gst['B6'] = "=SUM('Income Tracker'!G:G)"
ws_gst['C6'] = "=SUM('Income Tracker'!I:I)"
ws_gst['B6'].number_format = currency_format
ws_gst['C6'].number_format = currency_format

ws_gst['A7'] = "TOTAL GST COLLECTED"
ws_gst['A7'].font = Font(bold=True)
ws_gst['B7'] = "=B6"
ws_gst['C7'] = "=C6"
ws_gst['B7'].font = Font(bold=True, color="008000")
ws_gst['C7'].font = Font(bold=True, color="008000")
ws_gst['B7'].number_format = currency_format
ws_gst['C7'].number_format = currency_format

style_data_rows(ws_gst, 6, 7, 3)

ws_gst['A10'] = "GST PAID (INPUT TAX)"
ws_gst['A10'].font = Font(bold=True, size=12, color="FF0000")
ws_gst.append(["Source", "Base Amount (₹)", "GST Amount (₹)"])
style_header(ws_gst, 11, 3)

ws_gst['A12'] = "Expenses"
ws_gst['B12'] = "=SUM('Expense Tracker'!F:F)"
ws_gst['C12'] = "=SUM('Expense Tracker'!H:H)"
ws_gst['A13'] = "Stock Purchases"
ws_gst['B13'] = "=SUM('Stock Purchases'!K:K)"
ws_gst['C13'] = "=SUM('Stock Purchases'!M:M)"
ws_gst['A14'] = "Machine Purchases"
ws_gst['B14'] = "=SUM('Machines & Equipment'!F:F)"
ws_gst['C14'] = "=SUM('Machines & Equipment'!H:H)"
ws_gst['A15'] = "Machine Maintenance"
ws_gst['B15'] = "=SUM('Machine Maintenance'!F:F)"
ws_gst['C15'] = "=SUM('Machine Maintenance'!H:H)"

for row in [12, 13, 14, 15]:
    ws_gst.cell(row=row, column=2).number_format = currency_format
    ws_gst.cell(row=row, column=3).number_format = currency_format

ws_gst['A16'] = "TOTAL GST PAID"
ws_gst['A16'].font = Font(bold=True)
ws_gst['B16'] = "=SUM(B12:B15)"
ws_gst['C16'] = "=SUM(C12:C15)"
ws_gst['B16'].font = Font(bold=True, color="FF0000")
ws_gst['C16'].font = Font(bold=True, color="FF0000")
ws_gst['B16'].number_format = currency_format
ws_gst['C16'].number_format = currency_format

style_data_rows(ws_gst, 12, 16, 3)

ws_gst['A19'] = "NET GST PAYABLE"
ws_gst['A19'].font = Font(bold=True, size=14)
ws_gst.merge_cells('A19:B19')
ws_gst['C19'] = "=C7-C16"
ws_gst['C19'].font = Font(bold=True, size=14, color="1F4E79")
ws_gst['C19'].number_format = currency_format
ws_gst['C19'].fill = PatternFill("solid", fgColor="FFFFCC")

ws_gst['A21'] = "Note: Positive value = GST to pay to government"
ws_gst['A21'].font = Font(italic=True, color="666666")
ws_gst['A22'] = "Negative value = GST credit (Input Tax Credit)"
ws_gst['A22'].font = Font(italic=True, color="666666")

auto_width(ws_gst, 3)
ws_gst.column_dimensions['A'].width = 25
ws_gst.column_dimensions['B'].width = 18
ws_gst.column_dimensions['C'].width = 18

# ============ PROFITABILITY REPORT ============
ws_profit = wb.create_sheet("Profitability Report")

ws_profit['A1'] = "PROFITABILITY ANALYSIS"
ws_profit.merge_cells('A1:D1')
ws_profit['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_profit['A3'] = "OVERALL PROFITABILITY"
ws_profit['A3'].font = Font(bold=True, size=12)
ws_profit.append(["Metric", "Amount (₹)"])
style_header(ws_profit, 4, 2)

ws_profit['A5'] = "Total Revenue (excl. GST)"
ws_profit['B5'] = "=SUM('Income Tracker'!G:G)"
ws_profit['A6'] = "Total Expenses (excl. GST)"
ws_profit['B6'] = "=SUM('Expense Tracker'!F:F)"
ws_profit['A7'] = "Gross Profit"
ws_profit['B7'] = "=B5-B6"
ws_profit['A8'] = "Profit Margin %"
ws_profit['B8'] = "=IF(B5>0,B7/B5,0)"
ws_profit.cell(row=8, column=2).number_format = '0.0%'

for row in range(5, 8):
    ws_profit.cell(row=row, column=2).number_format = currency_format
ws_profit['B7'].font = Font(bold=True, color="008000")

style_data_rows(ws_profit, 5, 8, 2)
auto_width(ws_profit, 2)
ws_profit.column_dimensions['A'].width = 25
ws_profit.column_dimensions['B'].width = 18

# ============ STOCK REPORTS ============
ws_stock_rep = wb.create_sheet("Stock Reports")

ws_stock_rep['A1'] = "STOCK & INVENTORY REPORTS"
ws_stock_rep.merge_cells('A1:E1')
ws_stock_rep['A1'].font = Font(bold=True, size=14, color="1F4E79")

ws_stock_rep['A3'] = "TOTAL ASSET VALUE"
ws_stock_rep['A3'].font = Font(bold=True, size=12)
ws_stock_rep.append(["Category", "Base Value (₹)", "GST Paid (₹)", "Total Value (₹)", "Count"])
style_header(ws_stock_rep, 4, 5)

for i, cat in enumerate(machine_cats, start=5):
    ws_stock_rep.cell(row=i, column=1).value = cat
    ws_stock_rep.cell(row=i, column=2).value = f'=SUMIF(\'Machines & Equipment\'!D:D,A{i},\'Machines & Equipment\'!F:F)'
    ws_stock_rep.cell(row=i, column=3).value = f'=SUMIF(\'Machines & Equipment\'!D:D,A{i},\'Machines & Equipment\'!H:H)'
    ws_stock_rep.cell(row=i, column=4).value = f'=SUMIF(\'Machines & Equipment\'!D:D,A{i},\'Machines & Equipment\'!I:I)'
    ws_stock_rep.cell(row=i, column=5).value = f'=COUNTIF(\'Machines & Equipment\'!D:D,A{i})'
    ws_stock_rep.cell(row=i, column=2).number_format = currency_format
    ws_stock_rep.cell(row=i, column=3).number_format = currency_format
    ws_stock_rep.cell(row=i, column=4).number_format = currency_format

ws_stock_rep.cell(row=9, column=1).value = "TOTAL"
ws_stock_rep.cell(row=9, column=1).font = Font(bold=True)
ws_stock_rep.cell(row=9, column=2).value = "=SUM(B5:B8)"
ws_stock_rep.cell(row=9, column=3).value = "=SUM(C5:C8)"
ws_stock_rep.cell(row=9, column=4).value = "=SUM(D5:D8)"
ws_stock_rep.cell(row=9, column=5).value = "=SUM(E5:E8)"
for col in [2, 3, 4]:
    ws_stock_rep.cell(row=9, column=col).font = Font(bold=True)
    ws_stock_rep.cell(row=9, column=col).number_format = currency_format

style_data_rows(ws_stock_rep, 5, 9, 5)

ws_stock_rep['A12'] = "STOCK PURCHASE SUMMARY"
ws_stock_rep['A12'].font = Font(bold=True, size=12)
ws_stock_rep['A13'] = "Total Stock Purchases (Base)"
ws_stock_rep['B13'] = "=SUM('Stock Purchases'!K:K)"
ws_stock_rep['A14'] = "Total GST on Purchases"
ws_stock_rep['B14'] = "=SUM('Stock Purchases'!M:M)"
ws_stock_rep['A15'] = "Total with GST"
ws_stock_rep['B15'] = "=SUM('Stock Purchases'!N:N)"
for row in [13, 14, 15]:
    ws_stock_rep.cell(row=row, column=2).number_format = currency_format
ws_stock_rep['B15'].font = Font(bold=True)

auto_width(ws_stock_rep, 5)
ws_stock_rep.column_dimensions['A'].width = 25

# ============ DASHBOARD ============
ws_dash = wb.create_sheet("Dashboard")

ws_dash['A1'] = "MIS DASHBOARD"
ws_dash.merge_cells('A1:L1')
ws_dash['A1'].font = Font(bold=True, size=20, color="1F4E79")
ws_dash['A1'].alignment = center_align
ws_dash['A1'].fill = PatternFill("solid", fgColor="E8F4FD")

dash_header_fill = PatternFill("solid", fgColor="1F4E79")
ws_dash['A3'] = "KEY PERFORMANCE INDICATORS"
ws_dash['A3'].font = Font(bold=True, size=12, color="FFFFFF")
ws_dash['A3'].fill = dash_header_fill
ws_dash.merge_cells('A3:L3')

# KPIs Row 1
ws_dash['A5'] = "Total Revenue"
ws_dash['B5'] = "=SUM('Income Tracker'!J:J)"
ws_dash['C5'] = "Total Expenses"
ws_dash['D5'] = "=SUM('Expense Tracker'!I:I)"
ws_dash['E5'] = "Net Profit"
ws_dash['F5'] = "=B5-D5"
ws_dash['G5'] = "Profit Margin"
ws_dash['H5'] = "=IF(B5>0,F5/B5,0)"

for col in ['A', 'C', 'E', 'G']:
    ws_dash[f'{col}5'].font = Font(bold=True, size=10, color="666666")
for col in ['B', 'D', 'F']:
    ws_dash[f'{col}5'].font = Font(bold=True, size=14, color="1F4E79")
    ws_dash[f'{col}5'].number_format = currency_format
ws_dash['H5'].font = Font(bold=True, size=14, color="1F4E79")
ws_dash['H5'].number_format = '0.0%'
ws_dash['F5'].font = Font(bold=True, size=14, color="008000")

# KPIs Row 2
ws_dash['A7'] = "GST Collected"
ws_dash['B7'] = "=SUM('Income Tracker'!I:I)"
ws_dash['C7'] = "GST Paid"
ws_dash['D7'] = "=SUM('Expense Tracker'!H:H)+SUM('Stock Purchases'!M:M)"
ws_dash['E7'] = "Net GST Payable"
ws_dash['F7'] = "=B7-D7"
ws_dash['G7'] = "Total Assets"
ws_dash['H7'] = "=SUM('Machines & Equipment'!I:I)"

for col in ['A', 'C', 'E', 'G']:
    ws_dash[f'{col}7'].font = Font(bold=True, size=10, color="666666")
for col in ['B', 'D', 'F', 'H']:
    ws_dash[f'{col}7'].font = Font(bold=True, size=14, color="1F4E79")
    ws_dash[f'{col}7'].number_format = currency_format

# Charts Data
ws_dash['A10'] = "CHARTS DATA"
ws_dash['A10'].font = Font(bold=True, size=12, color="FFFFFF")
ws_dash['A10'].fill = dash_header_fill
ws_dash.merge_cells('A10:D10')

ws_dash['A11'] = "Category"
ws_dash['B11'] = "Amount"
ws_dash['A12'] = "Revenue"
ws_dash['B12'] = "=SUM('Income Tracker'!J:J)"
ws_dash['A13'] = "Expenses"
ws_dash['B13'] = "=SUM('Expense Tracker'!I:I)"

ws_dash['D11'] = "Service"
ws_dash['E11'] = "Revenue"
for i, svc in enumerate(services, start=12):
    ws_dash.cell(row=i, column=4).value = svc
    ws_dash.cell(row=i, column=5).value = f'=SUMIF(\'Income Tracker\'!D:D,D{i},\'Income Tracker\'!J:J)'

ws_dash['G11'] = "Expense Category"
ws_dash['H11'] = "Amount"
for i, cat in enumerate(expense_cats, start=12):
    ws_dash.cell(row=i, column=7).value = cat
    ws_dash.cell(row=i, column=8).value = f'=SUMIF(\'Expense Tracker\'!C:C,G{i},\'Expense Tracker\'!I:I)'

# Revenue vs Expenses Chart
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Revenue vs Expenses"
data1 = Reference(ws_dash, min_col=2, min_row=11, max_row=13)
cats1 = Reference(ws_dash, min_col=1, min_row=12, max_row=13)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.width = 12
chart1.height = 8
ws_dash.add_chart(chart1, "A21")

# Service Revenue Pie Chart
chart2 = PieChart()
chart2.title = "Revenue by Service"
data2 = Reference(ws_dash, min_col=5, min_row=11, max_row=19)
cats2 = Reference(ws_dash, min_col=4, min_row=12, max_row=19)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.dataLabels = DataLabelList()
chart2.dataLabels.showPercent = True
chart2.width = 12
chart2.height = 8
ws_dash.add_chart(chart2, "I21")

# Expense Category Doughnut Chart
chart3 = DoughnutChart()
chart3.title = "Expenses by Category"
data3 = Reference(ws_dash, min_col=8, min_row=11, max_row=17)
cats3 = Reference(ws_dash, min_col=7, min_row=12, max_row=17)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.dataLabels = DataLabelList()
chart3.dataLabels.showPercent = True
chart3.width = 12
chart3.height = 8
ws_dash.add_chart(chart3, "A38")

for col in range(1, 13):
    ws_dash.column_dimensions[get_column_letter(col)].width = 14

ws_dash.freeze_panes = 'A3'

# Save
wb.save('/var/lib/freelancer/projects/40182876/MIS_System_v3.xlsx')
print("MIS System v3 with GST created successfully!")
