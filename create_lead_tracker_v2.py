#!/usr/bin/env python3
"""
Home Services Lead Tracker — Excel Generator
Creates a comprehensive lead tracking workbook with:
  - HOW TO USE instructions sheet
  - Lead Tracker (main data sheet with formulas, dropdowns, conditional formatting)
  - Dashboard (analytics with charts)
  - Form Fields Reference (Google Form setup guide)
"""

import os
from datetime import datetime, date
from copy import copy

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── Color Constants ──────────────────────────────────────────────────────────
DARK_BLUE = "1B2A4A"
MEDIUM_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
GREEN_BG = "C6EFCE"
GREEN_FONT = "006100"
YELLOW_BG = "FFEB9C"
YELLOW_FONT = "9C6500"
RED_BG = "FFC7CE"
RED_FONT = "9C0006"
BLUE_BG = "B4C6E7"
BLUE_FONT = "1F4E79"
COMPLETED_BG = "548235"
COMPLETED_FONT_COLOR = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
BORDER_COLOR = "B4B4B4"

# ── Reusable style helpers ───────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

medium_blue_border = Border(
    left=Side(style="medium", color=MEDIUM_BLUE),
    right=Side(style="medium", color=MEDIUM_BLUE),
    top=Side(style="medium", color=MEDIUM_BLUE),
    bottom=Side(style="medium", color=MEDIUM_BLUE),
)

def make_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def make_font(color=None, bold=False, size=10, name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color)

center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
left_top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── Dropdown lists ───────────────────────────────────────────────────────────
CITIES = "Bangalore,Mumbai,Delhi,Hyderabad,Chennai,Pune,Other"
SERVICES = (
    "Bathroom Cleaning,Kitchen Cleaning,Full Home Cleaning,"
    "Rental Property Cleaning,Ready to Move In Cleaning,"
    "Painting,Pest Control,Plumbing,Electrician,Other"
)
TIME_SLOTS = (
    "Morning (8AM-10AM),Morning (10AM-12PM),"
    "Afternoon (12PM-2PM),Afternoon (2PM-4PM),"
    "Evening (4PM-6PM),Evening (6PM-8PM)"
)
ORDER_STATUSES = "Confirmed,Pending,Cancelled,Scheduled,Completed"
ADVANCE_STATUSES = "Received,NIL,Cleared"
PAYMENT_STATUSES = "Received,Pending"
PAYMENT_MODES = "Cash,UPI,Debit Card,Payment Gateway,Bank Transfer,Other"

SERVICES_LIST = [s.strip() for s in SERVICES.split(",")]
TIME_SLOTS_LIST = [s.strip() for s in TIME_SLOTS.split(",")]

# ── Column widths for Lead Tracker ───────────────────────────────────────────
LT_COL_WIDTHS = {
    "A": 6, "B": 18, "C": 12, "D": 20, "E": 16, "F": 16, "G": 22,
    "H": 14, "I": 18, "J": 30, "K": 20, "L": 14, "M": 20, "N": 14,
    "O": 20, "P": 14, "Q": 20, "R": 14, "S": 16, "T": 16, "U": 22,
    "V": 16, "W": 25, "X": 18, "Y": 16, "Z": 18, "AA": 18, "AB": 22,
    "AC": 18, "AD": 16, "AE": 16, "AF": 18, "AG": 16, "AH": 16,
    "AI": 16, "AJ": 20,
}

HEADERS = [
    "S.No", "Timestamp", "Order ID", "Customer Name", "Phone Number",
    "Alternate Phone", "Email", "City", "Area / Locality", "Full Address",
    "Service 1", "Price 1", "Service 2", "Price 2", "Service 3", "Price 3",
    "Service 4", "Price 4", "Total Value", "Preferred Date", "Slot Time",
    "Order Status", "Reason / Notes", "Vendor Name", "Vendor Contact",
    "Vendor Alternate No.", "Order Scheduled Date", "Scheduled Time",
    "Order Completed Date", "Advance Amount", "Advance Status",
    "Payment Value (Quoted)", "Payment Status", "Payment Mode",
    "Invoice Number", "Transaction Ref. No.",
]

CURRENCY_FMT = '₹#,##0.00'
DATE_FMT = 'DD-MMM-YYYY'
DATETIME_FMT = 'DD-MMM-YYYY HH:MM'

# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 0 — HOW TO USE
# ═════════════════════════════════════════════════════════════════════════════
def create_how_to_use(wb):
    ws = wb.active
    ws.title = "HOW TO USE"
    ws.sheet_properties.tabColor = DARK_BLUE

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100

    # Title
    cell = ws["B2"]
    cell.value = "HOW TO OPEN & USE THIS FILE"
    cell.font = Font(name="Calibri", size=20, bold=True, color=DARK_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 45

    # Separator
    ws.row_dimensions[3].height = 8

    instructions = [
        ("STEP 1 — OPEN IN GOOGLE SHEETS", MEDIUM_BLUE, True, 14, [
            "1. Upload this .xlsx file to your Google Drive.",
            "2. Double-click the file in Google Drive — it will open in Google Sheets.",
            "3. All formulas, dropdowns, and formatting will work automatically.",
        ]),
        ("STEP 2 — ENABLE FILTERS", MEDIUM_BLUE, True, 14, [
            "1. Go to the 'Lead Tracker' tab.",
            "2. Click on Data → Create a filter.",
            "3. You can now filter/sort by any column (Status, City, Date, etc.).",
        ]),
        ("STEP 3 — COLOR CODES", MEDIUM_BLUE, True, 14, [
            "• Green background = Confirmed order",
            "• Dark Green with white text = Completed order",
            "• Yellow background = Pending order",
            "• Red background = Cancelled order",
            "• Blue background = Scheduled order",
        ]),
        ("STEP 4 — HOW TO USE THE LEAD TRACKER", MEDIUM_BLUE, True, 14, [
            "• Google Form responses auto-fill columns B through M (Timestamp, Name, Phone, etc.).",
            "• Order ID (column C) is auto-generated in format ST-XXXX.",
            "• S.No (column A) is auto-generated.",
            "• Manually update: Order Status, Vendor details, Payment info.",
            "• Total Value (column S) auto-calculates from Price 1-4.",
        ]),
        ("STEP 4B — CALENDAR & TIME PICKERS", MEDIUM_BLUE, True, 14, [
            "• Date columns (Preferred Date, Scheduled Date, Completed Date) show a calendar picker in Google Sheets.",
            "• Time Slot columns have dropdown menus: Morning, Afternoon, Evening slots.",
            "• Simply click on a date cell and a calendar will appear (in Google Sheets).",
        ]),
        ("STEP 5 — DASHBOARD TAB", MEDIUM_BLUE, True, 14, [
            "• The Dashboard tab auto-calculates stats from the Lead Tracker.",
            "• Order counts, payment summaries, service breakdown, area distribution.",
            "• Date-wise report: enter From/To dates and see filtered stats.",
            "• Charts update automatically.",
        ]),
        ("STEP 6 — LINK GOOGLE FORM", MEDIUM_BLUE, True, 14, [
            "• See the 'Form Fields Reference' tab for field setup details.",
            "• Create a Google Form with matching fields.",
            "• Link form responses to the 'Lead Tracker' sheet.",
            "• Form responses will fill columns B onward; formulas handle the rest.",
        ]),
    ]

    row = 4
    for title, color, bold, size, lines in instructions:
        # Section title
        cell = ws.cell(row=row, column=2)
        cell.value = title
        cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 30
        row += 1

        for line in lines:
            cell = ws.cell(row=row, column=2)
            cell.value = line
            cell.font = Font(name="Calibri", size=11, color="333333")
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            ws.row_dimensions[row].height = 22
            row += 1

        row += 1  # spacer


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — LEAD TRACKER
# ═════════════════════════════════════════════════════════════════════════════
def create_lead_tracker(wb):
    ws = wb.create_sheet("Lead Tracker")
    ws.sheet_properties.tabColor = MEDIUM_BLUE

    # ── Column widths ────────────────────────────────────────────────────
    for col_letter, width in LT_COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    last_col = "AJ"
    last_col_idx = 36  # AJ = column 36

    # ── Row 1: Title ─────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = "HOME SERVICES — LEAD TRACKER & ORDER MANAGEMENT"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    title_cell.fill = make_fill(DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 45
    # Fill merged range background
    for c in range(2, last_col_idx + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = make_fill(DARK_BLUE)

    # ── Row 2: Subtitle ──────────────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col}2")
    sub_cell = ws["A2"]
    sub_cell.value = "Auto-linked with Google Form | Order ID Format: ST-XXXX | Use Filters for Quick Search"
    sub_cell.font = Font(name="Calibri", size=11, bold=False, color=WHITE)
    sub_cell.fill = make_fill(MEDIUM_BLUE)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28
    for c in range(2, last_col_idx + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = make_fill(MEDIUM_BLUE)

    # ── Row 3: Headers ───────────────────────────────────────────────────
    ws.row_dimensions[3].height = 40
    for idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=idx)
        cell.value = header
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = make_fill(MEDIUM_BLUE)
        cell.alignment = center_wrap
        cell.border = thin_border

    # ── Data validations ─────────────────────────────────────────────────
    dv_city = DataValidation(type="list", formula1=f'"{CITIES}"', allow_blank=True)
    dv_city.error = "Please select a valid city"
    dv_city.errorTitle = "Invalid City"
    ws.add_data_validation(dv_city)
    dv_city.add("H4:H1000")

    dv_service = DataValidation(type="list", formula1=f'"{SERVICES}"', allow_blank=True)
    dv_service.error = "Please select a valid service"
    ws.add_data_validation(dv_service)
    for col in ["K", "M", "O", "Q"]:
        dv_service.add(f"{col}4:{col}1000")

    dv_time = DataValidation(type="list", formula1=f'"{TIME_SLOTS}"', allow_blank=True)
    ws.add_data_validation(dv_time)
    dv_time.add("U4:U1000")
    dv_time.add("AB4:AB1000")

    dv_status = DataValidation(type="list", formula1=f'"{ORDER_STATUSES}"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("V4:V1000")

    dv_adv = DataValidation(type="list", formula1=f'"{ADVANCE_STATUSES}"', allow_blank=True)
    ws.add_data_validation(dv_adv)
    dv_adv.add("AE4:AE1000")

    dv_pay_status = DataValidation(type="list", formula1=f'"{PAYMENT_STATUSES}"', allow_blank=True)
    ws.add_data_validation(dv_pay_status)
    dv_pay_status.add("AG4:AG1000")

    dv_pay_mode = DataValidation(type="list", formula1=f'"{PAYMENT_MODES}"', allow_blank=True)
    ws.add_data_validation(dv_pay_mode)
    dv_pay_mode.add("AH4:AH1000")

    # Date validations for calendar picker
    dv_date = DataValidation(type="date", allow_blank=True)
    dv_date.error = "Please enter a valid date"
    ws.add_data_validation(dv_date)
    dv_date.add("T4:T1000")
    dv_date.add("AA4:AA1000")
    dv_date.add("AC4:AC1000")

    # ── Pre-format rows 4-100 with formulas, borders, alignment ──────────
    # Currency columns: L(12), N(14), P(16), R(18), S(19), AD(30), AF(32)
    currency_cols = [12, 14, 16, 18, 19, 30, 32]
    # Date columns: T(20), AA(27), AC(29)
    date_cols = [20, 27, 29]
    # Timestamp column: B(2)
    timestamp_col = 2

    for row in range(4, 101):
        for col in range(1, last_col_idx + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = center_wrap
            cell.border = thin_border

            if col in currency_cols:
                cell.number_format = CURRENCY_FMT
            elif col in date_cols:
                cell.number_format = DATE_FMT
            elif col == timestamp_col:
                cell.number_format = DATETIME_FMT

        # Column A: S.No formula
        ws.cell(row=row, column=1).value = f'=IF(B{row}<>"",ROW()-3,"")'
        # Column C: Order ID formula
        c_cell = ws.cell(row=row, column=3)
        c_cell.value = f'=IF(B{row}<>"","ST-"&TEXT(ROW()-3,"0000"),"")'
        c_cell.font = Font(name="Calibri", size=10, bold=True, color=MEDIUM_BLUE)
        # Column S: Total Value formula
        ws.cell(row=row, column=19).value = (
            f'=IF(OR(L{row}<>"",N{row}<>"",P{row}<>"",R{row}<>""),'
            f'SUM(L{row},N{row},P{row},R{row}),"")'
        )

    # ── Conditional Formatting ───────────────────────────────────────────

    # Order Status (column V) cell-level formatting
    status_range = "V4:V1000"
    ws.conditional_formatting.add(status_range, CellIsRule(
        operator="equal", formula=['"Confirmed"'],
        fill=make_fill(GREEN_BG),
        font=Font(bold=True, color=GREEN_FONT)
    ))
    ws.conditional_formatting.add(status_range, CellIsRule(
        operator="equal", formula=['"Pending"'],
        fill=make_fill(YELLOW_BG),
        font=Font(bold=True, color=YELLOW_FONT)
    ))
    ws.conditional_formatting.add(status_range, CellIsRule(
        operator="equal", formula=['"Cancelled"'],
        fill=make_fill(RED_BG),
        font=Font(bold=True, color=RED_FONT)
    ))
    ws.conditional_formatting.add(status_range, CellIsRule(
        operator="equal", formula=['"Scheduled"'],
        fill=make_fill(BLUE_BG),
        font=Font(bold=True, color=BLUE_FONT)
    ))
    ws.conditional_formatting.add(status_range, CellIsRule(
        operator="equal", formula=['"Completed"'],
        fill=make_fill(COMPLETED_BG),
        font=Font(bold=True, color=COMPLETED_FONT_COLOR)
    ))

    # Payment Status (column AG) conditional formatting
    pay_range = "AG4:AG1000"
    ws.conditional_formatting.add(pay_range, CellIsRule(
        operator="equal", formula=['"Received"'],
        fill=make_fill(GREEN_BG),
        font=Font(bold=True, color=GREEN_FONT)
    ))
    ws.conditional_formatting.add(pay_range, CellIsRule(
        operator="equal", formula=['"Pending"'],
        fill=make_fill(RED_BG),
        font=Font(bold=True, color=RED_FONT)
    ))

    # Advance Status (column AE) conditional formatting
    adv_range = "AE4:AE1000"
    ws.conditional_formatting.add(adv_range, CellIsRule(
        operator="equal", formula=['"Received"'],
        fill=make_fill(GREEN_BG),
        font=Font(bold=True, color=GREEN_FONT)
    ))
    ws.conditional_formatting.add(adv_range, CellIsRule(
        operator="equal", formula=['"NIL"'],
        fill=make_fill(YELLOW_BG),
        font=Font(bold=True, color=YELLOW_FONT)
    ))
    ws.conditional_formatting.add(adv_range, CellIsRule(
        operator="equal", formula=['"Cleared"'],
        fill=make_fill("A9D18E"),
        font=Font(bold=True, color="375623")
    ))

    # Row-level conditional formatting based on Order Status (column V)
    row_range = f"A4:{last_col}1000"
    ws.conditional_formatting.add(row_range, FormulaRule(
        formula=['$V4="Confirmed"'],
        fill=make_fill("E2EFDA")
    ))
    ws.conditional_formatting.add(row_range, FormulaRule(
        formula=['$V4="Pending"'],
        fill=make_fill("FFF2CC")
    ))
    ws.conditional_formatting.add(row_range, FormulaRule(
        formula=['$V4="Cancelled"'],
        fill=make_fill("FCE4EC")
    ))
    ws.conditional_formatting.add(row_range, FormulaRule(
        formula=['$V4="Scheduled"'],
        fill=make_fill("DAEEF3")
    ))
    ws.conditional_formatting.add(row_range, FormulaRule(
        formula=['$V4="Completed"'],
        fill=make_fill("E8D5F5")
    ))

    # ── Sample Data (rows 4-8) ───────────────────────────────────────────
    sample_data = [
        {  # Row 4 - Rajesh Kumar - Confirmed
            "B": datetime(2026, 1, 25, 9, 15),
            "D": "Rajesh Kumar", "E": "9876543210", "F": "", "G": "rajesh@email.com",
            "H": "Bangalore", "I": "Indiranagar", "J": "123 Main Road, Indiranagar, Bangalore",
            "K": "Bathroom Cleaning", "L": 2000,
            "M": "Pest Control", "N": 3500,
            "T": date(2026, 2, 1), "U": "Morning (10AM-12PM)",
            "V": "Confirmed", "W": "",
            "X": "Suresh", "Y": "9988776655", "Z": "",
            "AA": date(2026, 2, 1), "AB": "Morning (10AM-12PM)",
            "AD": 500, "AE": "Received",
            "AF": 5500, "AG": "Pending", "AH": "", "AI": "", "AJ": "",
        },
        {  # Row 5 - Priya Sharma - Pending
            "B": datetime(2026, 1, 25, 11, 45),
            "D": "Priya Sharma", "E": "8765432109", "F": "", "G": "priya@email.com",
            "H": "Bangalore", "I": "HSR Layout", "J": "456 2nd Cross, HSR Layout, Bangalore",
            "K": "Pest Control", "L": 3500,
            "T": date(2026, 2, 3), "U": "Afternoon (2PM-4PM)",
            "V": "Pending", "W": "Waiting for confirmation",
        },
        {  # Row 6 - Amit Patel - Cancelled
            "B": datetime(2026, 1, 26, 14, 20),
            "D": "Amit Patel", "E": "7654321098", "F": "", "G": "amit@email.com",
            "H": "Bangalore", "I": "Whitefield", "J": "789 Tech Park Road, Whitefield, Bangalore",
            "K": "Painting", "L": 15000,
            "M": "Full Home Cleaning", "N": 4500,
            "T": date(2026, 2, 5), "U": "Morning (8AM-10AM)",
            "V": "Cancelled", "W": "Price too high",
        },
        {  # Row 7 - Meera Reddy - Scheduled
            "B": datetime(2026, 1, 26, 16, 0),
            "D": "Meera Reddy", "E": "6543210987", "F": "", "G": "meera@email.com",
            "H": "Bangalore", "I": "Koramangala", "J": "321 5th Block, Koramangala, Bangalore",
            "K": "Full Home Cleaning", "L": 4500,
            "T": date(2026, 2, 7), "U": "Morning (10AM-12PM)",
            "V": "Scheduled", "W": "",
            "X": "Ramesh", "Y": "9876501234", "Z": "",
            "AA": date(2026, 2, 7), "AB": "Morning (10AM-12PM)",
            "AD": 1000, "AE": "Received",
            "AF": 4500, "AG": "Pending",
        },
        {  # Row 8 - Karthik Nair - Completed
            "B": datetime(2026, 1, 20, 8, 30),
            "D": "Karthik Nair", "E": "5432109876", "F": "", "G": "karthik@email.com",
            "H": "Bangalore", "I": "JP Nagar", "J": "654 6th Phase, JP Nagar, Bangalore",
            "K": "Plumbing", "L": 1500,
            "T": date(2026, 1, 22), "U": "Evening (4PM-6PM)",
            "V": "Completed", "W": "",
            "X": "Vijay", "Y": "9123456780", "Z": "",
            "AA": date(2026, 1, 22), "AB": "Evening (4PM-6PM)",
            "AC": date(2026, 1, 22),
            "AD": 0, "AE": "NIL",
            "AF": 1500, "AG": "Received", "AH": "Cash",
            "AI": "INV-0005", "AJ": "TXN-CASH-0005",
        },
    ]

    col_map = {}
    for i in range(1, last_col_idx + 1):
        col_map[get_column_letter(i)] = i

    for data_idx, data in enumerate(sample_data):
        row = 4 + data_idx
        for key, value in data.items():
            col_idx = col_map[key]
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            # Preserve formatting set in pre-format loop
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = center_wrap
            cell.border = thin_border
            if col_idx in currency_cols:
                cell.number_format = CURRENCY_FMT
            elif col_idx in date_cols:
                cell.number_format = DATE_FMT
            elif col_idx == timestamp_col:
                cell.number_format = DATETIME_FMT

        # Re-apply formula cells (A, C, S) so they aren't overwritten
        ws.cell(row=row, column=1).value = f'=IF(B{row}<>"",ROW()-3,"")'
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=row, column=1).alignment = center_wrap
        ws.cell(row=row, column=1).border = thin_border

        c_cell = ws.cell(row=row, column=3)
        c_cell.value = f'=IF(B{row}<>"","ST-"&TEXT(ROW()-3,"0000"),"")'
        c_cell.font = Font(name="Calibri", size=10, bold=True, color=MEDIUM_BLUE)
        c_cell.alignment = center_wrap
        c_cell.border = thin_border

        s_cell = ws.cell(row=row, column=19)
        s_cell.value = (
            f'=IF(OR(L{row}<>"",N{row}<>"",P{row}<>"",R{row}<>""),'
            f'SUM(L{row},N{row},P{row},R{row}),"")'
        )
        s_cell.number_format = CURRENCY_FMT
        s_cell.alignment = center_wrap
        s_cell.border = thin_border

    # ── Freeze panes & auto filter ───────────────────────────────────────
    ws.freeze_panes = "E4"
    ws.auto_filter.ref = f"A3:{last_col}1000"

    return ws


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════
def create_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = DARK_BLUE

    # Column widths
    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 18

    # ── Row 1: Title ─────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value = "DASHBOARD — BUSINESS ANALYTICS & STATS"
    t.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    t.fill = make_fill(DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 45
    for c in range(2, 13):
        cell = ws.cell(row=1, column=c)
        cell.fill = make_fill(DARK_BLUE)

    # ── Row 2: Subtitle ──────────────────────────────────────────────────
    ws.merge_cells("A2:L2")
    s = ws["A2"]
    s.value = "All stats auto-calculated from Lead Tracker sheet"
    s.font = Font(name="Calibri", size=11, color=WHITE)
    s.fill = make_fill(MEDIUM_BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28
    for c in range(2, 13):
        cell = ws.cell(row=2, column=c)
        cell.fill = make_fill(MEDIUM_BLUE)

    # ── Row 4-5: Color Legend ────────────────────────────────────────────
    ws.merge_cells("A4:E4")
    lg = ws["A4"]
    lg.value = "STATUS COLOR LEGEND"
    lg.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    lg.fill = make_fill(MEDIUM_BLUE)
    lg.alignment = center_wrap
    ws.row_dimensions[4].height = 28
    for c in range(2, 6):
        ws.cell(row=4, column=c).fill = make_fill(MEDIUM_BLUE)

    legend_items = [
        ("Confirmed", GREEN_BG, GREEN_FONT),
        ("Pending", YELLOW_BG, YELLOW_FONT),
        ("Cancelled", RED_BG, RED_FONT),
        ("Scheduled", BLUE_BG, BLUE_FONT),
        ("Completed", COMPLETED_BG, COMPLETED_FONT_COLOR),
    ]
    for i, (label, bg, fg) in enumerate(legend_items):
        cell = ws.cell(row=5, column=i + 1)
        cell.value = label
        cell.font = Font(name="Calibri", size=11, bold=True, color=fg)
        cell.fill = make_fill(bg)
        cell.alignment = center_wrap
        cell.border = thin_border
    ws.row_dimensions[5].height = 30

    # ── Section 1: ORDER STATISTICS (Row 7) ──────────────────────────────
    R_OS = 7
    ws.merge_cells(f"A{R_OS}:F{R_OS}")
    h = ws.cell(row=R_OS, column=1)
    h.value = "ORDER STATISTICS"
    h.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    h.fill = make_fill(MEDIUM_BLUE)
    h.alignment = center_wrap
    for c in range(2, 7):
        ws.cell(row=R_OS, column=c).fill = make_fill(MEDIUM_BLUE)
    ws.row_dimensions[R_OS].height = 30

    # Row 8: column headers
    os_headers = ["Total Orders", "Confirmed", "Pending", "Cancelled", "Scheduled", "Completed"]
    ws.row_dimensions[8].height = 28
    for i, hdr in enumerate(os_headers):
        cell = ws.cell(row=8, column=i + 1)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = make_fill(DARK_BLUE)
        cell.alignment = center_wrap
        cell.border = thin_border

    # Row 9: formulas (big numbers)
    os_formulas = [
        "=COUNTA('Lead Tracker'!C4:C1000)-COUNTBLANK('Lead Tracker'!C4:C1000)",
        '=COUNTIF(\'Lead Tracker\'!V4:V1000,"Confirmed")',
        '=COUNTIF(\'Lead Tracker\'!V4:V1000,"Pending")',
        '=COUNTIF(\'Lead Tracker\'!V4:V1000,"Cancelled")',
        '=COUNTIF(\'Lead Tracker\'!V4:V1000,"Scheduled")',
        '=COUNTIF(\'Lead Tracker\'!V4:V1000,"Completed")',
    ]
    os_colors = [LIGHT_BLUE, GREEN_BG, YELLOW_BG, RED_BG, BLUE_BG, COMPLETED_BG]
    os_font_colors = [DARK_BLUE, GREEN_FONT, YELLOW_FONT, RED_FONT, BLUE_FONT, COMPLETED_FONT_COLOR]
    ws.row_dimensions[9].height = 50
    for i, (formula, bg, fg) in enumerate(zip(os_formulas, os_colors, os_font_colors)):
        cell = ws.cell(row=9, column=i + 1)
        cell.value = formula
        cell.font = Font(name="Calibri", size=22, bold=True, color=fg)
        cell.fill = make_fill(bg)
        cell.alignment = center_wrap
        cell.border = thin_border

    # ── Section 2: PAYMENT SUMMARY (Row 11) ──────────────────────────────
    R_PS = 11
    ws.merge_cells(f"A{R_PS}:F{R_PS}")
    h = ws.cell(row=R_PS, column=1)
    h.value = "PAYMENT SUMMARY"
    h.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    h.fill = make_fill(MEDIUM_BLUE)
    h.alignment = center_wrap
    for c in range(2, 7):
        ws.cell(row=R_PS, column=c).fill = make_fill(MEDIUM_BLUE)
    ws.row_dimensions[R_PS].height = 30

    # Row 12: headers
    pay_headers = [
        "Total Quoted", "Total Received", "Total Pending",
        "Cash Received", "UPI Received", "Card/Gateway"
    ]
    ws.row_dimensions[12].height = 28
    for i, hdr in enumerate(pay_headers):
        cell = ws.cell(row=12, column=i + 1)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = make_fill(DARK_BLUE)
        cell.alignment = center_wrap
        cell.border = thin_border

    # Row 13: formulas
    pay_formulas = [
        "=SUMPRODUCT(('Lead Tracker'!S4:S1000)*1)",
        '=SUMPRODUCT((\'Lead Tracker\'!AG4:AG1000="Received")*(\'Lead Tracker\'!AF4:AF1000))',
        '=SUMPRODUCT((\'Lead Tracker\'!AG4:AG1000="Pending")*(\'Lead Tracker\'!AF4:AF1000))',
        '=SUMPRODUCT((\'Lead Tracker\'!AG4:AG1000="Received")*(\'Lead Tracker\'!AH4:AH1000="Cash")*(\'Lead Tracker\'!AF4:AF1000))',
        '=SUMPRODUCT((\'Lead Tracker\'!AG4:AG1000="Received")*(\'Lead Tracker\'!AH4:AH1000="UPI")*(\'Lead Tracker\'!AF4:AF1000))',
        '=SUMPRODUCT((\'Lead Tracker\'!AG4:AG1000="Received")*(\'Lead Tracker\'!AH4:AH1000="Debit Card")*(\'Lead Tracker\'!AF4:AF1000))+SUMPRODUCT((\'Lead Tracker\'!AG4:AG1000="Received")*(\'Lead Tracker\'!AH4:AH1000="Payment Gateway")*(\'Lead Tracker\'!AF4:AF1000))',
    ]
    ws.row_dimensions[13].height = 50
    for i, formula in enumerate(pay_formulas):
        cell = ws.cell(row=13, column=i + 1)
        cell.value = formula
        cell.font = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
        cell.fill = make_fill(LIGHT_BLUE)
        cell.alignment = center_wrap
        cell.border = thin_border
        cell.number_format = '₹#,##0'

    # ── Section 3: SERVICE-WISE BREAKDOWN (Row 15) ───────────────────────
    R_SB = 15
    ws.merge_cells(f"A{R_SB}:H{R_SB}")
    h = ws.cell(row=R_SB, column=1)
    h.value = "SERVICE-WISE BREAKDOWN"
    h.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    h.fill = make_fill(MEDIUM_BLUE)
    h.alignment = center_wrap
    for c in range(2, 9):
        ws.cell(row=R_SB, column=c).fill = make_fill(MEDIUM_BLUE)
    ws.row_dimensions[R_SB].height = 30

    # Row 16: table headers
    svc_headers = [
        "Service", "Total Orders", "Confirmed", "Scheduled",
        "Completed", "Pending", "Cancelled", "Revenue"
    ]
    ws.row_dimensions[16].height = 28
    for i, hdr in enumerate(svc_headers):
        cell = ws.cell(row=16, column=i + 1)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = make_fill(DARK_BLUE)
        cell.alignment = center_wrap
        cell.border = thin_border

    # Rows 17-26: service list
    for s_idx, service in enumerate(SERVICES_LIST):
        row = 17 + s_idx
        # Alternating row color
        bg = LIGHT_GRAY if s_idx % 2 == 1 else WHITE
        fill = make_fill(bg)

        # Col A: Service name
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = service
        cell_a.font = Font(name="Calibri", size=10)
        cell_a.alignment = left_wrap
        cell_a.border = thin_border
        cell_a.fill = fill

        # Col B: Total Orders (count across K, M, O, Q)
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = (
            f"=COUNTIF('Lead Tracker'!K4:K1000,A{row})"
            f"+COUNTIF('Lead Tracker'!M4:M1000,A{row})"
            f"+COUNTIF('Lead Tracker'!O4:O1000,A{row})"
            f"+COUNTIF('Lead Tracker'!Q4:Q1000,A{row})"
        )
        cell_b.font = Font(name="Calibri", size=10)
        cell_b.alignment = center_wrap
        cell_b.border = thin_border
        cell_b.fill = fill

        # Cols C-G: status-wise counts
        statuses_for_svc = ["Confirmed", "Scheduled", "Completed", "Pending", "Cancelled"]
        for st_idx, status in enumerate(statuses_for_svc):
            col = 3 + st_idx
            cell = ws.cell(row=row, column=col)
            cell.value = (
                f"=COUNTIFS('Lead Tracker'!K4:K1000,A{row},'Lead Tracker'!V4:V1000,\"{status}\")"
                f"+COUNTIFS('Lead Tracker'!M4:M1000,A{row},'Lead Tracker'!V4:V1000,\"{status}\")"
                f"+COUNTIFS('Lead Tracker'!O4:O1000,A{row},'Lead Tracker'!V4:V1000,\"{status}\")"
                f"+COUNTIFS('Lead Tracker'!Q4:Q1000,A{row},'Lead Tracker'!V4:V1000,\"{status}\")"
            )
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = center_wrap
            cell.border = thin_border
            cell.fill = fill

        # Col H: Revenue
        cell_h = ws.cell(row=row, column=8)
        cell_h.value = (
            f"=SUMPRODUCT(('Lead Tracker'!K4:K1000=A{row})*('Lead Tracker'!AG4:AG1000=\"Received\")*('Lead Tracker'!L4:L1000))"
            f"+SUMPRODUCT(('Lead Tracker'!M4:M1000=A{row})*('Lead Tracker'!AG4:AG1000=\"Received\")*('Lead Tracker'!N4:N1000))"
            f"+SUMPRODUCT(('Lead Tracker'!O4:O1000=A{row})*('Lead Tracker'!AG4:AG1000=\"Received\")*('Lead Tracker'!P4:P1000))"
            f"+SUMPRODUCT(('Lead Tracker'!Q4:Q1000=A{row})*('Lead Tracker'!AG4:AG1000=\"Received\")*('Lead Tracker'!R4:R1000))"
        )
        cell_h.font = Font(name="Calibri", size=10)
        cell_h.alignment = center_wrap
        cell_h.border = thin_border
        cell_h.fill = fill
        cell_h.number_format = '₹#,##0'

    # Row 27: TOTAL row
    total_row = 27
    ws.row_dimensions[total_row].height = 28
    cell_a = ws.cell(row=total_row, column=1)
    cell_a.value = "TOTAL"
    cell_a.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    cell_a.fill = make_fill(DARK_BLUE)
    cell_a.alignment = center_wrap
    cell_a.border = thin_border

    for col in range(2, 9):
        cell = ws.cell(row=total_row, column=col)
        col_letter = get_column_letter(col)
        cell.value = f"=SUM({col_letter}17:{col_letter}26)"
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = make_fill(DARK_BLUE)
        cell.alignment = center_wrap
        cell.border = thin_border
        if col == 8:
            cell.number_format = '₹#,##0'

    # ── Section 4: AREA-WISE BREAKDOWN ───────────────────────────────────
    area_start = 29  # 2 rows after service total (27)
    ws.merge_cells(f"A{area_start}:F{area_start}")
    h = ws.cell(row=area_start, column=1)
    h.value = "AREA-WISE ORDER DISTRIBUTION"
    h.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    h.fill = make_fill(MEDIUM_BLUE)
    h.alignment = center_wrap
    for c in range(2, 7):
        ws.cell(row=area_start, column=c).fill = make_fill(MEDIUM_BLUE)
    ws.row_dimensions[area_start].height = 30

    # Area table headers
    area_hdr_row = area_start + 1
    area_headers = ["Area", "Total Orders", "Confirmed", "Completed", "Pending", "Revenue"]
    ws.row_dimensions[area_hdr_row].height = 28
    for i, hdr in enumerate(area_headers):
        cell = ws.cell(row=area_hdr_row, column=i + 1)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = make_fill(DARK_BLUE)
        cell.alignment = center_wrap
        cell.border = thin_border

    areas = [
        "Indiranagar", "Koramangala", "HSR Layout", "Whitefield", "JP Nagar",
        "BTM Layout", "Electronic City", "MG Road", "Marathahalli",
        "Banashankari", "Other"
    ]
    area_data_start = area_hdr_row + 1
    for a_idx, area in enumerate(areas):
        row = area_data_start + a_idx
        bg = LIGHT_GRAY if a_idx % 2 == 1 else WHITE
        fill = make_fill(bg)

        # Col A: Area name
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = area
        cell_a.font = Font(name="Calibri", size=10)
        cell_a.alignment = left_wrap
        cell_a.border = thin_border
        cell_a.fill = fill

        # Col B: Total Orders
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = f"=COUNTIF('Lead Tracker'!I4:I1000,A{row})"
        cell_b.font = Font(name="Calibri", size=10)
        cell_b.alignment = center_wrap
        cell_b.border = thin_border
        cell_b.fill = fill

        # Col C: Confirmed
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = f'=COUNTIFS(\'Lead Tracker\'!I4:I1000,A{row},\'Lead Tracker\'!V4:V1000,"Confirmed")'
        cell_c.font = Font(name="Calibri", size=10)
        cell_c.alignment = center_wrap
        cell_c.border = thin_border
        cell_c.fill = fill

        # Col D: Completed
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = f'=COUNTIFS(\'Lead Tracker\'!I4:I1000,A{row},\'Lead Tracker\'!V4:V1000,"Completed")'
        cell_d.font = Font(name="Calibri", size=10)
        cell_d.alignment = center_wrap
        cell_d.border = thin_border
        cell_d.fill = fill

        # Col E: Pending
        cell_e = ws.cell(row=row, column=5)
        cell_e.value = f'=COUNTIFS(\'Lead Tracker\'!I4:I1000,A{row},\'Lead Tracker\'!V4:V1000,"Pending")'
        cell_e.font = Font(name="Calibri", size=10)
        cell_e.alignment = center_wrap
        cell_e.border = thin_border
        cell_e.fill = fill

        # Col F: Revenue
        cell_f = ws.cell(row=row, column=6)
        cell_f.value = (
            f"=SUMPRODUCT(('Lead Tracker'!I4:I1000=A{row})"
            f"*('Lead Tracker'!AG4:AG1000=\"Received\")"
            f"*('Lead Tracker'!AF4:AF1000))"
        )
        cell_f.font = Font(name="Calibri", size=10)
        cell_f.alignment = center_wrap
        cell_f.border = thin_border
        cell_f.fill = fill
        cell_f.number_format = '₹#,##0'

    # Area TOTAL row
    area_total_row = area_data_start + len(areas)
    ws.row_dimensions[area_total_row].height = 28
    cell_a = ws.cell(row=area_total_row, column=1)
    cell_a.value = "TOTAL"
    cell_a.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    cell_a.fill = make_fill(DARK_BLUE)
    cell_a.alignment = center_wrap
    cell_a.border = thin_border
    for col in range(2, 7):
        cell = ws.cell(row=area_total_row, column=col)
        cl = get_column_letter(col)
        cell.value = f"=SUM({cl}{area_data_start}:{cl}{area_total_row - 1})"
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = make_fill(DARK_BLUE)
        cell.alignment = center_wrap
        cell.border = thin_border
        if col == 6:
            cell.number_format = '₹#,##0'

    # ── Section 5: DATE-WISE REPORT ──────────────────────────────────────
    date_start = area_total_row + 2
    ws.merge_cells(f"A{date_start}:H{date_start}")
    h = ws.cell(row=date_start, column=1)
    h.value = "DATE-WISE REPORT (Enter dates below to filter)"
    h.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    h.fill = make_fill(MEDIUM_BLUE)
    h.alignment = center_wrap
    for c in range(2, 9):
        ws.cell(row=date_start, column=c).fill = make_fill(MEDIUM_BLUE)
    ws.row_dimensions[date_start].height = 30

    # From / To Date row
    fdr = date_start + 1  # from-date row
    ws.row_dimensions[fdr].height = 35

    cell_a = ws.cell(row=fdr, column=1)
    cell_a.value = "From Date:"
    cell_a.font = Font(name="Calibri", size=12, bold=True, color=MEDIUM_BLUE)
    cell_a.alignment = Alignment(horizontal="right", vertical="center")

    cell_b = ws.cell(row=fdr, column=2)
    cell_b.fill = make_fill(YELLOW_BG)
    cell_b.font = Font(name="Calibri", size=12, bold=True)
    cell_b.border = medium_blue_border
    cell_b.number_format = DATE_FMT
    cell_b.alignment = center_wrap

    cell_c = ws.cell(row=fdr, column=3)
    cell_c.value = "To Date:"
    cell_c.font = Font(name="Calibri", size=12, bold=True, color=MEDIUM_BLUE)
    cell_c.alignment = Alignment(horizontal="right", vertical="center")

    cell_d = ws.cell(row=fdr, column=4)
    cell_d.fill = make_fill(YELLOW_BG)
    cell_d.font = Font(name="Calibri", size=12, bold=True)
    cell_d.border = medium_blue_border
    cell_d.number_format = DATE_FMT
    cell_d.alignment = center_wrap

    cell_e = ws.cell(row=fdr, column=5)
    cell_e.value = "\u2190 Click the yellow cells, a calendar will appear in Google Sheets"
    cell_e.font = Font(name="Calibri", size=10, italic=True, color="666666")
    cell_e.alignment = Alignment(horizontal="left", vertical="center")

    # Date validation for calendar picker
    dv_from = DataValidation(type="date", allow_blank=True)
    ws.add_data_validation(dv_from)
    dv_from.add(f"B{fdr}")
    dv_to = DataValidation(type="date", allow_blank=True)
    ws.add_data_validation(dv_to)
    dv_to.add(f"D{fdr}")

    # Report headers row
    rpt_hdr_row = fdr + 1
    rpt_headers = [
        "Orders in Range", "Confirmed", "Completed", "Pending",
        "Cancelled", "Revenue in Range", "Payments Pending"
    ]
    ws.row_dimensions[rpt_hdr_row].height = 28
    for i, hdr in enumerate(rpt_headers):
        cell = ws.cell(row=rpt_hdr_row, column=i + 1)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = make_fill(DARK_BLUE)
        cell.alignment = center_wrap
        cell.border = thin_border

    # Report formula row
    rpt_data_row = rpt_hdr_row + 1
    ws.row_dimensions[rpt_data_row].height = 45

    date_formulas = [
        # Orders in Range
        (
            f'=IF(AND(B{fdr}<>"",D{fdr}<>""),'
            f'COUNTIFS(\'Lead Tracker\'!B4:B1000,">="&B{fdr},\'Lead Tracker\'!B4:B1000,"<="&D{fdr}),0)'
        ),
        # Confirmed
        (
            f'=IF(AND(B{fdr}<>"",D{fdr}<>""),'
            f'COUNTIFS(\'Lead Tracker\'!B4:B1000,">="&B{fdr},\'Lead Tracker\'!B4:B1000,"<="&D{fdr},'
            f'\'Lead Tracker\'!V4:V1000,"Confirmed"),0)'
        ),
        # Completed
        (
            f'=IF(AND(B{fdr}<>"",D{fdr}<>""),'
            f'COUNTIFS(\'Lead Tracker\'!B4:B1000,">="&B{fdr},\'Lead Tracker\'!B4:B1000,"<="&D{fdr},'
            f'\'Lead Tracker\'!V4:V1000,"Completed"),0)'
        ),
        # Pending
        (
            f'=IF(AND(B{fdr}<>"",D{fdr}<>""),'
            f'COUNTIFS(\'Lead Tracker\'!B4:B1000,">="&B{fdr},\'Lead Tracker\'!B4:B1000,"<="&D{fdr},'
            f'\'Lead Tracker\'!V4:V1000,"Pending"),0)'
        ),
        # Cancelled
        (
            f'=IF(AND(B{fdr}<>"",D{fdr}<>""),'
            f'COUNTIFS(\'Lead Tracker\'!B4:B1000,">="&B{fdr},\'Lead Tracker\'!B4:B1000,"<="&D{fdr},'
            f'\'Lead Tracker\'!V4:V1000,"Cancelled"),0)'
        ),
        # Revenue in Range
        (
            f'=IF(AND(B{fdr}<>"",D{fdr}<>""),'
            f'SUMPRODUCT((\'Lead Tracker\'!B4:B1000>=B{fdr})*(\'Lead Tracker\'!B4:B1000<=D{fdr})'
            f'*(\'Lead Tracker\'!AG4:AG1000="Received")*(\'Lead Tracker\'!AF4:AF1000)),0)'
        ),
        # Payments Pending
        (
            f'=IF(AND(B{fdr}<>"",D{fdr}<>""),'
            f'SUMPRODUCT((\'Lead Tracker\'!B4:B1000>=B{fdr})*(\'Lead Tracker\'!B4:B1000<=D{fdr})'
            f'*(\'Lead Tracker\'!AG4:AG1000="Pending")*(\'Lead Tracker\'!AF4:AF1000)),0)'
        ),
    ]
    for i, formula in enumerate(date_formulas):
        cell = ws.cell(row=rpt_data_row, column=i + 1)
        cell.value = formula
        cell.font = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
        cell.fill = make_fill(LIGHT_BLUE)
        cell.alignment = center_wrap
        cell.border = thin_border
        if i >= 5:  # Revenue and Payments Pending
            cell.number_format = '₹#,##0'

    # ── Charts ───────────────────────────────────────────────────────────
    # Bar Chart: Service-Wise Orders
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Service-Wise Orders"
    bar_chart.y_axis.title = "Number of Orders"
    bar_chart.x_axis.title = "Service"
    bar_chart.width = 20
    bar_chart.height = 12

    # Data: Total Orders column B, rows 16-26 (16 = header for titles_from_data)
    data_ref = Reference(ws, min_col=2, min_row=16, max_row=26)
    cats_ref = Reference(ws, min_col=1, min_row=17, max_row=26)
    bar_chart.add_data(data_ref, titles_from_data=True)
    bar_chart.set_categories(cats_ref)
    bar_chart.shape = 4
    ws.add_chart(bar_chart, f"J{R_SB}")

    # Pie Chart: Area-Wise Order Distribution
    pie_chart = PieChart()
    pie_chart.title = "Area-Wise Order Distribution"
    pie_chart.width = 16
    pie_chart.height = 12

    area_data_ref = Reference(ws, min_col=2, min_row=area_hdr_row, max_row=area_total_row - 1)
    area_cats_ref = Reference(ws, min_col=1, min_row=area_data_start, max_row=area_total_row - 1)
    pie_chart.add_data(area_data_ref, titles_from_data=True)
    pie_chart.set_categories(area_cats_ref)
    ws.add_chart(pie_chart, f"J{area_start}")

    # ── Freeze panes ─────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    return ws


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 3 — FORM FIELDS REFERENCE
# ═════════════════════════════════════════════════════════════════════════════
def create_form_fields_reference(wb):
    ws = wb.create_sheet("Form Fields Reference")
    ws.sheet_properties.tabColor = "2E5090"

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 60

    # ── Row 1: Title ─────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "GOOGLE FORM — FIELD SETUP REFERENCE"
    t.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    t.fill = make_fill(DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 45
    for c in range(2, 4):
        ws.cell(row=1, column=c).fill = make_fill(DARK_BLUE)

    # ── Row 3: Headers ───────────────────────────────────────────────────
    headers = ["Field Name", "Field Type", "Options / Notes"]
    for i, hdr in enumerate(headers):
        cell = ws.cell(row=3, column=i + 1)
        cell.value = hdr
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = make_fill(MEDIUM_BLUE)
        cell.alignment = center_wrap
        cell.border = thin_border
    ws.row_dimensions[3].height = 30

    # ── Field rows ───────────────────────────────────────────────────────
    services_str = ", ".join(SERVICES_LIST)
    time_slots_str = ", ".join(TIME_SLOTS_LIST)

    fields = [
        ("Customer Name", "Short Answer", "Required"),
        ("Phone Number", "Short Answer", "Required, Number validation"),
        ("Alternate Phone", "Short Answer", "Optional"),
        ("Email", "Short Answer", "Email validation"),
        ("City", "Dropdown", CITIES.replace(",", ", ")),
        ("Area / Locality", "Short Answer", "e.g. Indiranagar, Koramangala, HSR Layout"),
        ("Full Address", "Paragraph", "Required"),
        ("Service 1 Required", "Dropdown", services_str),
        ("Service 2 (Optional)", "Dropdown", services_str),
        ("Service 3 (Optional)", "Dropdown", services_str),
        ("Service 4 (Optional)", "Dropdown", services_str),
        ("Preferred Date", "Date", "Date picker"),
        ("Preferred Time Slot", "Dropdown", time_slots_str),
        ("Any Special Instructions", "Paragraph", "Optional"),
    ]

    for idx, (fname, ftype, notes) in enumerate(fields):
        row = 4 + idx
        bg = LIGHT_GRAY if idx % 2 == 1 else WHITE

        cell_a = ws.cell(row=row, column=1)
        cell_a.value = fname
        cell_a.font = Font(name="Calibri", size=10, bold=True)
        cell_a.alignment = left_wrap
        cell_a.border = thin_border
        cell_a.fill = make_fill(bg)

        cell_b = ws.cell(row=row, column=2)
        cell_b.value = ftype
        cell_b.font = Font(name="Calibri", size=10)
        cell_b.alignment = center_wrap
        cell_b.border = thin_border
        cell_b.fill = make_fill(bg)

        cell_c = ws.cell(row=row, column=3)
        cell_c.value = notes
        cell_c.font = Font(name="Calibri", size=10)
        cell_c.alignment = left_wrap
        cell_c.border = thin_border
        cell_c.fill = make_fill(bg)

    # ── Instructions section ─────────────────────────────────────────────
    instr_start = 4 + len(fields) + 2  # 2 blank rows

    ws.merge_cells(f"A{instr_start}:C{instr_start}")
    h = ws.cell(row=instr_start, column=1)
    h.value = "HOW TO LINK GOOGLE FORM TO THIS SHEET"
    h.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    h.fill = make_fill(MEDIUM_BLUE)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[instr_start].height = 35
    for c in range(2, 4):
        ws.cell(row=instr_start, column=c).fill = make_fill(MEDIUM_BLUE)

    steps = [
        "1. Create a new Google Form with the fields listed above.",
        "2. In the Google Form, click the Responses tab.",
        "3. Click the Google Sheets icon (\"Link to Sheets\").",
        "4. Select \"Select existing spreadsheet\" and choose this workbook.",
        "5. The form will create a new tab OR link to 'Lead Tracker' tab.",
        "6. If it creates a new tab, copy the response columns to match Lead Tracker columns B onward.",
        "7. Form timestamps auto-fill column B; formulas in A, C, and S handle the rest.",
        "8. Manually update Order Status, Vendor info, and Payment details after each lead.",
        "9. Use Data > Create a filter on the Lead Tracker for quick sorting.",
        "10. Check the Dashboard tab for real-time analytics and stats.",
    ]

    for s_idx, step in enumerate(steps):
        row = instr_start + 1 + s_idx
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = step
        cell.font = Font(name="Calibri", size=11, color="333333")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 24

    return ws


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()

    print("Creating HOW TO USE sheet...")
    create_how_to_use(wb)

    print("Creating Lead Tracker sheet...")
    create_lead_tracker(wb)

    print("Creating Dashboard sheet...")
    create_dashboard(wb)

    print("Creating Form Fields Reference sheet...")
    create_form_fields_reference(wb)

    # Set the first sheet as active
    wb.active = 0

    output_path = "/var/lib/freelancer/projects/40182876/Home_Services_Lead_Tracker.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\nWorkbook saved to: {output_path}")
    print("Done! Open in Google Sheets for best experience (calendar pickers, filters, etc.).")


if __name__ == "__main__":
    main()
