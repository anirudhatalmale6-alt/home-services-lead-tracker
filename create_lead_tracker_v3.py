#!/usr/bin/env python3
"""
Home Services Lead Tracker - Complete Excel Generator
Generates: Home_Services_Lead_Tracker.xlsx
"""

import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ── Color Constants ──────────────────────────────────────────────────────────
DARK_BLUE = "1B2A4A"
MEDIUM_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
GREEN_BG = "C6EFCE"
GREEN_FONT = "006100"
YELLOW_BG = "FFEB9C"
YELLOW_FONT = "9C6500"
RED_BG = "FFC7CE"
RED_FONT = "9C0006"
BLUE_BG = "B4C6E7"
BLUE_FONT = "1F4E79"
COMPLETED_BG = "548235"
COMPLETED_FONT_CLR = "FFFFFF"
REFUND_BG = "FFE0B2"
REFUND_FONT = "E65100"
LIGHT_GRAY = "F2F2F2"
BORDER_COLOR = "B4B4B4"

# ── Reusable style helpers ───────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

medium_blue_border = Border(
    left=Side(style="medium", color=MEDIUM_BLUE),
    right=Side(style="medium", color=MEDIUM_BLUE),
    top=Side(style="medium", color=MEDIUM_BLUE),
    bottom=Side(style="medium", color=MEDIUM_BLUE),
)

def make_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def apply_cell(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format

# ── Dropdown lists ───────────────────────────────────────────────────────────
CITY_LIST = "Bangalore,Mumbai,Delhi,Hyderabad,Chennai,Pune,Other"
BHK_LIST = "1BHK,2BHK,3BHK,4BHK,4+BHK,Not Applicable"
SERVICE_LIST = ("Bathroom Cleaning,Kitchen Cleaning,Full Home Cleaning,"
                "Rental Property Cleaning,Ready to Move In Cleaning,"
                "Painting,Pest Control,Plumbing,Electrician,Other")
SERVICES = [s.strip() for s in SERVICE_LIST.split(",")]
SLOT_LIST = "Morning 8-10,Morning 10-12,Afternoon 12-2,Afternoon 2-4,Evening 4-6,Evening 6-8"
SOURCE_LIST = "Website,Instagram,JustDial,Google,Referral,WhatsApp,Walk-in,Other"
SOURCES = [s.strip() for s in SOURCE_LIST.split(",")]
STATUS_LIST = "Confirmed,Pending,Cancelled,Scheduled,Completed,Refunded"
ADV_STATUS_LIST = "Received,NIL,Cleared"
PAY_STATUS_LIST = "Received,Pending"
PAY_MODE_LIST = "Cash,UPI,Debit Card,Payment Gateway,Bank Transfer,Other"

AREAS = [
    "Indiranagar", "Koramangala", "HSR Layout", "Whitefield", "JP Nagar",
    "BTM Layout", "Electronic City", "MG Road", "Marathahalli", "Banashankari", "Other"
]

# ── Column headers for Lead Tracker (A-AR = 44 columns) ────────────────────
HEADERS = [
    "S.No",              # A  (1)
    "Timestamp",          # B  (2)
    "Order ID",           # C  (3)
    "Customer Name",      # D  (4)
    "Phone Number",       # E  (5)
    "WhatsApp Number",    # F  (6)
    "Alternate Phone",    # G  (7)
    "Email",              # H  (8)
    "City",               # I  (9)
    "Area / Locality",    # J  (10)
    "Full Address",       # K  (11)
    "BHK",                # L  (12)
    "SQFT",               # M  (13)
    "Service 1",          # N  (14)
    "Price 1",            # O  (15)
    "Service 2",          # P  (16)
    "Price 2",            # Q  (17)
    "Service 3",          # R  (18)
    "Price 3",            # S  (19)
    "Service 4",          # T  (20)
    "Price 4",            # U  (21)
    "Total Value",        # V  (22)
    "Discount Amount",    # W  (23)
    "Discounted Total",   # X  (24)
    "Preferred Date",     # Y  (25)
    "Slot Time",          # Z  (26)
    "Order Source",       # AA (27)
    "Order Status",       # AB (28)
    "Reason / Notes",     # AC (29)
    "Vendor Name",        # AD (30)
    "Vendor Contact",     # AE (31)
    "Vendor Alternate No.", # AF (32)
    "Order Scheduled Date", # AG (33)
    "Scheduled Time",    # AH (34)
    "Order Completed Date", # AI (35)
    "Advance Amount",    # AJ (36)
    "Advance Status",    # AK (37)
    "Pending Balance",   # AL (38)
    "Payment Value",     # AM (39)
    "Payment Status",    # AN (40)
    "Payment Mode",      # AO (41)
    "Refund Amount",     # AP (42)
    "Invoice Number",    # AQ (43)
    "Transaction Ref. No." # AR (44)
]

COL_WIDTHS = {
    "A": 6, "B": 18, "C": 12, "D": 20, "E": 16, "F": 16, "G": 16, "H": 22,
    "I": 14, "J": 18, "K": 30, "L": 10, "M": 10,
    "N": 20, "O": 12, "P": 20, "Q": 12, "R": 20, "S": 12, "T": 20, "U": 12,
    "V": 14, "W": 14, "X": 14,
    "Y": 16, "Z": 22, "AA": 16, "AB": 16, "AC": 25,
    "AD": 18, "AE": 16, "AF": 18, "AG": 16, "AH": 22, "AI": 16,
    "AJ": 14, "AK": 14, "AL": 14, "AM": 16, "AN": 14, "AO": 16,
    "AP": 14, "AQ": 14, "AR": 18,
}

CURRENCY_FMT = '₹#,##0'
DATE_FMT = 'DD-MMM-YYYY'
DATETIME_FMT = 'DD-MMM-YYYY HH:MM'

# ═══════════════════════════════════════════════════════════════════════════
#  WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 0: HOW TO USE
# ═══════════════════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "HOW TO USE"
ws0.sheet_properties.tabColor = DARK_BLUE

title_font = Font(name="Calibri", bold=True, size=16, color=WHITE)
title_fill = make_fill(DARK_BLUE)
sub_font = Font(name="Calibri", bold=True, size=12, color=WHITE)
sub_fill = make_fill(MEDIUM_BLUE)
body_font = Font(name="Calibri", size=11, color="333333")
heading_font = Font(name="Calibri", bold=True, size=12, color=MEDIUM_BLUE)
center_al = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Title
ws0.merge_cells("A1:H1")
c = ws0["A1"]
c.value = "HOME SERVICES LEAD TRACKER — QUICK START GUIDE"
apply_cell(c, font=title_font, fill=title_fill, alignment=center_al)
ws0.row_dimensions[1].height = 45

ws0.merge_cells("A2:H2")
c = ws0["A2"]
c.value = "Follow these steps to get started with your tracker"
apply_cell(c, font=sub_font, fill=sub_fill, alignment=center_al)
ws0.row_dimensions[2].height = 30

instructions = [
    ("1. OPENING IN GOOGLE SHEETS",
     "Upload this file to Google Drive (drive.google.com). "
     "Right-click the file > Open with > Google Sheets. "
     "All formulas, dropdowns, and conditional formatting will work automatically."),
    ("2. ENABLE FILTERS",
     "Go to Data > Create a filter. This adds filter arrows to every column header "
     "so you can sort and filter by status, service, area, date, etc."),
    ("3. COLOR CODES EXPLAINED",
     "GREEN = Confirmed / Payment Received  |  YELLOW = Pending / Advance NIL  |  "
     "RED = Cancelled / Payment Pending  |  BLUE = Scheduled  |  PURPLE tint = Completed  |  "
     "ORANGE = Refunded"),
    ("4. USING THE LEAD TRACKER",
     "Start entering data from Row 4. The S.No, Order ID, Total Value, Discounted Total, "
     "Pending Balance, and Invoice Number columns auto-calculate via formulas. "
     "Use dropdowns for City, BHK, Services, Status, Payment, and Source fields."),
    ("5. CALENDAR / DATE PICKERS",
     "In Google Sheets, click any date cell (Preferred Date, Scheduled Date, Completed Date) "
     "and a calendar picker appears automatically. Dates are formatted DD-MMM-YYYY."),
    ("6. DASHBOARD TAB",
     "The Dashboard sheet auto-updates with order statistics, payment summary, "
     "service-wise breakdown, area-wise distribution, order source tracking, "
     "and a date-wise report with calendar inputs. Charts visualise the data."),
    ("7. LINKING A GOOGLE FORM",
     "Create a Google Form with fields listed in the 'Form Fields Reference' tab. "
     "Link the form response sheet, then copy-paste or IMPORTRANGE the data into "
     "the Lead Tracker. The Order ID and Invoice formulas populate automatically."),
    ("8. TIPS",
     "Keep the first 4 columns (A-D) visible — they are frozen. "
     "Use Order Source to track your best marketing channels. "
     "Regularly review the Dashboard for business insights. "
     "Back up your sheet weekly."),
]

row = 4
for heading, body in instructions:
    ws0.merge_cells(f"A{row}:H{row}")
    c = ws0[f"A{row}"]
    c.value = heading
    apply_cell(c, font=heading_font, alignment=left_al)
    ws0.row_dimensions[row].height = 25
    row += 1
    ws0.merge_cells(f"A{row}:H{row}")
    c = ws0[f"A{row}"]
    c.value = body
    apply_cell(c, font=body_font, alignment=left_al)
    ws0.row_dimensions[row].height = 55
    row += 2

for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
    ws0.column_dimensions[col_letter].width = 18

# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 1: LEAD TRACKER
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Lead Tracker")
ws1.sheet_properties.tabColor = MEDIUM_BLUE

num_cols = len(HEADERS)  # 44
last_col_letter = get_column_letter(num_cols)  # AR

# ── Row 1: Title ─────────────────────────────────────────────────────────
ws1.merge_cells(f"A1:{last_col_letter}1")
c = ws1["A1"]
c.value = "HOME SERVICES — LEAD TRACKER & ORDER MANAGEMENT"
apply_cell(c, font=Font(name="Calibri", bold=True, size=16, color=WHITE),
           fill=make_fill(DARK_BLUE), alignment=center_al)
ws1.row_dimensions[1].height = 45

# ── Row 2: Subtitle ─────────────────────────────────────────────────────
ws1.merge_cells(f"A2:{last_col_letter}2")
c = ws1["A2"]
c.value = "Comprehensive Lead Management for Home Services Business — Bangalore & Pan-India"
apply_cell(c, font=Font(name="Calibri", bold=True, size=11, color=WHITE),
           fill=make_fill(MEDIUM_BLUE), alignment=center_al)
ws1.row_dimensions[2].height = 28

# ── Row 3: Column Headers ───────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, size=10, color=WHITE)
header_fill = make_fill(MEDIUM_BLUE)
header_al = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[3].height = 40

for idx, hdr in enumerate(HEADERS, 1):
    cell = ws1.cell(row=3, column=idx, value=hdr)
    apply_cell(cell, font=header_font, fill=header_fill, alignment=header_al, border=thin_border)

# ── Column widths ────────────────────────────────────────────────────────
for col_letter, w in COL_WIDTHS.items():
    ws1.column_dimensions[col_letter].width = w

# ── Data Validations ────────────────────────────────────────────────────
def add_dropdown(ws, col_letter, formula1, start_row=4, end_row=1000):
    dv = DataValidation(type="list", formula1=f'"{formula1}"', allow_blank=True)
    dv.showDropDown = False
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")
    return dv

add_dropdown(ws1, "I", CITY_LIST)
# Area / Locality dropdown (column J)
AREA_LIST = ",".join(AREAS)
add_dropdown(ws1, "J", AREA_LIST)
add_dropdown(ws1, "L", BHK_LIST)
add_dropdown(ws1, "N", SERVICE_LIST)
add_dropdown(ws1, "P", SERVICE_LIST)
add_dropdown(ws1, "R", SERVICE_LIST)
add_dropdown(ws1, "T", SERVICE_LIST)
add_dropdown(ws1, "Z", SLOT_LIST)
add_dropdown(ws1, "AA", SOURCE_LIST)
add_dropdown(ws1, "AB", STATUS_LIST)
add_dropdown(ws1, "AH", SLOT_LIST)
add_dropdown(ws1, "AK", ADV_STATUS_LIST)
add_dropdown(ws1, "AN", PAY_STATUS_LIST)
add_dropdown(ws1, "AO", PAY_MODE_LIST)

# Date validations — use "between" operator for reliable Google Sheets calendar
from datetime import date as date_type
for date_col in ["Y", "AG", "AI"]:
    dv = DataValidation(
        type="date",
        operator="between",
        formula1=date_type(2020, 1, 1),
        formula2=date_type(2035, 12, 31),
        allow_blank=True
    )
    dv.showInputMessage = True
    dv.promptTitle = "Select Date"
    dv.prompt = "Click to open calendar picker"
    dv.showErrorMessage = True
    ws1.add_data_validation(dv)
    dv.add(f"{date_col}4:{date_col}1000")

# ── Pre-format rows 4-100: formulas, styles ─────────────────────────────
data_font = Font(name="Calibri", size=10)
data_al = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_data_al = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Columns that get currency format
currency_cols = ["O", "Q", "S", "U", "V", "W", "X", "AJ", "AL", "AM", "AP"]
# Columns that get date format
date_cols = ["Y", "AG", "AI"]
# Timestamp column
ts_col = "B"

for row in range(4, 101):
    for col_idx in range(1, num_cols + 1):
        cell = ws1.cell(row=row, column=col_idx)
        col_l = get_column_letter(col_idx)
        apply_cell(cell, font=data_font, alignment=data_al, border=thin_border)

        # Number formats
        if col_l in currency_cols:
            cell.number_format = CURRENCY_FMT
        elif col_l in date_cols:
            cell.number_format = DATE_FMT
        elif col_l == ts_col:
            cell.number_format = DATETIME_FMT
        elif col_l == "M":
            cell.number_format = '#,##0'

        # Left-align text-heavy columns
        if col_l in ["D", "H", "J", "K", "AC"]:
            cell.alignment = left_data_al

    # Formulas
    r = row
    # A: S.No
    ws1[f"A{r}"] = f'=IF(B{r}<>"",ROW()-3,"")'
    ws1[f"A{r}"].number_format = '0'
    # C: Order ID
    ws1[f"C{r}"] = f'=IF(B{r}<>"","ST-"&TEXT(ROW()-3,"0000"),"")'
    ws1[f"C{r}"].font = Font(name="Calibri", size=10, bold=True, color=MEDIUM_BLUE)
    # V: Total Value
    ws1[f"V{r}"] = f'=IF(COUNTA(N{r})>0,SUM(O{r},Q{r},S{r},U{r}),"")'
    # X: Discounted Total
    ws1[f"X{r}"] = f'=IF(V{r}<>"",V{r}-IF(W{r}<>"",W{r},0),"")'
    # AL: Pending Balance (0 when payment received, otherwise Discounted Total - Advance)
    ws1[f"AL{r}"] = f'=IF(X{r}="","",IF(AN{r}="Received",0,X{r}-IF(AJ{r}<>"",AJ{r},0)))'
    # AQ: Invoice Number
    ws1[f"AQ{r}"] = f'=IF(B{r}<>"","INV-"&TEXT(ROW()-3,"0000"),"")'

# ── Conditional Formatting ───────────────────────────────────────────────
# Order Status column AB (col 28)
ws1.conditional_formatting.add(
    "AB4:AB1000",
    CellIsRule(operator="equal", formula=['"Confirmed"'],
              fill=make_fill(GREEN_BG), font=Font(color=GREEN_FONT, bold=True)))
ws1.conditional_formatting.add(
    "AB4:AB1000",
    CellIsRule(operator="equal", formula=['"Pending"'],
              fill=make_fill(YELLOW_BG), font=Font(color=YELLOW_FONT, bold=True)))
ws1.conditional_formatting.add(
    "AB4:AB1000",
    CellIsRule(operator="equal", formula=['"Cancelled"'],
              fill=make_fill(RED_BG), font=Font(color=RED_FONT, bold=True)))
ws1.conditional_formatting.add(
    "AB4:AB1000",
    CellIsRule(operator="equal", formula=['"Scheduled"'],
              fill=make_fill(BLUE_BG), font=Font(color=BLUE_FONT, bold=True)))
ws1.conditional_formatting.add(
    "AB4:AB1000",
    CellIsRule(operator="equal", formula=['"Completed"'],
              fill=make_fill(COMPLETED_BG), font=Font(color=COMPLETED_FONT_CLR, bold=True)))
ws1.conditional_formatting.add(
    "AB4:AB1000",
    CellIsRule(operator="equal", formula=['"Refunded"'],
              fill=make_fill(REFUND_BG), font=Font(color=REFUND_FONT, bold=True)))

# Payment Status column AN (col 40)
ws1.conditional_formatting.add(
    "AN4:AN1000",
    CellIsRule(operator="equal", formula=['"Received"'],
              fill=make_fill(GREEN_BG), font=Font(color=GREEN_FONT, bold=True)))
ws1.conditional_formatting.add(
    "AN4:AN1000",
    CellIsRule(operator="equal", formula=['"Pending"'],
              fill=make_fill(RED_BG), font=Font(color=RED_FONT, bold=True)))

# Advance Status column AK (col 37)
ws1.conditional_formatting.add(
    "AK4:AK1000",
    CellIsRule(operator="equal", formula=['"Received"'],
              fill=make_fill(GREEN_BG), font=Font(color=GREEN_FONT, bold=True)))
ws1.conditional_formatting.add(
    "AK4:AK1000",
    CellIsRule(operator="equal", formula=['"NIL"'],
              fill=make_fill(YELLOW_BG), font=Font(color=YELLOW_FONT, bold=True)))
ws1.conditional_formatting.add(
    "AK4:AK1000",
    CellIsRule(operator="equal", formula=['"Cleared"'],
              fill=make_fill("A9D18E"), font=Font(color="375623", bold=True)))

# Row-level conditional formatting based on $AB column
row_range = f"A4:{last_col_letter}1000"
ws1.conditional_formatting.add(
    row_range,
    FormulaRule(formula=['$AB4="Confirmed"'], fill=make_fill("E2EFDA")))
ws1.conditional_formatting.add(
    row_range,
    FormulaRule(formula=['$AB4="Pending"'], fill=make_fill("FFF2CC")))
ws1.conditional_formatting.add(
    row_range,
    FormulaRule(formula=['$AB4="Cancelled"'], fill=make_fill("FCE4EC")))
ws1.conditional_formatting.add(
    row_range,
    FormulaRule(formula=['$AB4="Scheduled"'], fill=make_fill("DAEEF3")))
ws1.conditional_formatting.add(
    row_range,
    FormulaRule(formula=['$AB4="Completed"'], fill=make_fill("E8D5F5")))
ws1.conditional_formatting.add(
    row_range,
    FormulaRule(formula=['$AB4="Refunded"'], fill=make_fill("FFF3E0")))

# ── Sample Data (rows 4-8) ──────────────────────────────────────────────
now = datetime.now()

# Column letter -> column index mapping
col_map = {}
for i in range(1, num_cols + 1):
    col_map[get_column_letter(i)] = i

# Formula columns that must NOT be overwritten with static data
formula_cols = {"A", "C", "V", "X", "AL", "AQ"}

samples = [
    {   # Row 4 - Rajesh Kumar, Confirmed
        "B": now - timedelta(days=5),
        "D": "Rajesh Kumar", "E": "9876543210", "F": "9876543210", "G": "",
        "H": "rajesh.k@email.com", "I": "Bangalore", "J": "Indiranagar",
        "K": "123, 12th Main, Indiranagar, Bangalore - 560038",
        "L": "2BHK", "M": 1200,
        "N": "Bathroom Cleaning", "O": 2000, "P": "Pest Control", "Q": 3500,
        "W": 500,
        "Y": now + timedelta(days=2), "Z": "Morning 10-12",
        "AA": "Website", "AB": "Confirmed", "AC": "",
        "AD": "Suresh", "AE": "9988776655",
        "AG": now + timedelta(days=2), "AH": "Morning 10-12",
        "AJ": 500, "AK": "Received",
    },
    {   # Row 5 - Priya Sharma, Pending
        "B": now - timedelta(days=3),
        "D": "Priya Sharma", "E": "8765432109", "F": "", "G": "",
        "H": "priya.s@email.com", "I": "Bangalore", "J": "HSR Layout",
        "K": "45, Sector 2, HSR Layout, Bangalore - 560102",
        "L": "3BHK", "M": 1800,
        "N": "Full Home Cleaning", "O": 4500,
        "Y": now + timedelta(days=5), "Z": "Afternoon 2-4",
        "AA": "Instagram", "AB": "Pending",
    },
    {   # Row 6 - Amit Patel, Cancelled
        "B": now - timedelta(days=7),
        "D": "Amit Patel", "E": "7654321098", "F": "", "G": "",
        "H": "amit.p@email.com", "I": "Bangalore", "J": "Whitefield",
        "K": "78, ITPL Road, Whitefield, Bangalore - 560066",
        "L": "2BHK", "M": 1100,
        "N": "Painting", "O": 15000, "P": "Plumbing", "Q": 2000,
        "W": 2000,
        "AA": "Google", "AB": "Cancelled", "AC": "Price too high",
    },
    {   # Row 7 - Meera Reddy, Scheduled
        "B": now - timedelta(days=2),
        "D": "Meera Reddy", "E": "6543210987", "F": "6543210987", "G": "",
        "H": "meera.r@email.com", "I": "Bangalore", "J": "Koramangala",
        "K": "22, 5th Block, Koramangala, Bangalore - 560095",
        "L": "3BHK", "M": 1600,
        "N": "Full Home Cleaning", "O": 4500,
        "Y": now + timedelta(days=3), "Z": "Morning 8-10",
        "AA": "JustDial", "AB": "Scheduled",
        "AD": "Ramesh", "AE": "9876501234",
        "AG": now + timedelta(days=3), "AH": "Morning 8-10",
        "AJ": 1000, "AK": "Received",
    },
    {   # Row 8 - Karthik Nair, Completed
        "B": now - timedelta(days=10),
        "D": "Karthik Nair", "E": "5432109876", "F": "", "G": "",
        "H": "karthik.n@email.com", "I": "Bangalore", "J": "JP Nagar",
        "K": "56, 6th Phase, JP Nagar, Bangalore - 560078",
        "L": "1BHK", "M": 650,
        "N": "Plumbing", "O": 1500,
        "Y": now - timedelta(days=8), "Z": "Evening 4-6",
        "AA": "Website", "AB": "Completed",
        "AD": "Vijay", "AE": "9123456780",
        "AG": now - timedelta(days=8), "AH": "Evening 4-6",
        "AI": now - timedelta(days=8),
        "AJ": 0, "AK": "NIL",
        "AM": 1500, "AN": "Received", "AO": "Cash", "AP": 0,
    },
]

for s_idx, sample in enumerate(samples):
    row = 4 + s_idx
    for col_letter, value in sample.items():
        if col_letter in formula_cols:
            continue
        col_idx = col_map[col_letter]
        cell = ws1.cell(row=row, column=col_idx)
        cell.value = value

# ── Freeze panes & auto filter ──────────────────────────────────────────
ws1.freeze_panes = "E4"
ws1.auto_filter.ref = f"A3:{last_col_letter}1000"

# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 2: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Dashboard")
ws2.sheet_properties.tabColor = DARK_BLUE

# Helper references
LT = "'Lead Tracker'"

# ── Row 1-2: Title ───────────────────────────────────────────────────────
ws2.merge_cells("A1:N1")
c = ws2["A1"]
c.value = "DASHBOARD — BUSINESS ANALYTICS & STATS"
apply_cell(c, font=Font(name="Calibri", bold=True, size=16, color=WHITE),
           fill=make_fill(DARK_BLUE), alignment=center_al)
ws2.row_dimensions[1].height = 45

ws2.merge_cells("A2:N2")
c = ws2["A2"]
c.value = "Auto-updated analytics from Lead Tracker data"
apply_cell(c, font=Font(name="Calibri", bold=True, size=11, color=WHITE),
           fill=make_fill(MEDIUM_BLUE), alignment=center_al)
ws2.row_dimensions[2].height = 28

# Column widths A-H = 18
for col_l in ["A", "B", "C", "D", "E", "F", "G", "H"]:
    ws2.column_dimensions[col_l].width = 18
for col_l in ["I", "J", "K", "L", "M", "N"]:
    ws2.column_dimensions[col_l].width = 16

# ── Row 4-5: Color Legend ────────────────────────────────────────────────
legend_row = 4
ws2[f"A{legend_row}"].value = "STATUS COLORS:"
apply_cell(ws2[f"A{legend_row}"], font=Font(name="Calibri", bold=True, size=10, color=DARK_BLUE),
           alignment=left_al)

legend_data = [
    ("B", "Confirmed", GREEN_BG, GREEN_FONT),
    ("C", "Pending", YELLOW_BG, YELLOW_FONT),
    ("D", "Cancelled", RED_BG, RED_FONT),
    ("E", "Scheduled", BLUE_BG, BLUE_FONT),
    ("F", "Completed", COMPLETED_BG, COMPLETED_FONT_CLR),
    ("G", "Refunded", REFUND_BG, REFUND_FONT),
]
for col_l, label, bg, fg in legend_data:
    cell = ws2[f"{col_l}{legend_row}"]
    cell.value = label
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=10, color=fg),
               fill=make_fill(bg), alignment=center_al, border=thin_border)

# ── SECTION 1: ORDER STATISTICS (Row 6-8) ───────────────────────────────
sec1_title_row = 6
sec1_header_row = 7
sec1_data_row = 8

ws2.merge_cells(f"A{sec1_title_row}:G{sec1_title_row}")
ws2[f"A{sec1_title_row}"].value = "ORDER STATISTICS"
apply_cell(ws2[f"A{sec1_title_row}"], font=Font(name="Calibri", bold=True, size=12, color=WHITE),
           fill=make_fill(DARK_BLUE), alignment=center_al)

stat_headers = ["Total Orders", "Confirmed", "Pending", "Cancelled", "Scheduled", "Completed", "Refunded"]
stat_colors = [LIGHT_BLUE, GREEN_BG, YELLOW_BG, RED_BG, BLUE_BG, COMPLETED_BG, REFUND_BG]
stat_font_colors = [DARK_BLUE, GREEN_FONT, YELLOW_FONT, RED_FONT, BLUE_FONT, COMPLETED_FONT_CLR, REFUND_FONT]

stat_formulas = [
    f'=COUNTIF({LT}!C4:C1000,"ST-*")',
    f'=COUNTIF({LT}!AB4:AB1000,"Confirmed")',
    f'=COUNTIF({LT}!AB4:AB1000,"Pending")',
    f'=COUNTIF({LT}!AB4:AB1000,"Cancelled")',
    f'=COUNTIF({LT}!AB4:AB1000,"Scheduled")',
    f'=COUNTIF({LT}!AB4:AB1000,"Completed")',
    f'=COUNTIF({LT}!AB4:AB1000,"Refunded")',
]

for i, (hdr, bg, fg) in enumerate(zip(stat_headers, stat_colors, stat_font_colors)):
    col_l = get_column_letter(i + 1)
    # Header
    cell = ws2[f"{col_l}{sec1_header_row}"]
    cell.value = hdr
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=10, color=WHITE),
               fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border)
    # Data
    cell = ws2[f"{col_l}{sec1_data_row}"]
    cell.value = stat_formulas[i]
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=22, color=fg),
               fill=make_fill(bg), alignment=center_al, border=thin_border)

ws2.row_dimensions[sec1_data_row].height = 50

# ── SECTION 2: PAYMENT SUMMARY ──────────────────────────────────────────
R_PS = sec1_data_row + 2  # row 10
ws2.merge_cells(f"A{R_PS}:H{R_PS}")
ws2[f"A{R_PS}"].value = "PAYMENT SUMMARY"
apply_cell(ws2[f"A{R_PS}"], font=Font(name="Calibri", bold=True, size=12, color=WHITE),
           fill=make_fill(DARK_BLUE), alignment=center_al)

pay_headers = ["Total Quoted", "Total Received", "Total Pending", "Cash Received",
               "UPI Received", "Card / Gateway", "Total Advance", "Total Refunds"]
pay_formulas = [
    # Total Quoted: sum of Discounted Total (X) where not blank
    f'=SUMPRODUCT(({LT}!X4:X1000<>"")*({LT}!X4:X1000))',
    # Total Received: sum AM where AN=Received
    f'=SUMPRODUCT(({LT}!AN4:AN1000="Received")*({LT}!AM4:AM1000))',
    # Total Pending: sum AM where AN=Pending
    f'=SUMPRODUCT(({LT}!AN4:AN1000="Pending")*({LT}!AM4:AM1000))',
    # Cash Received
    f'=SUMPRODUCT(({LT}!AO4:AO1000="Cash")*({LT}!AN4:AN1000="Received")*({LT}!AM4:AM1000))',
    # UPI Received
    f'=SUMPRODUCT(({LT}!AO4:AO1000="UPI")*({LT}!AN4:AN1000="Received")*({LT}!AM4:AM1000))',
    # Card/Gateway
    f'=SUMPRODUCT((({LT}!AO4:AO1000="Debit Card")+({LT}!AO4:AO1000="Payment Gateway"))*({LT}!AN4:AN1000="Received")*({LT}!AM4:AM1000))',
    # Total Advance where AK=Received
    f'=SUMPRODUCT(({LT}!AK4:AK1000="Received")*({LT}!AJ4:AJ1000))',
    # Total Refunds
    f'=SUM({LT}!AP4:AP1000)',
]

r_pay_hdr = R_PS + 1
r_pay_data = R_PS + 2
for i, (hdr, formula) in enumerate(zip(pay_headers, pay_formulas)):
    col_l = get_column_letter(i + 1)
    cell = ws2[f"{col_l}{r_pay_hdr}"]
    cell.value = hdr
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=10, color=WHITE),
               fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border)
    cell = ws2[f"{col_l}{r_pay_data}"]
    cell.value = formula
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=14, color=DARK_BLUE),
               fill=make_fill(LIGHT_BLUE), alignment=center_al, border=thin_border,
               number_format=CURRENCY_FMT)

ws2.row_dimensions[r_pay_data].height = 40

# ── SECTION 3: DISCOUNT SUMMARY ─────────────────────────────────────────
R_DS = r_pay_data + 2
ws2.merge_cells(f"A{R_DS}:C{R_DS}")
ws2[f"A{R_DS}"].value = "DISCOUNT SUMMARY"
apply_cell(ws2[f"A{R_DS}"], font=Font(name="Calibri", bold=True, size=12, color=WHITE),
           fill=make_fill(DARK_BLUE), alignment=center_al)

disc_headers = ["Orders with Discount", "Total Discount Given", "Avg Discount"]
disc_formulas = [
    f'=COUNTIF({LT}!W4:W1000,">"&0)',
    f'=SUM({LT}!W4:W1000)',
    f'=IFERROR(AVERAGEIF({LT}!W4:W1000,">"&0),0)',
]

r_disc_hdr = R_DS + 1
r_disc_data = R_DS + 2
for i, (hdr, formula) in enumerate(zip(disc_headers, disc_formulas)):
    col_l = get_column_letter(i + 1)
    cell = ws2[f"{col_l}{r_disc_hdr}"]
    cell.value = hdr
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=10, color=WHITE),
               fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border)
    cell = ws2[f"{col_l}{r_disc_data}"]
    cell.value = formula
    fmt = '0' if i == 0 else CURRENCY_FMT
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=14, color=DARK_BLUE),
               fill=make_fill(LIGHT_BLUE), alignment=center_al, border=thin_border,
               number_format=fmt)

ws2.row_dimensions[r_disc_data].height = 40

# ── SECTION 4: ORDER SOURCE TRACKING ────────────────────────────────────
R_SRC = r_disc_data + 2
ws2.merge_cells(f"A{R_SRC}:E{R_SRC}")
ws2[f"A{R_SRC}"].value = "ORDER SOURCE TRACKING"
apply_cell(ws2[f"A{R_SRC}"], font=Font(name="Calibri", bold=True, size=12, color=WHITE),
           fill=make_fill(DARK_BLUE), alignment=center_al)

src_col_headers = ["Source", "Total Orders", "Confirmed", "Completed", "Revenue"]
r_src_hdr = R_SRC + 1
for i, hdr in enumerate(src_col_headers):
    col_l = get_column_letter(i + 1)
    cell = ws2[f"{col_l}{r_src_hdr}"]
    cell.value = hdr
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=10, color=WHITE),
               fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border)

r_src_start = r_src_hdr + 1
for s_idx, source in enumerate(SOURCES):
    r = r_src_start + s_idx
    ws2[f"A{r}"].value = source
    apply_cell(ws2[f"A{r}"], font=Font(name="Calibri", bold=True, size=10),
               alignment=left_al, border=thin_border, fill=make_fill(LIGHT_GRAY))
    # Total Orders
    ws2[f"B{r}"].value = f'=COUNTIF({LT}!AA4:AA1000,A{r})'
    apply_cell(ws2[f"B{r}"], font=data_font, alignment=center_al, border=thin_border)
    # Confirmed
    ws2[f"C{r}"].value = f'=COUNTIFS({LT}!AA4:AA1000,A{r},{LT}!AB4:AB1000,"Confirmed")'
    apply_cell(ws2[f"C{r}"], font=data_font, alignment=center_al, border=thin_border)
    # Completed
    ws2[f"D{r}"].value = f'=COUNTIFS({LT}!AA4:AA1000,A{r},{LT}!AB4:AB1000,"Completed")'
    apply_cell(ws2[f"D{r}"], font=data_font, alignment=center_al, border=thin_border)
    # Revenue
    ws2[f"E{r}"].value = (
        f'=SUMPRODUCT(({LT}!AA4:AA1000=A{r})*({LT}!AN4:AN1000="Received")*({LT}!AM4:AM1000))'
    )
    apply_cell(ws2[f"E{r}"], font=data_font, alignment=center_al, border=thin_border,
               number_format=CURRENCY_FMT)

# Total row
r_src_total = r_src_start + len(SOURCES)
ws2[f"A{r_src_total}"].value = "TOTAL"
apply_cell(ws2[f"A{r_src_total}"], font=Font(name="Calibri", bold=True, size=10, color=WHITE),
           fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border)
for col_l in ["B", "C", "D", "E"]:
    ws2[f"{col_l}{r_src_total}"].value = f"=SUM({col_l}{r_src_start}:{col_l}{r_src_total - 1})"
    fmt = CURRENCY_FMT if col_l == "E" else '0'
    apply_cell(ws2[f"{col_l}{r_src_total}"],
               font=Font(name="Calibri", bold=True, size=10, color=WHITE),
               fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border,
               number_format=fmt)

# ── SECTION 5: SERVICE-WISE BREAKDOWN ────────────────────────────────────
R_SVC = r_src_total + 2
ws2.merge_cells(f"A{R_SVC}:H{R_SVC}")
ws2[f"A{R_SVC}"].value = "SERVICE-WISE BREAKDOWN"
apply_cell(ws2[f"A{R_SVC}"], font=Font(name="Calibri", bold=True, size=12, color=WHITE),
           fill=make_fill(DARK_BLUE), alignment=center_al)

svc_col_headers = ["Service", "Total Orders", "Confirmed", "Scheduled", "Completed",
                   "Pending", "Cancelled", "Revenue"]
r_svc_hdr = R_SVC + 1
for i, hdr in enumerate(svc_col_headers):
    col_l = get_column_letter(i + 1)
    cell = ws2[f"{col_l}{r_svc_hdr}"]
    cell.value = hdr
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=10, color=WHITE),
               fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border)

r_svc_start = r_svc_hdr + 1
for s_idx, svc in enumerate(SERVICES):
    r = r_svc_start + s_idx
    ws2[f"A{r}"].value = svc
    apply_cell(ws2[f"A{r}"], font=Font(name="Calibri", bold=True, size=10),
               alignment=left_al, border=thin_border, fill=make_fill(LIGHT_GRAY))

    # Total Orders across 4 service columns
    ws2[f"B{r}"].value = (
        f'=COUNTIF({LT}!N4:N1000,A{r})+COUNTIF({LT}!P4:P1000,A{r})'
        f'+COUNTIF({LT}!R4:R1000,A{r})+COUNTIF({LT}!T4:T1000,A{r})'
    )
    apply_cell(ws2[f"B{r}"], font=data_font, alignment=center_al, border=thin_border)

    # Status-wise counts across all 4 service columns
    for si, (status_name, col_l) in enumerate(
            [("Confirmed", "C"), ("Scheduled", "D"), ("Completed", "E"),
             ("Pending", "F"), ("Cancelled", "G")]):
        ws2[f"{col_l}{r}"].value = (
            f'=COUNTIFS({LT}!N4:N1000,A{r},{LT}!AB4:AB1000,"{status_name}")'
            f'+COUNTIFS({LT}!P4:P1000,A{r},{LT}!AB4:AB1000,"{status_name}")'
            f'+COUNTIFS({LT}!R4:R1000,A{r},{LT}!AB4:AB1000,"{status_name}")'
            f'+COUNTIFS({LT}!T4:T1000,A{r},{LT}!AB4:AB1000,"{status_name}")'
        )
        apply_cell(ws2[f"{col_l}{r}"], font=data_font, alignment=center_al, border=thin_border)

    # Revenue: SUMPRODUCT matching service in each col * price where AN=Received
    ws2[f"H{r}"].value = (
        f'=SUMPRODUCT(({LT}!N4:N1000=A{r})*({LT}!AN4:AN1000="Received")*({LT}!O4:O1000))'
        f'+SUMPRODUCT(({LT}!P4:P1000=A{r})*({LT}!AN4:AN1000="Received")*({LT}!Q4:Q1000))'
        f'+SUMPRODUCT(({LT}!R4:R1000=A{r})*({LT}!AN4:AN1000="Received")*({LT}!S4:S1000))'
        f'+SUMPRODUCT(({LT}!T4:T1000=A{r})*({LT}!AN4:AN1000="Received")*({LT}!U4:U1000))'
    )
    apply_cell(ws2[f"H{r}"], font=data_font, alignment=center_al, border=thin_border,
               number_format=CURRENCY_FMT)

# Service total row
r_svc_total = r_svc_start + len(SERVICES)
ws2[f"A{r_svc_total}"].value = "TOTAL"
apply_cell(ws2[f"A{r_svc_total}"], font=Font(name="Calibri", bold=True, size=10, color=WHITE),
           fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border)
for col_l in ["B", "C", "D", "E", "F", "G", "H"]:
    ws2[f"{col_l}{r_svc_total}"].value = f"=SUM({col_l}{r_svc_start}:{col_l}{r_svc_total - 1})"
    fmt = CURRENCY_FMT if col_l == "H" else '0'
    apply_cell(ws2[f"{col_l}{r_svc_total}"],
               font=Font(name="Calibri", bold=True, size=10, color=WHITE),
               fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border,
               number_format=fmt)

# ── SECTION 6: AREA-WISE ORDER DISTRIBUTION ─────────────────────────────
R_AREA = r_svc_total + 2
ws2.merge_cells(f"A{R_AREA}:F{R_AREA}")
ws2[f"A{R_AREA}"].value = "AREA-WISE ORDER DISTRIBUTION"
apply_cell(ws2[f"A{R_AREA}"], font=Font(name="Calibri", bold=True, size=12, color=WHITE),
           fill=make_fill(DARK_BLUE), alignment=center_al)

area_col_headers = ["Area", "Total Orders", "Confirmed", "Completed", "Pending", "Revenue"]
r_area_hdr = R_AREA + 1
for i, hdr in enumerate(area_col_headers):
    col_l = get_column_letter(i + 1)
    cell = ws2[f"{col_l}{r_area_hdr}"]
    cell.value = hdr
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=10, color=WHITE),
               fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border)

r_area_start = r_area_hdr + 1
for a_idx, area in enumerate(AREAS):
    r = r_area_start + a_idx
    ws2[f"A{r}"].value = area
    apply_cell(ws2[f"A{r}"], font=Font(name="Calibri", bold=True, size=10),
               alignment=left_al, border=thin_border, fill=make_fill(LIGHT_GRAY))

    # Total Orders - COUNTIF on J column (Area/Locality)
    ws2[f"B{r}"].value = f'=COUNTIF({LT}!J4:J1000,A{r})'
    apply_cell(ws2[f"B{r}"], font=data_font, alignment=center_al, border=thin_border)

    # Confirmed
    ws2[f"C{r}"].value = f'=COUNTIFS({LT}!J4:J1000,A{r},{LT}!AB4:AB1000,"Confirmed")'
    apply_cell(ws2[f"C{r}"], font=data_font, alignment=center_al, border=thin_border)

    # Completed
    ws2[f"D{r}"].value = f'=COUNTIFS({LT}!J4:J1000,A{r},{LT}!AB4:AB1000,"Completed")'
    apply_cell(ws2[f"D{r}"], font=data_font, alignment=center_al, border=thin_border)

    # Pending
    ws2[f"E{r}"].value = f'=COUNTIFS({LT}!J4:J1000,A{r},{LT}!AB4:AB1000,"Pending")'
    apply_cell(ws2[f"E{r}"], font=data_font, alignment=center_al, border=thin_border)

    # Revenue
    ws2[f"F{r}"].value = (
        f'=SUMPRODUCT(({LT}!J4:J1000=A{r})*({LT}!AN4:AN1000="Received")*({LT}!AM4:AM1000))'
    )
    apply_cell(ws2[f"F{r}"], font=data_font, alignment=center_al, border=thin_border,
               number_format=CURRENCY_FMT)

# Area total row
r_area_total = r_area_start + len(AREAS)
ws2[f"A{r_area_total}"].value = "TOTAL"
apply_cell(ws2[f"A{r_area_total}"], font=Font(name="Calibri", bold=True, size=10, color=WHITE),
           fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border)
for col_l in ["B", "C", "D", "E", "F"]:
    ws2[f"{col_l}{r_area_total}"].value = f"=SUM({col_l}{r_area_start}:{col_l}{r_area_total - 1})"
    fmt = CURRENCY_FMT if col_l == "F" else '0'
    apply_cell(ws2[f"{col_l}{r_area_total}"],
               font=Font(name="Calibri", bold=True, size=10, color=WHITE),
               fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border,
               number_format=fmt)

# ── SECTION 7: DATE-WISE REPORT ─────────────────────────────────────────
R_DATE = r_area_total + 2
ws2.merge_cells(f"A{R_DATE}:H{R_DATE}")
ws2[f"A{R_DATE}"].value = "DATE-WISE REPORT"
apply_cell(ws2[f"A{R_DATE}"], font=Font(name="Calibri", bold=True, size=12, color=WHITE),
           fill=make_fill(DARK_BLUE), alignment=center_al)

# From / To date inputs
r_date_input = R_DATE + 1
ws2[f"A{r_date_input}"].value = "From Date"
apply_cell(ws2[f"A{r_date_input}"], font=Font(name="Calibri", bold=True, size=11, color=DARK_BLUE),
           alignment=center_al, border=thin_border)
ws2[f"B{r_date_input}"].value = None  # user enters date
apply_cell(ws2[f"B{r_date_input}"], font=Font(name="Calibri", bold=True, size=12, color=DARK_BLUE),
           fill=make_fill(YELLOW_BG), alignment=center_al, border=medium_blue_border,
           number_format=DATE_FMT)

ws2[f"C{r_date_input}"].value = "To Date"
apply_cell(ws2[f"C{r_date_input}"], font=Font(name="Calibri", bold=True, size=11, color=DARK_BLUE),
           alignment=center_al, border=thin_border)
ws2[f"D{r_date_input}"].value = None
apply_cell(ws2[f"D{r_date_input}"], font=Font(name="Calibri", bold=True, size=12, color=DARK_BLUE),
           fill=make_fill(YELLOW_BG), alignment=center_al, border=medium_blue_border,
           number_format=DATE_FMT)

# Date validation on the input cells — use "between" operator with wide range
# This more reliably triggers Google Sheets' built-in calendar picker
from datetime import date as date_type
for date_input_cell in [f"B{r_date_input}", f"D{r_date_input}"]:
    dv = DataValidation(
        type="date",
        operator="between",
        formula1=date_type(2020, 1, 1),
        formula2=date_type(2035, 12, 31),
        allow_blank=True
    )
    dv.showInputMessage = True
    dv.promptTitle = "Select Date"
    dv.prompt = "Click to open calendar picker"
    dv.showErrorMessage = True
    dv.errorTitle = "Invalid Date"
    dv.error = "Please select a valid date"
    ws2.add_data_validation(dv)
    dv.add(date_input_cell)

ws2[f"F{r_date_input}"].value = "\u2190 Click yellow cells for calendar in Google Sheets"
apply_cell(ws2[f"F{r_date_input}"], font=Font(name="Calibri", italic=True, size=10, color=MEDIUM_BLUE),
           alignment=left_al)

# Report headers
r_date_hdr = r_date_input + 1
date_report_headers = ["Orders in Range", "Confirmed", "Completed", "Pending",
                       "Cancelled", "Revenue", "Pending Payments"]
for i, hdr in enumerate(date_report_headers):
    col_l = get_column_letter(i + 1)
    cell = ws2[f"{col_l}{r_date_hdr}"]
    cell.value = hdr
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=10, color=WHITE),
               fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border)

# Report data
r_date_data = r_date_hdr + 1
ws2.row_dimensions[r_date_data].height = 45

from_cell = f"B{r_date_input}"
to_cell = f"D{r_date_input}"

# Date-wise formulas use INT() to strip time from Timestamp (col B) for reliable date comparison
# INT(datetime) returns just the date part in Google Sheets/Excel

# Orders in Range
ws2[f"A{r_date_data}"].value = (
    f'=IF(OR({from_cell}="",{to_cell}=""),"Enter dates",'
    f'SUMPRODUCT((INT({LT}!B4:B1000)>={from_cell})*(INT({LT}!B4:B1000)<={to_cell})*({LT}!B4:B1000<>"")))'
)
apply_cell(ws2[f"A{r_date_data}"], font=Font(name="Calibri", bold=True, size=16, color=DARK_BLUE),
           fill=make_fill(LIGHT_BLUE), alignment=center_al, border=thin_border)

# Confirmed in range
ws2[f"B{r_date_data}"].value = (
    f'=IF(OR({from_cell}="",{to_cell}=""),"",'
    f'SUMPRODUCT((INT({LT}!B4:B1000)>={from_cell})*(INT({LT}!B4:B1000)<={to_cell})*({LT}!AB4:AB1000="Confirmed")))'
)
apply_cell(ws2[f"B{r_date_data}"], font=Font(name="Calibri", bold=True, size=16, color=GREEN_FONT),
           fill=make_fill(LIGHT_BLUE), alignment=center_al, border=thin_border)

# Completed in range
ws2[f"C{r_date_data}"].value = (
    f'=IF(OR({from_cell}="",{to_cell}=""),"",'
    f'SUMPRODUCT((INT({LT}!B4:B1000)>={from_cell})*(INT({LT}!B4:B1000)<={to_cell})*({LT}!AB4:AB1000="Completed")))'
)
apply_cell(ws2[f"C{r_date_data}"], font=Font(name="Calibri", bold=True, size=16, color=DARK_BLUE),
           fill=make_fill(LIGHT_BLUE), alignment=center_al, border=thin_border)

# Pending in range
ws2[f"D{r_date_data}"].value = (
    f'=IF(OR({from_cell}="",{to_cell}=""),"",'
    f'SUMPRODUCT((INT({LT}!B4:B1000)>={from_cell})*(INT({LT}!B4:B1000)<={to_cell})*({LT}!AB4:AB1000="Pending")))'
)
apply_cell(ws2[f"D{r_date_data}"], font=Font(name="Calibri", bold=True, size=16, color=YELLOW_FONT),
           fill=make_fill(LIGHT_BLUE), alignment=center_al, border=thin_border)

# Cancelled in range
ws2[f"E{r_date_data}"].value = (
    f'=IF(OR({from_cell}="",{to_cell}=""),"",'
    f'SUMPRODUCT((INT({LT}!B4:B1000)>={from_cell})*(INT({LT}!B4:B1000)<={to_cell})*({LT}!AB4:AB1000="Cancelled")))'
)
apply_cell(ws2[f"E{r_date_data}"], font=Font(name="Calibri", bold=True, size=16, color=RED_FONT),
           fill=make_fill(LIGHT_BLUE), alignment=center_al, border=thin_border)

# Revenue in range
ws2[f"F{r_date_data}"].value = (
    f'=IF(OR({from_cell}="",{to_cell}=""),"",'
    f'SUMPRODUCT((INT({LT}!B4:B1000)>={from_cell})*(INT({LT}!B4:B1000)<={to_cell})'
    f'*({LT}!AN4:AN1000="Received")*({LT}!AM4:AM1000)))'
)
apply_cell(ws2[f"F{r_date_data}"], font=Font(name="Calibri", bold=True, size=16, color=DARK_BLUE),
           fill=make_fill(LIGHT_BLUE), alignment=center_al, border=thin_border,
           number_format=CURRENCY_FMT)

# Pending Payments in range
ws2[f"G{r_date_data}"].value = (
    f'=IF(OR({from_cell}="",{to_cell}=""),"",'
    f'SUMPRODUCT((INT({LT}!B4:B1000)>={from_cell})*(INT({LT}!B4:B1000)<={to_cell})'
    f'*({LT}!AN4:AN1000="Pending")*({LT}!AM4:AM1000)))'
)
apply_cell(ws2[f"G{r_date_data}"], font=Font(name="Calibri", bold=True, size=16, color=RED_FONT),
           fill=make_fill(LIGHT_BLUE), alignment=center_al, border=thin_border,
           number_format=CURRENCY_FMT)

# ── CHARTS ───────────────────────────────────────────────────────────────

# Chart 1: Service-Wise Orders (BarChart)
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Service-Wise Orders"
chart1.style = 10
chart1.y_axis.title = "Number of Orders"
chart1.x_axis.title = "Service"
chart1.width = 20
chart1.height = 12

data_ref1 = Reference(ws2, min_col=2, min_row=r_svc_hdr,
                      max_row=r_svc_start + len(SERVICES) - 1, max_col=2)
cats_ref1 = Reference(ws2, min_col=1, min_row=r_svc_start,
                      max_row=r_svc_start + len(SERVICES) - 1)
chart1.add_data(data_ref1, titles_from_data=True)
chart1.set_categories(cats_ref1)
chart1.shape = 4
ws2.add_chart(chart1, f"J{R_SVC}")

# Chart 2: Area-Wise Order Distribution (PieChart)
chart2 = PieChart()
chart2.title = "Area-Wise Order Distribution"
chart2.style = 10
chart2.width = 16
chart2.height = 12

data_ref2 = Reference(ws2, min_col=2, min_row=r_area_hdr,
                      max_row=r_area_start + len(AREAS) - 1, max_col=2)
cats_ref2 = Reference(ws2, min_col=1, min_row=r_area_start,
                      max_row=r_area_start + len(AREAS) - 1)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref2)
ws2.add_chart(chart2, f"J{R_AREA}")

# Chart 3: Order Source Distribution (BarChart)
chart3 = BarChart()
chart3.type = "col"
chart3.title = "Order Source Distribution"
chart3.style = 10
chart3.y_axis.title = "Number of Orders"
chart3.x_axis.title = "Source"
chart3.width = 18
chart3.height = 10

data_ref3 = Reference(ws2, min_col=2, min_row=r_src_hdr,
                      max_row=r_src_start + len(SOURCES) - 1, max_col=2)
cats_ref3 = Reference(ws2, min_col=1, min_row=r_src_start,
                      max_row=r_src_start + len(SOURCES) - 1)
chart3.add_data(data_ref3, titles_from_data=True)
chart3.set_categories(cats_ref3)
chart3.shape = 4
ws2.add_chart(chart3, f"K{R_SRC}")

# Freeze pane
ws2.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 3: FORM FIELDS REFERENCE
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Form Fields Reference")
ws3.sheet_properties.tabColor = MEDIUM_BLUE

# Title
ws3.merge_cells("A1:C1")
c = ws3["A1"]
c.value = "GOOGLE FORM — FIELD REFERENCE"
apply_cell(c, font=Font(name="Calibri", bold=True, size=16, color=WHITE),
           fill=make_fill(DARK_BLUE), alignment=center_al)
ws3.row_dimensions[1].height = 40

ws3.merge_cells("A2:C2")
c = ws3["A2"]
c.value = "Use this as a reference when creating your Google Form for lead capture"
apply_cell(c, font=Font(name="Calibri", bold=True, size=11, color=WHITE),
           fill=make_fill(MEDIUM_BLUE), alignment=center_al)
ws3.row_dimensions[2].height = 28

# Headers
form_headers = ["Field Name", "Field Type", "Options / Notes"]
for i, hdr in enumerate(form_headers):
    col_l = get_column_letter(i + 1)
    cell = ws3[f"{col_l}3"]
    cell.value = hdr
    apply_cell(cell, font=Font(name="Calibri", bold=True, size=10, color=WHITE),
               fill=make_fill(MEDIUM_BLUE), alignment=center_al, border=thin_border)

ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 55

form_fields = [
    ("Customer Name", "Short Answer", "Required. Full name of the customer."),
    ("Phone Number", "Short Answer", "Required. Primary contact number (10 digits)."),
    ("WhatsApp Number", "Short Answer", "Optional. WhatsApp contact if different from phone."),
    ("Alternate Phone", "Short Answer", "Optional. Secondary contact number."),
    ("Email", "Short Answer", "Optional. Customer email for invoices/updates."),
    ("City", "Dropdown", CITY_LIST),
    ("Area / Locality", "Short Answer", "e.g. Indiranagar, Koramangala, HSR Layout, Whitefield, JP Nagar"),
    ("Full Address", "Paragraph", "Complete address with landmark and pincode."),
    ("BHK", "Dropdown", BHK_LIST),
    ("SQFT (Approx)", "Short Answer", "Approximate area in square feet."),
    ("Service 1", "Dropdown", SERVICE_LIST),
    ("Service 2", "Dropdown", f"Optional. {SERVICE_LIST}"),
    ("Service 3", "Dropdown", f"Optional. {SERVICE_LIST}"),
    ("Service 4", "Dropdown", f"Optional. {SERVICE_LIST}"),
    ("Preferred Date", "Date", "Customer's preferred service date."),
    ("Time Slot", "Dropdown", SLOT_LIST),
    ("Order Source", "Dropdown", SOURCE_LIST),
    ("Special Instructions", "Paragraph", "Any special requirements or notes from the customer."),
]

for i, (field, ftype, notes) in enumerate(form_fields):
    r = 4 + i
    ws3[f"A{r}"].value = field
    apply_cell(ws3[f"A{r}"], font=Font(name="Calibri", bold=True, size=10),
               alignment=left_al, border=thin_border,
               fill=make_fill(LIGHT_GRAY if i % 2 == 0 else WHITE))
    ws3[f"B{r}"].value = ftype
    apply_cell(ws3[f"B{r}"], font=Font(name="Calibri", size=10),
               alignment=center_al, border=thin_border,
               fill=make_fill(LIGHT_GRAY if i % 2 == 0 else WHITE))
    ws3[f"C{r}"].value = notes
    apply_cell(ws3[f"C{r}"], font=Font(name="Calibri", size=10),
               alignment=left_al, border=thin_border,
               fill=make_fill(LIGHT_GRAY if i % 2 == 0 else WHITE))

# Instructions for linking Google Form
r_inst = 4 + len(form_fields) + 1
ws3.merge_cells(f"A{r_inst}:C{r_inst}")
ws3[f"A{r_inst}"].value = "HOW TO LINK GOOGLE FORM"
apply_cell(ws3[f"A{r_inst}"], font=Font(name="Calibri", bold=True, size=12, color=WHITE),
           fill=make_fill(DARK_BLUE), alignment=center_al)

link_instructions = [
    "1. Create a new Google Form at forms.google.com with the fields listed above.",
    "2. Go to Responses tab in the Form > Click the Google Sheets icon to link responses.",
    "3. A new response sheet is created automatically in Google Sheets.",
    "4. In the Lead Tracker sheet, use IMPORTRANGE or copy-paste data from the response sheet.",
    '5. Example: =IMPORTRANGE("spreadsheet_url", "Form Responses 1!A2:R100")',
    "6. Map the form fields to the correct Lead Tracker columns (D, E, F, G, H, I, J, K, L, M, N, P, R, T, Y, Z, AA).",
    "7. The Order ID (C), S.No (A), Total Value (V), Discounted Total (X), Pending Balance (AL), and Invoice (AQ) auto-populate.",
    "8. Manually update: Order Status (AB), Vendor details (AD-AF), Payment info (AJ-AO), etc.",
]

for li, text in enumerate(link_instructions):
    r = r_inst + 1 + li
    ws3.merge_cells(f"A{r}:C{r}")
    ws3[f"A{r}"].value = text
    apply_cell(ws3[f"A{r}"], font=Font(name="Calibri", size=10, color="333333"),
               alignment=left_al, border=thin_border)
    ws3.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════════
#  SAVE
# ═══════════════════════════════════════════════════════════════════════════
output_dir = "/var/lib/freelancer/projects/40182876"
output_file = os.path.join(output_dir, "Home_Services_Lead_Tracker.xlsx")
os.makedirs(output_dir, exist_ok=True)
wb.save(output_file)
print(f"Successfully created: {output_file}")
print(f"Sheets: {wb.sheetnames}")
print(f"Lead Tracker columns: {num_cols} (A to {last_col_letter})")
print(f"Dashboard sections: Order Stats, Payment Summary, Discount Summary, "
      f"Source Tracking, Service Breakdown, Area Distribution, Date-wise Report + 3 Charts")
print("Done!")
